import os
import pandas as pd
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def update_long_forecast_log_with_formatting(production_df, actuals_df, excel_output_path):
    today = actuals_df['date'].max().normalize()
    today_date = date.today()
    records = []

    for h in [7, 14, 30, 45]:
        tag_col = f'Confidence: {h}-Day'
        target_date = today + pd.Timedelta(days=h)

        actuals_on_target = actuals_df[actuals_df['date'] == target_date][['indicator', 'seen']]
        actuals_map = dict(zip(actuals_on_target['indicator'], actuals_on_target['seen']))

        for _, row in production_df.iterrows():
            indicator = row['Indicator']
            forecast_tag = row[tag_col]
            was_seen = actuals_map.get(indicator, None)

            if target_date > today:
                outcome = "Pending"
            elif was_seen == 1:
                outcome = "Seen"
            elif was_seen == 0:
                outcome = "Not Seen"
            else:
                outcome = "Unknown"

            records.append({
                'Indicator': indicator,
                'Forecast Date': today,
                'Forecast Type': f'{h}-Day',
                'Confidence': forecast_tag,
                'Forecasted Check Date': target_date,
                'Outcome': outcome
            })

    df = pd.DataFrame(records)
    df['Forecast Date'] = pd.to_datetime(df['Forecast Date'])
    df['Forecasted Check Date'] = pd.to_datetime(df['Forecasted Check Date'])

    expected_columns = ['Indicator', 'Forecast Date', 'Forecast Type', 'Confidence', 'Forecasted Check Date', 'Outcome']
    df = df[expected_columns]

    # Append existing logs if provided (Excel version)
    if excel_output_path and os.path.exists(excel_output_path):
        existing = pd.read_excel(excel_output_path, parse_dates=['Forecast Date', 'Forecasted Check Date'])
        df = pd.concat([existing, df], ignore_index=True).drop_duplicates()
        df = df[expected_columns]

    if excel_output_path:
        df.to_excel(excel_output_path, index=False)

        wb = load_workbook(excel_output_path)
        ws = wb.active

        green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        bold_font = Font(bold=True)
        default_fill = PatternFill(fill_type=None)
        default_font = Font(bold=False)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            row_data = {ws.cell(row=1, column=col_idx + 1).value: cell for col_idx, cell in enumerate(row)}
            confidence = row_data['Confidence'].value
            ftype = row_data['Forecast Type'].value
            outcome = row_data['Outcome'].value
            check_date = row_data['Forecasted Check Date'].value

            if not isinstance(check_date, (datetime, pd.Timestamp)) or check_date.date() != today_date:
                for cell in row:
                    cell.fill = default_fill
                    cell.font = default_font
                continue

            is_7day = ftype == '7-Day'
            is_highly_likely = 'Highly likely' in (confidence or '')
            is_low_confidence = 'Low confidence' in (confidence or '')
            is_seen = outcome == 'Seen'
            is_not_seen = outcome == 'Not Seen'

            apply_green = (is_low_confidence and is_not_seen) or (is_7day and is_highly_likely and is_seen)
            apply_red = (is_low_confidence and is_seen) or (is_7day and is_highly_likely and is_not_seen)

            if apply_green:
                for cell in row:
                    cell.fill = green_fill
            elif apply_red:
                for cell in row:
                    cell.fill = red_fill
            else:
                for cell in row:
                    cell.fill = default_fill

            if is_seen:
                for cell in row:
                    cell.font = bold_font
            else:
                for cell in row:
                    cell.font = default_font

        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    val = str(cell.value) if cell.value else ''
                    max_len = max(max_len, len(val))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 2

        wb.save(excel_output_path)

    return df
