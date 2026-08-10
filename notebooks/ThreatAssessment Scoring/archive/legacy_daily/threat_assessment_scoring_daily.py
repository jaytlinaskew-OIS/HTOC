# pip install openpyxl

import sys
import os
import re
import ast
import glob
import warnings
import urllib3
import urllib.parse
from configparser import ConfigParser
from datetime import date, datetime, timedelta, UTC
from concurrent.futures import ThreadPoolExecutor, as_completed

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import load_workbook

warnings.simplefilter(action='ignore', category=pd.errors.SettingWithCopyWarning)

# ── Paths & constants ─────────────────────────────────────────────────────────
HTOC_SHARE_ROOT = r"\\10.1.4.22\data\HTOC"

# Look for TC SDK relative to this script (checks up to 3 parent levels),
# then fall back to the network share location.
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TC_SDK_PATH = next(
    (os.path.join(d, 'threatconnect') for d in [
        _SCRIPT_DIR,
        os.path.dirname(_SCRIPT_DIR),
        os.path.dirname(os.path.dirname(_SCRIPT_DIR)),
    ] if os.path.isdir(os.path.join(d, 'threatconnect'))),
    HTOC_SHARE_ROOT + r"\Data_Analytics\threatconnect"
)
PROJECT_ROOT    = r"\\cscso1fsappv01\home\jaskew\HTOC\scripts\Data Movement\ThrearConnect-api-pull"
CONFIG_PATH     = os.path.join(PROJECT_ROOT, "utils", "config.json")

OBSERVED_INDICATORS_CSV  = HTOC_SHARE_ROOT + r"\Data_Analytics\Data\Observed_Indicators\htoc_observed_indicators.csv"
TAGS_FILE_PATH           = HTOC_SHARE_ROOT + r"\Data_Analytics\Data\Observed_Tags\htoc_observed_indicator_tags.csv"
OPDIV_BASE_PATH          = HTOC_SHARE_ROOT + r"\Data_Analytics\Data\OpDiv_Observations\htoc_opdiv_obs_d{date}.csv"
OUTPUT_DIR               = HTOC_SHARE_ROOT + r"\Data_Analytics\Data\Threat Assessment Scores"
EXCEL_FILENAME           = "Threat_Assessment_Scores.xlsx"

INDICATOR_TYPE_NAMES = [
    "Address", "EmailAddress", "File", "Host", "URL", "ASN", "CIDR",
    "Email Subject", "Hashtag", "Mutex", "Registry Key", "User Agent",
]
OWNER_NAMES = [
    'HTOC Org', 'CISA Federal Feed', 'CMS_CTI', 'Crowdstrike Falcon Intelligence',
    'DHS CISCP', 'Intel 471 Intelligence', 'Mandiant Advantage Threat Intelligence',
    'VA_TIP Data', 'Google Threat Intelligence',
]
THREAT_CATEGORY_FILTER  = 'THREAT ACTOR'
RESULT_PAGE_SIZE        = 500
FIRSTSEEN_LOOKBACK_DAYS = 7

OPDIV_DATE_FORMAT  = "%Y%m%d"
OPDIV_LOOKBACK_DAYS = 365
OPDIV_USE_COLS     = ['indicator', 'observations', 'OpDiv', 'obs_date']
OPDIV_DTYPES       = {'indicator': 'str', 'observations': 'int32', 'OpDiv': 'category', 'obs_date': 'str'}
OPDIV_WORKERS      = 16

MASS_SCANNER_TIER1_OBS = 10_000
MASS_SCANNER_TIER2_OBS = 100_000
MASS_SCANNER_OPDIV_MIN = 5

# ── ThreatConnect init ────────────────────────────────────────────────────────

sys.path.insert(0, TC_SDK_PATH)
from ThreatConnect import ThreatConnect
from RequestObject import RequestObject
from Owners import Owners

if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)
from utils.config_loader import load_config

api_secret_key, api_access_id, api_base_url, api_default_org = load_config(CONFIG_PATH)
print(f"Loaded config — Base URL: {api_base_url} | Access ID: {api_access_id}")

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
tc = ThreatConnect(api_access_id, api_secret_key, api_default_org, api_base_url)
print("ThreatConnect initialized.")

# ── Step 1: Load observed indicators CSV, filter to today ─────────────────────

df_observed = pd.read_csv(OBSERVED_INDICATORS_CSV)
print(f"Loaded {len(df_observed):,} rows and {df_observed.shape[1]} columns from observed indicators")

df_observed["ioc_current_period_firstseen"] = pd.to_datetime(df_observed["ioc_current_period_firstseen"])
today = pd.Timestamp(date.today())
df_today = df_observed[df_observed["ioc_current_period_firstseen"].dt.normalize() == today]
print(f"Records first seen today ({today.date()}): {len(df_today):,}")

# ── Step 2: Query ThreatConnect for today's indicators ────────────────────────

today_indicators = df_today['indicator'].dropna().unique().tolist()
print(f"Distinct indicators from today's observed data: {len(today_indicators)}")

type_name_condition = ", ".join([f'"{t}"' for t in INDICATOR_TYPE_NAMES])
owner_condition     = ", ".join([f'"{o}"' for o in OWNER_NAMES])
summary_condition   = ", ".join([f'"{ind}"' for ind in today_indicators])

tql_raw = (
    f'ownerName IN ({owner_condition}) AND '
    f'typeName IN ({type_name_condition}) AND '
    f'summary IN ({summary_condition})'
)
tql_encoded  = urllib.parse.quote(tql_raw)
final_results = []
ro_multi = RequestObject()
ro_multi.set_http_method('GET')
result_start = 0

while True:
    try:
        ro_multi.set_request_uri(
            f'/v3/indicators?tql={tql_encoded}'
            f'&fields=tags,observations,associatedGroups,falsePositives,threatAssess'
            f'&resultStart={result_start}&resultLimit={RESULT_PAGE_SIZE}'
        )
        response   = tc.api_request(ro_multi)
        ct         = response.headers.get('content-type', '')
        if not ct.startswith('application/json'):
            raise RuntimeError(f"Non-JSON response ({ct}): {response.content[:200]}")
        results    = response.json()
        data_items = results.get('data', []) or []
        if not data_items:
            break
        final_results.append(results)
        result_start += RESULT_PAGE_SIZE
    except Exception as e:
        print(f"Failed to query indicators (start={result_start}): {e}")
        break

normalized_data = []
for result in final_results:
    data_items = result.get('data', [])
    for item in data_items:
        if isinstance(item, dict) and 'summary' in item:
            normalized_data.append(item)

if normalized_data:
    observed_src = pd.json_normalize(normalized_data)
    observed_src['indicator'] = observed_src['summary'].astype(str).str.split().str[0].str.strip()

    sources_per_indicator = (
        observed_src.groupby('indicator')['ownerName']
        .apply(lambda x: ', '.join(sorted(set(x))))
        .reset_index()
        .rename(columns={'ownerName': 'sources'})
    )
    observed_src = observed_src.merge(sources_per_indicator, on='indicator', how='left')

    observed_src['_htoc_first'] = (observed_src['ownerName'] == 'HTOC Org').astype(int)
    observed_src = observed_src.sort_values(['indicator', '_htoc_first'], ascending=[True, False])
    observed_src = observed_src.drop(columns=['_htoc_first'])
    observed_src = observed_src.groupby('indicator', as_index=False).first()
    observed_src = observed_src[observed_src['ownerName'] == 'HTOC Org'].copy()
else:
    print("No valid indicator data found.")
    observed_src = pd.DataFrame()

print(f"observed_src: {len(observed_src):,} rows x {len(observed_src.columns)} cols")

# ── Step 3: Merge threat actor tags ──────────────────────────────────────────

htoc_observed_tags  = pd.read_csv(TAGS_FILE_PATH)
threat_actor_records = htoc_observed_tags[htoc_observed_tags['threat_category'] == THREAT_CATEGORY_FILTER]

threat_actor_condensed = threat_actor_records.groupby('indicator').agg({
    'type':                 'first',
    'orig_tag':             lambda x: ', '.join(sorted(set(x.dropna()))),
    'tag':                  lambda x: ', '.join(sorted(set(x.dropna()))),
    'threat_category':      lambda x: ', '.join(sorted(set(x.dropna()))),
    'NATION STATE':         lambda x: ', '.join(sorted(set(x.dropna()))) if x.notna().any() else None,
    'SECURITY ORGANIZATION':lambda x: ', '.join(sorted(set(x.dropna()))) if x.notna().any() else None,
    'MALWARE CLASS':        lambda x: ', '.join(sorted(set(x.dropna()))) if x.notna().any() else None,
    'CVE_NBR':              lambda x: ', '.join(sorted(set(x.dropna()))) if x.notna().any() else None,
}).reset_index()

print(f"Loaded {len(threat_actor_records)} {THREAT_CATEGORY_FILTER} rows, condensed to {len(threat_actor_condensed)} unique indicators")

threat_actor_tags = (
    threat_actor_condensed[['indicator', 'tag', 'orig_tag', 'threat_category',
                             'NATION STATE', 'SECURITY ORGANIZATION', 'MALWARE CLASS', 'CVE_NBR']]
    .copy()
    .rename(columns={
        'tag':                  'threat_actor',
        'orig_tag':             'threat_actor_orig_tag',
        'threat_category':      'threat_actor_category',
        'NATION STATE':         'threat_nation_state',
        'SECURITY ORGANIZATION':'threat_security_org',
        'MALWARE CLASS':        'threat_malware_class',
        'CVE_NBR':              'threat_cve_nbr',
    })
)
observed_src = observed_src.merge(threat_actor_tags, on='indicator', how='left')
print(f"Threat actor matches: {observed_src['threat_actor'].notna().sum():,} indicators")

# ── Step 4: Add firstseen date + incidents/events ─────────────────────────────

df_obs_firstseen = df_observed.copy()
df_obs_firstseen['firstseen_dt'] = pd.to_datetime(df_obs_firstseen['firstseen_dt'], errors='coerce')
cutoff = today - pd.Timedelta(days=FIRSTSEEN_LOOKBACK_DAYS)
df_last14 = df_obs_firstseen[
    (df_obs_firstseen['firstseen_dt'] >= cutoff) &
    (df_obs_firstseen['firstseen_dt'] <= today)
][['indicator', 'firstseen_dt']].rename(columns={'firstseen_dt': 'firstseen_date'})
observed_src = observed_src.merge(df_last14, on='indicator', how='left')

INCIDENT_COLUMN_NAME   = "incidents/events"
GROUP_TYPES_OF_INTEREST = {"incident", "event"}
INCIDENT_ID_REGEX       = re.compile(r"\bINC\d+\b", re.IGNORECASE)

if observed_src.empty:
    observed_src[INCIDENT_COLUMN_NAME] = []
    incidents_df = pd.DataFrame()
elif 'associatedGroups.data' not in observed_src.columns:
    observed_src[INCIDENT_COLUMN_NAME] = "None"
    incidents_df = pd.DataFrame()
else:
    def _extract_groups_of_interest(val):
        if isinstance(val, list):
            return [i for i in val if isinstance(i, dict) and str(i.get('type', '')).lower() in GROUP_TYPES_OF_INTEREST]
        if isinstance(val, dict):
            return [val] if str(val.get('type', '')).lower() in GROUP_TYPES_OF_INTEREST else []
        return []

    def _format_group_label(item):
        if not isinstance(item, dict):
            return str(item)
        typ    = str(item.get('type', '')).title()
        number = item.get('id') or item.get('xid') or item.get('name')
        return f"{typ}:{number}" if number is not None else typ

    def _extract_incidents_from_description(text):
        if not isinstance(text, str):
            return []
        unique_ids = list(dict.fromkeys(m.upper() for m in INCIDENT_ID_REGEX.findall(text)))
        return [{"type": "incident", "id": inc_id} for inc_id in unique_ids]

    groups = observed_src['associatedGroups.data'].apply(_extract_groups_of_interest)
    desc_groups = (
        observed_src['description'].apply(_extract_incidents_from_description)
        if 'description' in observed_src.columns
        else pd.Series([[] for _ in range(len(observed_src))], index=observed_src.index)
    )
    combined_groups = groups.combine(desc_groups, lambda a, b: (a or []) + (b or []))
    observed_src[INCIDENT_COLUMN_NAME] = combined_groups.apply(
        lambda lst: ";".join(_format_group_label(it) for it in lst) if isinstance(lst, list) and len(lst) > 0 else "None"
    )
    incidents_df = observed_src[combined_groups.apply(lambda lst: isinstance(lst, list) and len(lst) > 0)].copy()
    print(f"Annotated incidents/events for {len(observed_src)} indicators; found {len(incidents_df)} with matches.")

# ── Step 5: Tag extraction (PB / botnet) ──────────────────────────────────────

tags_exploded = (
    observed_src[['indicator', 'tags.data']]
    .explode('tags.data')
    .dropna(subset=['tags.data'])
)
tags_exploded['tag_name'] = tags_exploded['tags.data'].apply(lambda x: x.get('name') if isinstance(x, dict) else None)

indicator_to_tags = (
    tags_exploded.groupby('indicator')['tag_name']
    .apply(lambda x: [t for t in x if t])
    .to_dict()
)
observed_src['tag_list'] = observed_src['indicator'].map(lambda ind: indicator_to_tags.get(ind, []))

PB_START_LOWER_TAGS = {"soar indicator pb", "scanning cdn pb", "known scanning pb", "web scanner", "active scanning"}

def extract_pb_lower_tags(tag_list):
    if not isinstance(tag_list, list):
        return []
    return [t for t in tag_list if isinstance(t, str) and t.strip().lower() in PB_START_LOWER_TAGS]

observed_src['pb_lower_tags']   = observed_src['indicator'].map(lambda ind: extract_pb_lower_tags(indicator_to_tags.get(ind, [])))
observed_src['pb_lower_flag']   = observed_src['pb_lower_tags'].apply(lambda x: isinstance(x, list) and len(x) > 0)
observed_src['pb_lower_reason'] = observed_src['pb_lower_tags'].apply(
    lambda tags: f"pb_tag:{tags[0].strip()}" if isinstance(tags, list) and len(tags) > 0 else None
)

BOTNET_TAGS_OF_INTEREST = [
    "Scanning", "DDoS", "Spam", "Phishing", "Cryptojacking",
    "Credential Stuffing", "Ransomware", "Data Theft",
    "Cross Site Scripting Attacks", "SQL Injections",
]
botnet_tags_lower = {t.lower() for t in BOTNET_TAGS_OF_INTEREST}

def extract_botnet_tags(tag_list):
    if not isinstance(tag_list, list):
        return []
    return [t for t in tag_list if isinstance(t, str) and t.strip().lower() in botnet_tags_lower]

observed_src['Botnet'] = observed_src['indicator'].map(lambda ind: extract_botnet_tags(indicator_to_tags.get(ind, [])))

# ── Step 6: Load OpDiv observation files ──────────────────────────────────────

def get_file_paths(base_path, days=365):
    dir_path  = os.path.normpath(os.path.dirname(base_path))
    pattern   = os.path.join(dir_path, 'htoc_opdiv_obs_d*.csv')
    available = {os.path.normpath(p) for p in glob.glob(pattern)}
    today_dt  = datetime.now(UTC).replace(tzinfo=None)
    existing  = [
        p for p in
        (os.path.normpath(base_path.format(date=(today_dt - timedelta(days=i)).strftime(OPDIV_DATE_FORMAT))) for i in range(days))
        if p in available
    ]
    print(f"Found {len(existing)} of {days} OpDiv files to load")
    return existing

def _read_one(path):
    try:
        return pd.read_csv(path, usecols=OPDIV_USE_COLS, dtype=OPDIV_DTYPES)
    except Exception as e:
        print(f"Error reading {os.path.basename(path)}: {e}")
        return None

def load_observed_data(file_paths):
    with ThreadPoolExecutor(max_workers=OPDIV_WORKERS) as executor:
        frames = list(executor.map(_read_one, file_paths))
    frames = [f for f in frames if f is not None]
    if frames:
        df = pd.concat(frames, ignore_index=True)
        print(f"Loaded {len(df):,} rows from {len(frames)} files")
    else:
        df = pd.DataFrame()
    return df

file_paths       = get_file_paths(OPDIV_BASE_PATH, days=OPDIV_LOOKBACK_DAYS)
observed_data_df = load_observed_data(file_paths)

# Mass scanner detection
if not observed_data_df.empty and 'obs_date' in observed_data_df.columns:
    cutoff_7d = (pd.Timestamp.utcnow() - pd.Timedelta(days=7)).strftime('%Y-%m-%d')
    obs_7d    = observed_data_df[observed_data_df['obs_date'] >= cutoff_7d].copy()
    mass_scanner_agg = (
        obs_7d.groupby('indicator')
        .agg(
            total_obs_7d    =('observations', 'sum'),
            unique_opdivs_7d=('OpDiv', 'nunique'),
            unique_days_7d  =('obs_date', 'nunique'),
        )
        .reset_index()
    )
    broad_spread = mass_scanner_agg['unique_opdivs_7d'] >= MASS_SCANNER_OPDIV_MIN
    mass_scanner_agg['mass_scanner_tier1'] = (
        (mass_scanner_agg['total_obs_7d'] >= MASS_SCANNER_TIER1_OBS) &
        (mass_scanner_agg['total_obs_7d'] <  MASS_SCANNER_TIER2_OBS) & broad_spread
    )
    mass_scanner_agg['mass_scanner_tier2'] = (
        (mass_scanner_agg['total_obs_7d'] >= MASS_SCANNER_TIER2_OBS) & broad_spread
    )
    n1 = mass_scanner_agg['mass_scanner_tier1'].sum()
    n2 = mass_scanner_agg['mass_scanner_tier2'].sum()
    print(f"Mass scanner — Tier 1: {n1} indicators | Tier 2: {n2} indicators")
else:
    mass_scanner_agg = pd.DataFrame(columns=[
        'indicator', 'total_obs_7d', 'unique_opdivs_7d', 'unique_days_7d',
        'mass_scanner_tier1', 'mass_scanner_tier2'
    ])
    print("Mass scanner detection skipped — observed_data_df is empty or missing obs_date.")

# ── Step 7: Partner extraction ────────────────────────────────────────────────

cutoff_naive = pd.Timestamp.utcnow().tz_convert(None)
KNOWN_PARTNERS = {'DHA', 'OS', 'FDA', 'CMS', 'VA', 'HRSA', 'NIH', 'IHS', 'HHS', 'CDC'}

def get_all_partner_indicators_from_obs(observed_data_df, cutoff_naive):
    if observed_data_df.empty or 'OpDiv' not in observed_data_df.columns:
        return pd.DataFrame()
    observed_data_df['obs_date'] = pd.to_datetime(observed_data_df['obs_date'], errors='coerce')
    recent_obs = observed_data_df[observed_data_df['obs_date'] >= cutoff_naive - timedelta(days=60)].copy()
    if recent_obs.empty:
        return pd.DataFrame()
    partner_counts = (
        recent_obs.groupby('indicator')['OpDiv']
        .agg(['nunique', lambda s: ', '.join(sorted(set(s.dropna())))]).reset_index()
        .rename(columns={'nunique': 'partner_count_obs', '<lambda_0>': 'partners_from_obs'})
    )
    return partner_counts[partner_counts['partner_count_obs'] >= 1].copy()

def extract_partners_from_tags(observed_src):
    df = observed_src.copy()
    tags_exploded = df[['indicator', 'tags.data']].explode('tags.data').dropna(subset=['tags.data'])
    tags_norm = pd.json_normalize(tags_exploded['tags.data'])
    tags_norm.columns = [f"tag_{c}" for c in tags_norm.columns]
    tags_norm['tag_name'] = tags_norm['tag_name'].str.replace('VA CSOC CTS Splunk', 'VA Splunk API', regex=False)
    tags_expanded = tags_exploded.reset_index(drop=True).join(tags_norm)
    tags_expanded['partner'] = tags_expanded['tag_name'].map(
        lambda n: n[:-len(' Splunk API')] if isinstance(n, str) and n.endswith(' Splunk API') else None
    )
    tag_fields = list(tags_norm.columns)
    tag_agg = (
        tags_expanded.groupby('indicator', as_index=False)
        .agg({**{f: list for f in tag_fields}, 'partner': lambda x: [p for p in dict.fromkeys(x) if p]})
        .rename(columns={'partner': 'partners_from_tags'})
    )
    if 'tag_name' in tag_agg.columns:
        def extract_standalone(tag_list):
            if not isinstance(tag_list, list):
                return []
            return [t.strip() for t in tag_list if isinstance(t, str) and t.strip() in KNOWN_PARTNERS]
        tag_agg['standalone_partners'] = tag_agg['tag_name'].apply(extract_standalone)
        def combine_tag_partners(row):
            sp = row.get('partners_from_tags', [])
            ss = row.get('standalone_partners', [])
            if isinstance(sp, str):
                sp = [p.strip() for p in sp.split(',') if p.strip()]
            if not isinstance(sp, list):
                sp = []
            return list(dict.fromkeys(sp + ss))
        tag_agg['partners_from_tags'] = tag_agg.apply(combine_tag_partners, axis=1)
        tag_agg = tag_agg.drop(columns=['standalone_partners'], errors='ignore')
    return tag_agg, tag_fields

def combine_partners_from_sources(base_agg, tag_agg, all_partner_indicators):
    agg_df = base_agg.merge(tag_agg, on='indicator', how='left')
    if not all_partner_indicators.empty:
        agg_df = agg_df.merge(
            all_partner_indicators[['indicator', 'partners_from_obs', 'partner_count_obs']],
            on='indicator', how='left'
        )
    else:
        agg_df['partners_from_obs'] = ''
        agg_df['partner_count_obs'] = 0

    def combine_all(row):
        obs = row.get('partners_from_obs', '')
        tag = row.get('partners_from_tags', [])
        combined = set()
        if pd.notna(obs) and obs:
            for p in str(obs).split(', '):
                if p.strip():
                    combined.add(p.strip())
        if isinstance(tag, list):
            for p in tag:
                if p and p.strip():
                    combined.add(p.strip())
        elif pd.notna(tag) and tag:
            for p in str(tag).split(','):
                if p.strip():
                    combined.add(p.strip())
        return ', '.join(sorted(combined)) if combined else ''

    agg_df['partners'] = agg_df.apply(combine_all, axis=1)
    agg_df['partner_count'] = agg_df['partners'].apply(
        lambda x: len([p for p in str(x).split(', ') if p.strip()]) if pd.notna(x) and x else 0
    )
    agg_df = agg_df.drop(columns=[c for c in ['partners_from_obs', 'partner_count_obs', 'partners_from_tags'] if c in agg_df.columns], errors='ignore')
    return agg_df

print("Starting partner extraction pipeline...")
all_partner_indicators = get_all_partner_indicators_from_obs(observed_data_df, cutoff_naive)
print(f"Found {len(all_partner_indicators)} indicators with partners from observation data")

tag_agg, tag_fields = extract_partners_from_tags(observed_src)
print(f"Processed tags for {len(tag_agg)} indicators")

df = observed_src.copy()
first_cols = [
    'id','dateAdded','ownerId','ownerName','webLink','type','lastModified','falsePositives',
    'rating','confidence','description','summary','observations',
    'lastObserved','privateFlag','active','activeLocked','ip',
    'legacyLink','source','address','url','threatAssessScore','calScore',
    'incidents/events','sources','threat_actor','firstseen_date',
    'tag_list','pb_lower_tags','pb_lower_flag','pb_lower_reason',
]
if 'Botnet' in observed_src.columns:
    df['Botnet'] = observed_src['Botnet']
    first_cols.append('Botnet')

def clean_list(lst):
    if not isinstance(lst, list):
        return []
    cleaned = []
    for v in lst:
        try:
            if pd.isna(v):
                continue
        except Exception:
            pass
        if isinstance(v, str) and v == "":
            continue
        cleaned.append(v)
    return cleaned

def list_to_csv(lst):
    return ", ".join(str(v) for v in lst) if lst else ""

base_agg = (
    df.drop(columns=[
        'createdBy.id','createdBy.userName','createdBy.firstName','createdBy.lastName',
        'createdBy.pseudonym','createdBy.owner','xid','eventType','documentDateAdded',
        'documentType','fileSize','fileName','downVoteCount','upVoteCount','type_group',
        'webLink_group','ownerName_group','ownerId_group','dateAdded_group','id_group',
        'platforms.count','tactics.count',
    ], errors='ignore')
    .groupby('indicator', as_index=False)[[c for c in first_cols if c in df.columns]]
    .first()
)

agg_df = combine_partners_from_sources(base_agg, tag_agg, all_partner_indicators)

for col in ['group_ids', 'group_names'] + tag_fields:
    if col in agg_df.columns:
        agg_df[col] = agg_df[col].apply(clean_list).apply(list_to_csv)

print(f"Processing complete — {len(agg_df)} indicators")

if not mass_scanner_agg.empty:
    agg_df = agg_df.merge(
        mass_scanner_agg[['indicator','total_obs_7d','unique_opdivs_7d','mass_scanner_tier1','mass_scanner_tier2']],
        on='indicator', how='left'
    )
    agg_df['mass_scanner_tier1'] = agg_df['mass_scanner_tier1'].fillna(False).astype(bool)
    agg_df['mass_scanner_tier2'] = agg_df['mass_scanner_tier2'].fillna(False).astype(bool)
    agg_df['total_obs_7d']       = agg_df['total_obs_7d'].fillna(0).astype(int)
    agg_df['unique_opdivs_7d']   = agg_df['unique_opdivs_7d'].fillna(0).astype(int)
else:
    agg_df['mass_scanner_tier1'] = False
    agg_df['mass_scanner_tier2'] = False
    agg_df['total_obs_7d']       = 0
    agg_df['unique_opdivs_7d']   = 0

print(f"agg_df: {len(agg_df):,} rows x {len(agg_df.columns)} cols")

# ── Step 8: Enrichment (VT / Shodan) ─────────────────────────────────────────

COL_PATH    = "data.enrichment.data"
key_col     = 'indicator' if 'indicator' in agg_df.columns else 'summary'
VT_TYPES    = {'Address','IPv4','IPv6','Host','Domain','URL','File','SHA1','SHA256','MD5'}
SHODAN_TYPES= {'Address','IPv4','IPv6'}

cols = [key_col, 'type'] + (['id'] if 'id' in agg_df.columns else [])
candidates = (
    agg_df[cols].dropna(subset=[key_col]).astype({key_col: str})
    .drop_duplicates(subset=[key_col])
)
candidates = candidates[candidates['type'].astype(str).str.strip().isin(VT_TYPES | SHODAN_TYPES)].copy()
indicator_values = candidates[key_col].tolist()
print(f"Enriching {len(indicator_values)} indicators (VT; Shodan for IP types only)...")

def _enrich_one(row_series):
    value  = row_series[key_col]
    typ    = str(row_series.get('type', '') or '')
    row_id = row_series.get('id')
    use_id = pd.notna(row_id) and str(row_id).strip().isdigit()
    try:
        iid = str(int(float(row_id))) if use_id else urllib.parse.quote(value, safe="")
        providers = []
        if typ in VT_TYPES:    providers.append("VirusTotalV3")
        if typ in SHODAN_TYPES: providers.append("Shodan")
        if not providers:      providers.append("VirusTotalV3")
        q   = urllib.parse.urlencode({"type": providers}, doseq=True)
        ro  = RequestObject()
        ro.set_http_method("POST")
        ro.set_request_uri(f"/v3/indicators/{iid}/enrich?{q}")
        ro.set_body({})
        resp = tc.api_request(ro)
        try:
            data = resp.json()
        except Exception:
            data = {"status": getattr(resp, 'status_code', 'n/a'), "raw": getattr(resp, 'text', None)}
        data[key_col] = value
        return (data, None)
    except Exception as e:
        return (None, {key_col: value, "type": typ, "error": str(e)})

enriched_results = []
failed = []
with ThreadPoolExecutor(max_workers=8) as executor:
    futures = {executor.submit(_enrich_one, row): row for _, row in candidates.iterrows()}
    for future in as_completed(futures):
        result, err = future.result()
        if result is not None:
            enriched_results.append(result)
        else:
            failed.append(err)

if not enriched_results:
    print("No enrichment data retrieved.")
    recent_tags = agg_df.copy()
else:
    df_enriched = pd.json_normalize(enriched_results).drop_duplicates(subset=[key_col], keep="last")
    recent_tags = agg_df.merge(df_enriched, on=key_col, how="left")

    if COL_PATH in recent_tags.columns:
        exploded    = recent_tags[[key_col, COL_PATH]].explode(COL_PATH).dropna(subset=[COL_PATH])
        enrich_flat = pd.json_normalize(exploded[COL_PATH]).add_prefix("enrich_")
        enrich_flat[key_col] = exploded[key_col].values

        def _agg_obj(series):
            vals = [v for v in series.dropna()]
            if not vals:
                return None
            flat = []
            for v in vals:
                if isinstance(v, list):
                    flat.extend(v)
                else:
                    flat.append(v)
            if all(not isinstance(v, (list, dict)) for v in flat):
                return list(pd.Series(flat).unique())
            return flat

        obj_cols = enrich_flat.select_dtypes("object").columns.difference([key_col])
        num_cols = enrich_flat.columns.difference(obj_cols.union({key_col}))
        agg_dict = {c: _agg_obj for c in obj_cols}
        agg_dict.update({c: "max" for c in num_cols})
        enrich_wide = enrich_flat.groupby(key_col, as_index=False).agg(agg_dict)
        recent_tags = (
            recent_tags.drop(columns=[COL_PATH], errors="ignore")
            .drop_duplicates(subset=[key_col])
            .merge(enrich_wide, on=key_col, how="left")
        )

    print(f"Enrichment complete for {recent_tags[key_col].notna().sum()} indicators.")

if failed:
    fail_df = pd.DataFrame(failed)
    print(f"{len(failed)} indicators failed enrichment (showing up to 10):")
    print(fail_df.head(10).to_string())

recent_tags.drop(columns=[
    'tag_id','tag_lastUsed','tag_lastModified','tag_ownerId','tag_ownerName',
    'tag_dateAdded','tag_description','tag_tactics.count','tag_platform.data',
    'tag_platform.count','data.id','data.dateAdded','data.ownerId','data.webLink',
    'data.ownerName','data.lastModified','data.summary','data.ip','data.legacyLink',
    'data.source','enrich_cloudProvider','enrich_cloudRegion','enrich_type','id',
], inplace=True, errors='ignore')

# Enrichment domain counts
if 'enrich_domains' in exploded.columns:
    flat_df = (
        exploded[['indicator','enrich_domains']].explode('enrich_domains')
        .dropna(subset=['enrich_domains']).rename(columns={'enrich_domains':'domain'})
    )
    domain_counts = (
        flat_df.groupby('domain')['indicator'].nunique().reset_index()
        .rename(columns={'indicator':'indicator_count'}).sort_values('indicator_count', ascending=False)
    )
    print(domain_counts.to_string())

# Yearly observation count per indicator
agg_by_indicator = (
    observed_data_df.groupby('indicator', as_index=False)['obs_date']
    .nunique().rename(columns={'obs_date':'obs_days_count'})
)
agg_by_indicator = agg_by_indicator[agg_by_indicator['indicator'].isin(recent_tags['indicator'])]
recent_tags = recent_tags.merge(
    agg_by_indicator.rename(columns={'obs_days_count':'obs_count'}),
    on='indicator', how='left'
)
print(f"obs_count merged: {len(agg_by_indicator):,} indicators")
print(f"recent_tags: {len(recent_tags):,} rows x {len(recent_tags.columns)} cols")

# ── Step 9: PRISM Scoring ─────────────────────────────────────────────────────

df_scored = recent_tags.copy()

VT_COL          = 'enrich_vtMaliciousCount'
VT_EFFECTIVE_MAX = 40

if VT_COL in df_scored.columns:
    df_scored[VT_COL]       = pd.to_numeric(df_scored[VT_COL], errors='coerce')
    df_scored['vt_present'] = df_scored[VT_COL].notna()
else:
    df_scored[VT_COL]       = np.nan
    df_scored['vt_present'] = False

df_scored['vt_present']              = df_scored['vt_present'].astype(bool)
df_scored['vt_display']              = np.where(df_scored['vt_present'], df_scored[VT_COL], 'No VT Score')
df_scored['vt_numeric_for_scoring']  = df_scored[VT_COL].fillna(0).clip(0, VT_EFFECTIVE_MAX)

scoring_bins = [0, 200, 500, 800, 1001]
label_bins   = ['low', 'medium', 'high', 'critical']

Weights = {
    "MALICIOUS_WEIGHT":         7.50,
    "OBSERVATION_COUNT_WEIGHT": 0.02,
    "CONTINUITY_WEIGHT":        0.90,
    "TC_RATING":                0.01,
    "TC_CONFIDENCE":            0.025,
    "TOR_ACTIVITY_WEIGHT":      9.00,
    "CAL_SCORE_WEIGHT":         2.75,
    "TC_THREAT_SCORE_WEIGHT":   0.75,
    "INCIDENTS_EVENTS_WEIGHT":  8.00,
    "PARTNER_WEIGHT":           2.10,
    "SOURCES_WEIGHT":           2.80,
    "THREAT_ACTOR_WEIGHT":      10.00,
    "FIRST_OBS_WEIGHT":         2.00,
}

BOTNET_ACTIONS           = {'scanning','ddos','spam','phishing','cryptojacking','credential stuffing','ransomware'}
TOR_ACTIVITY             = {'tor','tor activity'}
MAX_OBS_REALISTIC        = 365
MAX_RATING               = 5
MAX_CONFIDENCE           = 100
FALSE_POSITIVE_WEIGHT    = 0.9
BOTNET_MULTIPLIER        = 0.4
SCANNER_PENALTY_MULTIPLIER = 0.80
DATA_QUALITY_FLOOR       = 0.85
PB_BASE_SCORE_MULTIPLIER = 0.45
THREAT_TAG_MIN_FLOOR     = 560

def convert_to_list(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return []
    if isinstance(val, (list, set, tuple)):
        return list(val)
    if isinstance(val, str):
        s = val.strip()
        if s.startswith('[') and s.endswith(']'):
            try:
                parsed = ast.literal_eval(s)
                if isinstance(parsed, (list, tuple)):
                    return list(parsed)
            except Exception:
                pass
        return [x.strip() for x in val.split(',') if x.strip()]
    return [val] if val else []

def extract_pb_lower_tags_from_val(val):
    return [str(x).strip() for x in convert_to_list(val) if str(x).strip().lower() in PB_START_LOWER_TAGS]

def has_pb_lower_tag(val):
    return len(extract_pb_lower_tags_from_val(val)) > 0

def normalize_token(val):
    s = str(val).strip().lower().replace('_', ' ')
    s = re.sub(r'[()]', lambda m: f" {m.group(0)} ", s)
    s = re.sub(r'\s*/\s*', ' / ', s)
    return " ".join(s.split())

def flatten_tag_tokens(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return []
    values = val if isinstance(val, (list, set, tuple)) else [val]
    out = []
    for item in values:
        if item is None or (isinstance(item, float) and pd.isna(item)):
            continue
        if isinstance(item, dict):
            tag_name = item.get('name')
            if tag_name is not None:
                out.append(normalize_token(tag_name))
            continue
        if isinstance(item, str):
            s = item.strip()
            if not s:
                continue
            parsed = convert_to_list(s)
            if isinstance(parsed, list) and len(parsed) > 1:
                out.extend(normalize_token(x) for x in parsed if str(x).strip())
            else:
                out.extend(normalize_token(x) for x in s.split(',') if str(x).strip())
            continue
        out.append(normalize_token(item))
    return [t for t in out if t not in {'', 'none', 'nan'}]

def token_matches_pattern(token, pattern):
    p = normalize_token(pattern)
    return token.endswith(p[1:]) if p.startswith('*') else token == p

def has_threat_actor(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    if isinstance(val, str):
        s = val.strip()
        return not (s == '' or s.lower() in {'none', 'nan'})
    if isinstance(val, (list, set, tuple)):
        return len(val) > 0
    return False

# Input caps
for col, cap in [('obs_count', MAX_OBS_REALISTIC), ('rating', MAX_RATING), ('confidence', MAX_CONFIDENCE), ('calScore', 1000)]:
    df_scored[col] = pd.to_numeric(
        df_scored[col] if col in df_scored.columns else pd.Series(0, index=df_scored.index),
        errors='coerce'
    ).fillna(0).clip(0, cap)

df_scored['type'] = (
    df_scored['type'] if 'type' in df_scored.columns else pd.Series('', index=df_scored.index)
).astype(str)

# PB lower flag
if 'pb_lower_flag' in df_scored.columns:
    df_scored['pb_lower_flag'] = df_scored['pb_lower_flag'].fillna(False).astype(bool)
elif 'tag_list' in df_scored.columns:
    df_scored['pb_lower_flag'] = df_scored['tag_list'].apply(has_pb_lower_tag).astype(bool)
else:
    df_scored['pb_lower_flag'] = False

if 'pb_lower_tags' not in df_scored.columns:
    df_scored['pb_lower_tags'] = (
        df_scored['tag_list'].apply(extract_pb_lower_tags_from_val) if 'tag_list' in df_scored.columns
        else [[] for _ in range(len(df_scored))]
    )

if 'pb_lower_reason' not in df_scored.columns:
    df_scored['pb_lower_reason'] = df_scored['pb_lower_tags'].apply(
        lambda tags: f"pb_tag:{str(tags[0]).strip()}" if isinstance(tags, list) and len(tags) > 0 else None
    )

# First-seen recency boost
FIRST_OBS_MAX_DAYS = 14
firstseen_dt = pd.to_datetime(df_scored.get('firstseen_date', pd.Series(pd.NaT, index=df_scored.index)), errors='coerce')
today_ts     = pd.Timestamp.today().normalize()
age_days     = (today_ts - firstseen_dt).dt.days.clip(lower=0)
freshness    = ((FIRST_OBS_MAX_DAYS - age_days) / FIRST_OBS_MAX_DAYS).clip(lower=0.0, upper=1.0)
freshness    = freshness.where(firstseen_dt.notna(), 0.0)
df_scored['first_obs_raw_score'] = freshness * Weights['FIRST_OBS_WEIGHT']

# Base additive evidence
MALICIOUS_EXPONENT = 0.75
df_scored['w_malicious_eff']   = Weights['MALICIOUS_WEIGHT']
df_scored['w_tc_rating_eff']   = Weights['TC_RATING']
df_scored['malicious_scaled']  = np.power(df_scored['vt_numeric_for_scoring'], MALICIOUS_EXPONENT)
df_scored['malicious_raw_score'] = df_scored['malicious_scaled'] * Weights['MALICIOUS_WEIGHT']

FILE_TYPES = {'SHA1','SHA256','MD5','file','File'}
df_scored['is_file_type'] = df_scored['type'].isin(FILE_TYPES)
df_scored['continuity_val'] = df_scored['type'].map({
    'Address':1,'IPv4':1,'IPv6':1,
    'Domain':2,'Host':2,'URL':2,'Stripped URL':2,'EmailAddress':2,'EmailSubject':2,
    'SHA1':3,'SHA256':3,'MD5':3,'file':3,'File':3,
}).fillna(0)
df_scored['continuity_raw_score'] = df_scored['continuity_val'] * Weights['CONTINUITY_WEIGHT']
df_scored.loc[df_scored['is_file_type'], 'continuity_raw_score'] = 900
df_scored['tc_raw_rating']     = df_scored['rating'] * df_scored['w_tc_rating_eff']
df_scored['tc_raw_confidence'] = np.sqrt(df_scored['confidence']) * Weights['TC_CONFIDENCE']
df_scored['cal_raw_score']     = (df_scored['calScore'] / 1000.0) * Weights['CAL_SCORE_WEIGHT']

TC_THREAT_COL = 'threatAssessScore'
if TC_THREAT_COL in df_scored.columns:
    df_scored[TC_THREAT_COL]         = pd.to_numeric(df_scored.get(TC_THREAT_COL, 0), errors='coerce').fillna(0).clip(0, 1000)
    df_scored['tc_threat_raw_score'] = (df_scored[TC_THREAT_COL] / 1000.0) * Weights['TC_THREAT_SCORE_WEIGHT']
else:
    df_scored['tc_threat_raw_score'] = 0.0

# Incidents/events bonus
INCIDENTS_EVENTS_COL = 'incidents/events'
def has_incident_event(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    if isinstance(val, (list, set, tuple)):
        return len(val) > 0
    s = str(val).strip()
    return not (s == '' or s.lower() in {'none', 'nan'})

df_scored['incidents_events_flag'] = (
    df_scored[INCIDENTS_EVENTS_COL].apply(has_incident_event).astype(int)
    if INCIDENTS_EVENTS_COL in df_scored.columns else 0
)
df_scored['incidents_events_score'] = np.where(
    df_scored['pb_lower_flag'], 0.0, df_scored['incidents_events_flag'] * Weights['INCIDENTS_EVENTS_WEIGHT']
)

# Sources / partners
def count_sources(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return 0
    if isinstance(v, str):
        return len(set(s.strip() for s in v.split(',') if s.strip()))
    if isinstance(v, (list, set, tuple)):
        return len(set(str(s).strip() for s in v if str(s).strip()))
    return 0

def count_partners(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return 0
    if isinstance(v, (list, set, tuple)):
        return len(set(str(x).strip() for x in v if str(x).strip()))
    s = str(v).strip()
    if s == '' or s.lower() in {'none', 'nan'}:
        return 0
    return len(set(str(x).strip() for x in convert_to_list(s) if str(x).strip()))

df_scored['sources_count']      = df_scored['sources'].apply(count_sources) if 'sources' in df_scored.columns else 1
df_scored['sources_count_safe'] = df_scored['sources_count'].clip(lower=1)
df_scored['sources_raw_score']  = np.log1p(df_scored['sources_count_safe'] - 1) * Weights['SOURCES_WEIGHT']
df_scored['partners_count']     = df_scored['partners'].apply(count_partners) if 'partners' in df_scored.columns else 0
df_scored['partners_count_safe']= df_scored['partners_count'].clip(lower=1)
df_scored['partner_raw_score']  = np.log1p(df_scored['partners_count_safe'] - 1) * Weights['PARTNER_WEIGHT']

# Tagging boost
PAIRED_COUNTRY_TAGS = {
    'iran':        ['MOIS','IRGC','IRGC CEC','IRGC EWCD','IRGC QF','*Sandstorm','*Kitten'],
    'russia':      ['SVR','FSB','GRU','*Blizzard','*Bear'],
    'china':       ['MSS','ShSSB','TSSB','GSSB','JSSB','SSSB','HuSSB','HaSSB','SiSSB','PLA','PLA CSF','PLA SSF','PLAN','*Panda','*Typhoon'],
    'north korea': ['RGB','*Chollima','*Sleet'],
    'palestine':   ['Hamas'],
    'lebanon':     ['Lebanese Hizballah'],
}
STANDALONE_BOOST_TAGS = {
    'vietnam','belarus','palestine','pakistan','india',
    'teampcp','compromised trivy','operational relay box',
    'command and control','command and control (c2)','c2','data exfiltration',
    'wiper','destructive wiper','data wiper',
    'dropper','loader','loader/dropper','loader / dropper',
    'backdoor','rat','backdoor/rat','backdoor / rat','ransomware',
    'remote code execution','remote code execution (rce)','rce',
    'cisco','fortigate','fortinet','sap netweaver',
    'apt & targeted attack','spacehop orb','superjumper',
}
CVE_PATTERN = re.compile(r'^cve-\d{1,4}-\d{1,7}$', re.IGNORECASE)

def evaluate_tagging_boost_reason(row):
    token_fields = ['adversary','threat_actor','threat_actor_orig_tag','threat_actor_category',
                    'threat_nation_state','threat_security_org','threat_malware_class','threat_cve_nbr',
                    'tag_name','enrich_tags','tags.data']
    tokens    = []
    for col in token_fields:
        if col in row.index:
            tokens.extend(flatten_tag_tokens(row.get(col)))
    token_set = set(tokens)
    for country, pair_tags in PAIRED_COUNTRY_TAGS.items():
        pair_hits = sorted({tok for tok in token_set for patt in pair_tags if token_matches_pattern(tok, patt)})
        if country in token_set and pair_hits:
            return f"pair:{country}+{pair_hits[0]}"
    standalone_hits = sorted(tok for tok in token_set if tok in STANDALONE_BOOST_TAGS)
    if standalone_hits:
        return f"standalone:{standalone_hits[0]}"
    cve_hits = sorted(tok for tok in token_set if CVE_PATTERN.match(tok))
    if cve_hits:
        return f"cve:{cve_hits[0]}"
    return None

threat_actor_flag = pd.Series(False, index=df_scored.index)
if 'adversary' in df_scored.columns:
    threat_actor_flag = df_scored['adversary'].apply(has_threat_actor)
elif 'threat_actor' in df_scored.columns:
    threat_actor_flag = df_scored['threat_actor'].apply(has_threat_actor)

df_scored['threat_actor_score']    = threat_actor_flag.astype(int) * Weights['THREAT_ACTOR_WEIGHT']
df_scored['Tagging_Boost_Reason']  = df_scored.apply(evaluate_tagging_boost_reason, axis=1)
df_scored['Tagging_Boost']         = df_scored['Tagging_Boost_Reason'].notna().astype(bool)

# TOR activity
def has_tor_activity(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    if isinstance(val, (list, set, tuple)):
        t = " ".join(map(str, val)).lower()
    elif isinstance(val, str):
        if not val.strip():
            return False
        t = " ".join(x.strip() for x in val.split(',')).lower()
    else:
        t = str(val).lower()
        if t in ['nan','none','']:
            return False
    return any(k in t for k in TOR_ACTIVITY)

tor_mask_enrich = df_scored['enrich_tags'].apply(has_tor_activity) if 'enrich_tags' in df_scored.columns else pd.Series(False, index=df_scored.index)
tor_mask_tag    = df_scored['tag_name'].apply(convert_to_list).apply(has_tor_activity) if 'tag_name' in df_scored.columns else pd.Series(False, index=df_scored.index)
tor_flag        = (tor_mask_enrich | tor_mask_tag).astype(int)
df_scored['tor_activity_score'] = tor_flag * Weights['TOR_ACTIVITY_WEIGHT']
boost_mask      = df_scored['vt_present'] & (pd.to_numeric(df_scored['vt_numeric_for_scoring'], errors='coerce').fillna(0) >= 10) & tor_flag.astype(bool)
df_scored.loc[boost_mask, 'tor_activity_score'] *= 2

# Stacked context bonus
df_scored['stacked_context_count'] = (
    (df_scored['threat_actor_score'] > 0).astype(int) +
    (df_scored['tor_activity_score'] > 0).astype(int) +
    (df_scored['incidents_events_score'] > 0).astype(int) +
    (df_scored['sources_count'] >= 2).astype(int) +
    (df_scored['partners_count'] >= 2).astype(int)
)
df_scored['stacked_context_bonus'] = np.select(
    [df_scored['stacked_context_count'] >= 4, df_scored['stacked_context_count'] == 3, df_scored['stacked_context_count'] == 2],
    [25.0, 15.0, 7.0], default=0.0
)

# Raw score
threat_boost_mask = df_scored['Tagging_Boost'].astype(bool)
df_scored.loc[threat_boost_mask, ['malicious_raw_score','tc_raw_rating','tc_raw_confidence','tc_threat_raw_score']] = 0.0
df_scored['raw_score'] = (
    df_scored['malicious_raw_score'] + df_scored['continuity_raw_score'] +
    df_scored['tc_raw_rating'] + df_scored['tc_raw_confidence'] +
    df_scored['tor_activity_score'] + df_scored['incidents_events_score'] +
    df_scored['sources_raw_score'] + df_scored['partner_raw_score'] +
    df_scored['threat_actor_score'] + df_scored['cal_raw_score'] +
    df_scored['tc_threat_raw_score'] + df_scored['first_obs_raw_score'] +
    df_scored['stacked_context_bonus']
)
df_scored['pb_base_multiplier'] = np.where(df_scored['pb_lower_flag'], PB_BASE_SCORE_MULTIPLIER, 1.0)
df_scored['raw_score'] *= df_scored['pb_base_multiplier']

# Penalties
OBS_MIN_MULTIPLIER = 0.50
obs_frac = df_scored['obs_count'] / MAX_OBS_REALISTIC
df_scored['obs_penalty_multiplier'] = (1.0 - Weights['OBSERVATION_COUNT_WEIGHT'] * obs_frac).clip(OBS_MIN_MULTIPLIER, 1.0)
df_scored['raw_score'] *= df_scored['obs_penalty_multiplier']

present_frac = df_scored[['type','rating','confidence']].notna().sum(axis=1) / 3
df_scored['data_quality_multiplier'] = present_frac.clip(DATA_QUALITY_FLOOR, 1.0)
df_scored['raw_score'] *= df_scored['data_quality_multiplier']

def is_botnet(val):
    text = " ".join(map(str, val)).lower() if isinstance(val, (list, set, tuple)) else str(val).lower()
    return int(any(action in text for action in BOTNET_ACTIONS))

df_scored['botnet_flag'] = (
    pd.Series(df_scored['Botnet']).apply(is_botnet).astype(int)
    if 'Botnet' in df_scored.columns else 0
)
botnet_penalty_mask = (df_scored['botnet_flag'] == 1) & (~df_scored['Tagging_Boost']) & (~df_scored['is_file_type'])
df_scored['botnet_penalty_multiplier'] = 1.0
df_scored.loc[botnet_penalty_mask, 'botnet_penalty_multiplier'] = BOTNET_MULTIPLIER
df_scored['raw_score'] *= df_scored['botnet_penalty_multiplier']

if 'falsePositives' in df_scored.columns:
    df_scored['falsePositives'] = pd.to_numeric(df_scored['falsePositives'], errors='coerce').fillna(0)
    mask_fp = df_scored['falsePositives'] > 0
    df_scored['false_positive_raw_score'] = df_scored['raw_score'] * FALSE_POSITIVE_WEIGHT
    df_scored.loc[mask_fp, 'raw_score'] = df_scored.loc[mask_fp, 'false_positive_raw_score']
else:
    df_scored['falsePositives'] = 0
    df_scored['false_positive_raw_score'] = df_scored['raw_score']

def has_scanner_tag(val):
    scanners = {'scanner','masscan','zmap','shodan','censys','active scanning: scanning ip blocks','web scanner','active scanning'}
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    t = " ".join(map(str, val)).lower() if isinstance(val, (list, set, tuple)) else str(val).lower() if val else ''
    if t in ['nan','none','']:
        return False
    return any(s in t for s in scanners)

scanner_mask_enrich = df_scored['enrich_tags'].apply(has_scanner_tag) if 'enrich_tags' in df_scored.columns else pd.Series(False, index=df_scored.index)
scanner_mask_tag    = df_scored['tag_name'].apply(convert_to_list).apply(has_scanner_tag) if 'tag_name' in df_scored.columns else pd.Series(False, index=df_scored.index)
scanner_mask = (scanner_mask_enrich | scanner_mask_tag) & ~df_scored['is_file_type']
df_scored['scanner_penalty_multiplier'] = np.where(scanner_mask, SCANNER_PENALTY_MULTIPLIER, 1.0)
df_scored['raw_score'] *= df_scored['scanner_penalty_multiplier']

MASS_SCANNER_TIER1_MULTIPLIER = 0.40
MASS_SCANNER_TIER2_MULTIPLIER = 0.05
tier1_col = df_scored.get('mass_scanner_tier1', None)
tier2_col = df_scored.get('mass_scanner_tier2', None)
mass_scanner_tier1_mask = tier1_col.astype(bool) & ~df_scored['is_file_type'] if tier1_col is not None else pd.Series(False, index=df_scored.index)
mass_scanner_tier2_mask = tier2_col.astype(bool) & ~df_scored['is_file_type'] if tier2_col is not None else pd.Series(False, index=df_scored.index)
df_scored['mass_scanner_penalty_multiplier'] = np.select(
    [mass_scanner_tier2_mask, mass_scanner_tier1_mask],
    [MASS_SCANNER_TIER2_MULTIPLIER, MASS_SCANNER_TIER1_MULTIPLIER], default=1.0
)
df_scored['raw_score'] *= df_scored['mass_scanner_penalty_multiplier']

# Normalize to 0-1000
MAX_SOURCES_REALISTIC = 8
MAX_PARTNERS_REALISTIC = 10
BASE_CAP = (
    np.power(VT_EFFECTIVE_MAX, MALICIOUS_EXPONENT) * Weights['MALICIOUS_WEIGHT'] +
    3 * Weights['CONTINUITY_WEIGHT'] +
    (MAX_RATING * Weights['TC_RATING']) +
    (np.sqrt(MAX_CONFIDENCE) * Weights['TC_CONFIDENCE']) +
    (Weights['TOR_ACTIVITY_WEIGHT'] * 2) +
    Weights['INCIDENTS_EVENTS_WEIGHT'] +
    (np.log1p(MAX_SOURCES_REALISTIC - 1) * Weights['SOURCES_WEIGHT']) +
    (np.log1p(MAX_PARTNERS_REALISTIC - 1) * Weights['PARTNER_WEIGHT']) +
    Weights['THREAT_ACTOR_WEIGHT'] +
    Weights['CAL_SCORE_WEIGHT'] +
    Weights['TC_THREAT_SCORE_WEIGHT'] +
    Weights['FIRST_OBS_WEIGHT'] +
    25.0  # MAX_STACKED_CONTEXT_BONUS
)
FILE_BASELINE_RAW = 900.0
df_scored['raw_score_cap_row'] = np.where(df_scored['is_file_type'], BASE_CAP + FILE_BASELINE_RAW, BASE_CAP)
df_scored['PRISM_Score'] = (
    np.minimum(1000 * (df_scored['raw_score'] / df_scored['raw_score_cap_row']).clip(0, 1) * 1.40, 1000)
    .round().fillna(0).astype(int)
)

# VT ceilings / floors
vt_present_mask     = df_scored['vt_present']
vt_counts_present   = df_scored['vt_numeric_for_scoring']
low_cap_mask        = vt_present_mask & (vt_counts_present <= 3)
high_floor_mask     = vt_present_mask & (vt_counts_present >= 13)
tor_present_mask    = df_scored['tor_activity_score'] > 0
threat_actor_present_mask = df_scored['threat_actor_score'] > 0
low_cap_final_mask  = low_cap_mask & ~(tor_present_mask | threat_actor_present_mask | threat_boost_mask)
df_scored.loc[low_cap_final_mask, 'PRISM_Score'] = df_scored.loc[low_cap_final_mask, 'PRISM_Score'].clip(upper=499)
high_floor_final_mask = high_floor_mask & ~df_scored['pb_lower_flag']
df_scored['vt_high_floor_bypassed'] = high_floor_mask & df_scored['pb_lower_flag']
df_scored.loc[high_floor_final_mask, 'PRISM_Score'] = df_scored.loc[high_floor_final_mask, 'PRISM_Score'].clip(lower=500)

df_scored['Severity'] = pd.cut(df_scored['PRISM_Score'], bins=scoring_bins, labels=label_bins, right=False)
file_hash_mask = df_scored['is_file_type']
df_scored.loc[file_hash_mask, 'PRISM_Score'] = df_scored.loc[file_hash_mask, 'PRISM_Score'].clip(lower=scoring_bins[3])
df_scored.loc[file_hash_mask, 'Severity']   = 'critical'

# Tag floor band
min_tag_floor_score  = THREAT_TAG_MIN_FLOOR
tag_floor_band_width = 60
below_floor_mask     = threat_boost_mask & (df_scored['PRISM_Score'] < min_tag_floor_score)
if below_floor_mask.any():
    vals    = df_scored.loc[below_floor_mask, 'PRISM_Score'].astype(float)
    min_val = vals.min()
    max_val = vals.max()
    if max_val > min_val:
        norm = (vals - min_val) / (max_val - min_val)
        df_scored.loc[below_floor_mask, 'PRISM_Score'] = min_tag_floor_score + norm * tag_floor_band_width
    else:
        df_scored.loc[below_floor_mask, 'PRISM_Score'] = min_tag_floor_score + (tag_floor_band_width / 2.0)

df_scored['PRISM_Score']    = df_scored['PRISM_Score'].clip(upper=1000).round().astype(int)
df_scored['Severity']       = pd.cut(df_scored['PRISM_Score'], bins=scoring_bins, labels=label_bins, right=False)
df_scored.loc[df_scored['is_file_type'], 'Severity'] = 'critical'
df_scored['PRISM_Score_Final'] = df_scored['PRISM_Score'].astype(int)
df_scored['Severity_Final']    = pd.cut(df_scored['PRISM_Score_Final'], bins=scoring_bins, labels=label_bins, right=False)
df_scored.loc[df_scored['is_file_type'], 'Severity_Final'] = 'critical'

# Explanation
_NAME_MAP = {
    'malicious_raw_score':   'VT malicious (log-scaled)',
    'continuity_raw_score':  'Continuity (indicator type)',
    'tc_raw_rating':         'TC rating',
    'tc_raw_confidence':     'TC confidence',
    'tor_activity_score':    'TOR activity',
    'incidents_events_score':'Incident/Event association',
    'sources_raw_score':     'Multi-source validation',
    'partner_raw_score':     'Partner coverage bonus',
    'threat_actor_score':    'Threat actor association',
    'cal_raw_score':         'CAL score',
    'tc_threat_raw_score':   'TC threat assessment',
    'first_obs_raw_score':   'Recent first-seen activity',
    'stacked_context_bonus': 'Reinforcing context bonus',
}

def build_explanation(row):
    parts       = {k: float(row.get(k, 0) or 0) for k in _NAME_MAP.keys()}
    final       = row.get('PRISM_Score_Final')
    score       = float(final) if pd.notna(final) else float(row.get('PRISM_Score', 0) or 0)
    sev         = str(row.get('Severity_Final', row.get('Severity', 'nan')))
    current_date= datetime.now(UTC).strftime('%Y-%m-%d')
    vt_note     = (f"VT score: {int(row.get('vt_numeric_for_scoring', 0))}." if bool(row.get('vt_present', False)) else "VT score not available (neutral).")
    contrib     = sorted(parts.items(), key=lambda kv: abs(kv[1]), reverse=True)[:3]
    drivers_text= "; ".join(_NAME_MAP.get(k, k) for k, v in contrib if v != 0) or "No significant drivers"
    pb_reason   = row.get('pb_lower_reason')
    pb_flag     = bool(row.get('pb_lower_flag', False))
    pb_mult     = float(row.get('pb_base_multiplier', 1.0) or 1.0)
    pb_vt_bypass= bool(row.get('vt_high_floor_bypassed', False))
    pb_note     = (f"PB lower-start rule applied{f' ({pb_reason})' if pd.notna(pb_reason) and str(pb_reason).strip() else ''}; base score multiplier {pb_mult:.2f}; incident/event boost suppressed{'; VT high floor bypassed' if pb_vt_bypass else ''}."
                   if pb_flag else "No PB lower-start rule.")
    threat_actor_val = row.get('adversary')
    if pd.isna(threat_actor_val) or str(threat_actor_val).strip().lower() in {'none','nan',''}:
        threat_actor_val = row.get('threat_actor')
    actor_sentence = (f" Associated threat actor(s): {str(threat_actor_val).strip()}."
                      if threat_actor_val is not None and str(threat_actor_val).strip().lower() not in {'none','nan',''} else "")
    inc_flag    = int(row.get('incidents_events_flag', 0) or 0)
    inc_events  = str(row.get('incidents/events', '')).strip()
    inc_note    = (f"Linked to incident/event: {inc_events}." if inc_flag == 1 and not pb_flag and inc_events and inc_events.lower() not in {'none','nan',''}
                   else "Linked to incident/event, but incident/event boost suppressed by PB rule." if inc_flag == 1 and pb_flag
                   else "No incident/event link.")
    tagging_reason_val = row.get('Tagging_Boost_Reason')
    boost_reason_note  = (f"Tagging boost: {str(tagging_reason_val).strip()}." if pd.notna(tagging_reason_val) and str(tagging_reason_val).strip().lower() not in {'none','nan',''}
                          else "Tagging boost: criteria matched." if bool(row.get('Tagging_Boost', False)) else "")
    return (
        f"[{current_date}] Severity: {sev}. {vt_note} Contextual Drivers: {drivers_text}. "
        f"Observed across {int(row.get('partners_count', 0) or 0)} partner(s). "
        f"Observed by {int(row.get('sources_count', 1) or 1)} sources. "
        f"{pb_note} "
        f"{'Botnet penalty applied.' if float(row.get('botnet_penalty_multiplier', 1.0)) < 1.0 else 'No botnet penalty.'} "
        f"{'Scanner penalty applied.' if float(row.get('scanner_penalty_multiplier', 1.0)) < 1.0 else 'No scanner penalty.'} "
        f"{'TOR activity detected.' if float(row.get('tor_activity_score', 0) or 0) > 0 else 'No TOR activity.'} "
        f"{inc_note}{actor_sentence} {boost_reason_note} Score: {score:.0f}/1000."
    )

df_scored['AI_Adjustment'] = 0
df_scored['Explanation']   = df_scored.apply(build_explanation, axis=1)

if 'indicator' in df_scored.columns:
    df_scored.drop_duplicates(subset='indicator', inplace=True)

column_rename_map = {
    'indicator':'Indicator','type':'Indicator Type','lastObserved':'Last Observed',
    'vt_display':'VT Display','obs_count':'Observation Yearly Count',
    'rating':'ThreatConnect Rating','obs_penalty_multiplier':'Observation Penalty Multiplier',
    'botnet_flag':'Botnet Flag','falsePositives':'False Positives','partners':'Partners',
    'partners_count':'Partner Count','sources_count':'Source Count','adversary':'Adversary',
    'threat_actor':'Threat Actor','threat_nation_state':'Threat Nation State',
    'threat_security_org':'Threat Security Org','threat_cve_nbr':'Threat CVE',
    'Tagging_Boost':'Tagging Boost','Tagging_Boost_Reason':'Tagging Boost Reason',
    'pb_lower_flag':'PB Lower Flag','pb_lower_tags':'PB Lower Tags','pb_lower_reason':'PB Lower Reason',
    'pb_base_multiplier':'PB Base Multiplier','vt_high_floor_bypassed':'VT High Floor Bypassed',
    'calScore':'CAL Score','threatAssessScore':'ThreatConnect Score',
    'PRISM_Score':'PRISM Score','PRISM_Score_Final':'PRISM Score (Final)',
    'Severity':'Severity','Severity_Final':'Severity (Final)','Explanation':'Explanation',
}
df_scored.rename(columns=column_rename_map, inplace=True)
print(f"Scoring complete — {len(df_scored)} indicators scored")

# ── Step 10: Save to Excel ────────────────────────────────────────────────────

os.makedirs(OUTPUT_DIR, exist_ok=True)
excel_path = os.path.join(OUTPUT_DIR, EXCEL_FILENAME)

columns_to_save = [c for c in [
    'Indicator','Last Observed','Indicator Type','VirusTotal Malicious Score',
    'Observation Yearly Count','ThreatConnect Rating','Observation Penalty Multiplier',
    'Botnet Flag','False Positives','Partners','incidents/events','Threat Actor',
    'Threat Nation State','Threat Security Org','Threat CVE','Tagging Boost','Tagging Boost Reason',
    'CAL Score','ThreatConnect Score','PRISM Score','Severity','Explanation',
] if c in df_scored.columns]

df_export = df_scored[columns_to_save].copy()
if 'PRISM Score (Final)' in df_scored.columns:
    df_export['PRISM Score'] = df_scored['PRISM Score (Final)']
if 'Severity (Final)' in df_scored.columns:
    df_export['Severity'] = df_scored['Severity (Final)']

for col in df_export.columns:
    if pd.api.types.is_datetime64_any_dtype(df_export[col]) and df_export[col].dt.tz is not None:
        df_export[col] = df_export[col].dt.tz_convert('UTC').dt.tz_localize(None)

if os.path.exists(excel_path):
    df_existing = pd.read_excel(excel_path, engine='openpyxl')
    rename_map_existing = {old: new for old, new in column_rename_map.items() if old in df_existing.columns and new not in df_existing.columns}
    if rename_map_existing:
        df_existing.rename(columns=rename_map_existing, inplace=True)
    for col in columns_to_save:
        if col not in df_existing.columns:
            df_existing[col] = pd.NaT if col == 'Last Observed' else (0 if col in ['VirusTotal Malicious Score','Observation Yearly Count','ThreatConnect Rating','Observation Penalty Multiplier','Botnet Flag','False Positives','PRISM Score','ThreatConnect Score'] else '')
    df_existing      = df_existing[columns_to_save].copy()
    existing_set     = set(df_existing['Indicator'].values)
    new_set          = set(df_export['Indicator'].values)
    df_existing_idx  = df_existing.set_index('Indicator').sort_index()
    df_export_idx    = df_export.set_index('Indicator').sort_index()
    indicators_to_update   = [i for i in existing_set & new_set if not df_existing_idx.loc[i].equals(df_export_idx.loc[i])]
    indicators_unchanged   = [i for i in existing_set & new_set if i not in indicators_to_update]
    df_combined = pd.concat([
        df_existing[df_existing['Indicator'].isin(indicators_unchanged)],
        df_export[df_export['Indicator'].isin(indicators_to_update)],
        df_export[df_export['Indicator'].isin(new_set - existing_set)],
        df_existing[~df_existing['Indicator'].isin(new_set)],
    ], ignore_index=True).drop_duplicates(subset='Indicator', keep='last')
    print(f"Updated: {len(indicators_to_update)} | Added: {len(new_set - existing_set)} | Total: {len(df_combined)}")
else:
    df_combined = df_export.drop_duplicates(subset='Indicator', keep='last').copy()
    print(f"Created new file with {len(df_combined)} indicators")

df_complete_history = None
if os.path.exists(excel_path):
    try:
        df_complete_history = pd.read_excel(excel_path, sheet_name='Complete History', engine='openpyxl')
    except Exception:
        pass

fills = {
    'low':      PatternFill(start_color='83de85', end_color='83de85', fill_type='solid'),
    'medium':   PatternFill(start_color='eef084', end_color='eef084', fill_type='solid'),
    'high':     PatternFill(start_color='f29953', end_color='f29953', fill_type='solid'),
    'critical': PatternFill(start_color='e83f3f', end_color='e83f3f', fill_type='solid'),
}

with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    df_combined.to_excel(writer, index=False, sheet_name='PRISM Scores')
    if 'ThreatConnect Score' in df_combined.columns and 'PRISM Score' in df_combined.columns:
        df_comp = df_combined[['Indicator','ThreatConnect Score','PRISM Score']].copy()
        df_comp['ThreatConnect Score'] = pd.to_numeric(df_comp['ThreatConnect Score'], errors='coerce').fillna(0)
        df_comp['PRISM Score']         = pd.to_numeric(df_comp['PRISM Score'], errors='coerce').fillna(0)
        df_comp['Difference']          = df_comp['PRISM Score'] - df_comp['ThreatConnect Score']
        df_comp.to_excel(writer, index=False, sheet_name='Score Comparison')
    if df_complete_history is not None:
        df_complete_history.to_excel(writer, index=False, sheet_name='Complete History')
    worksheet = writer.sheets['PRISM Scores']
    for row_idx, severity in enumerate(df_combined['Severity'], start=2):
        fill = fills.get(str(severity).lower())
        if fill:
            for col_idx in range(1, len(df_combined.columns) + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = fill

print(f"Saved {len(df_combined)} indicators to {excel_path}")

# ── Step 11: Scoring history ──────────────────────────────────────────────────

run_timestamp   = datetime.now(UTC).strftime('%Y-%m-%d %H:%M:%S')
history_columns = ['Scoring Date','Indicator','Indicator Type','PRISM Score','Severity','Explanation']
column_mapping  = {
    'Indicator':'Indicator','Indicator Type':'Indicator Type',
    'PRISM Score (Final)':'PRISM Score','Severity (Final)':'Severity','Explanation':'Explanation',
}

cols_available     = {k: v for k, v in column_mapping.items() if k in df_scored.columns}
df_hist_slice      = df_scored[list(cols_available.keys())].rename(columns=cols_available)
# convert any categorical columns before fillna('') to avoid pandas restriction
for _col in df_hist_slice.select_dtypes(['category']).columns:
    df_hist_slice[_col] = df_hist_slice[_col].astype(str)
df_history_current = df_hist_slice.fillna('').assign(**{'Scoring Date': run_timestamp})
df_history_current = df_history_current[[c for c in history_columns if c in df_history_current.columns]]

if os.path.exists(excel_path):
    try:
        df_history_all = pd.read_excel(excel_path, sheet_name='Complete History', engine='openpyxl')
        for col in df_history_all.columns:
            if pd.api.types.is_datetime64_any_dtype(df_history_all[col]):
                if getattr(df_history_all[col].dtype, 'tz', None) is not None:
                    df_history_all[col] = df_history_all[col].dt.tz_convert('UTC').dt.tz_localize(None)
        df_history_all = pd.concat([
            df_history_all.dropna(axis=1, how='all'),
            df_history_current.dropna(axis=1, how='all'),
        ], ignore_index=True)
    except Exception:
        df_history_all = df_history_current.copy()
else:
    df_history_all = df_history_current.copy()

df_history_all['_date_only'] = pd.to_datetime(df_history_all['Scoring Date']).dt.date
df_history_all = df_history_all.drop_duplicates(subset=['Indicator','_date_only'], keep='last').drop(columns=['_date_only'])
df_history_all = df_history_all.sort_values(['Indicator','Scoring Date'], ascending=[True, False])
df_history_all = df_history_all[[c for c in history_columns if c in df_history_all.columns]]

existing_sheets = {}
if os.path.exists(excel_path):
    try:
        ef = pd.ExcelFile(excel_path, engine='openpyxl')
        for sn in ef.sheet_names:
            if sn not in ['Complete History','Latest Scores']:
                existing_sheets[sn] = pd.read_excel(excel_path, sheet_name=sn, engine='openpyxl')
        ef.close()
    except Exception as e:
        print(f"Warning: Could not read existing sheets: {e}")

with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    for sn, sdf in existing_sheets.items():
        sdf.to_excel(writer, index=False, sheet_name=sn)
        if sn == 'PRISM Scores' and 'Severity' in sdf.columns:
            ws = writer.sheets['PRISM Scores']
            for row_idx, severity in enumerate(sdf['Severity'], start=2):
                fill = fills.get(str(severity).lower())
                if fill:
                    for col_idx in range(1, len(sdf.columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill
    df_history_all.to_excel(writer, sheet_name='Complete History', index=False)
    ws_h = writer.sheets['Complete History']
    for dim, width in [('A', 20), ('B', 20), ('C', 15), ('D', 18), ('E', 12), ('F', 60)]:
        ws_h.column_dimensions[dim].width = width
    for cell in ws_h[1]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

print(f"Scoring history updated — {len(df_history_all)} total records, {df_history_all['Indicator'].nunique()} unique indicators")
print("Done.")
