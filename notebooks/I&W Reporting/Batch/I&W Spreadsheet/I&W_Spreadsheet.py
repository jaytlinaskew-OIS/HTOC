#!/usr/bin/env python3
"""
I&W Spreadsheet Processing Script

This script processes I&W (Indications & Warnings) indicators from ThreatConnect API
and generates Excel spreadsheets with clickable hyperlinks.

Converted from: I&W_Spreadsheet.ipynb
Author: GitHub Copilot
Created: October 8, 2025
"""

import sys
import os
import urllib3
import pandas as pd
from datetime import datetime, timedelta
import pytz
import urllib.parse
import warnings
import time
import logging
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter


def main():
    """Main execution function."""
    print("Starting I&W Spreadsheet Processing...")
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # CONFIGURATION & SETUP
    # ═══════════════════════════════════════════════════════════════════════════════
    
    # Add your local ThreatConnect SDK to path
    sys.path.append("Z:/HTOC/Data_Analytics/threatconnect")
    from ThreatConnect import ThreatConnect
    from RequestObject import RequestObject

    # Load API config - using the same approach as main.py
    project_root = r"Z:\HTOC\HTOC Reports\I&W Reports\5. I&W Staging\I&W Report Processing Scripts"
    if project_root not in sys.path:
        sys.path.append(project_root)

    # Add the scripts directory to the path to import config_loader
    scripts_path = os.path.join(project_root, "scripts")
    if scripts_path not in sys.path:
        sys.path.append(scripts_path)

    from config_loader import get_threatconnect_config

    config_path = os.path.join(project_root, "utils", "config.json")
    try:
        tc_config = get_threatconnect_config(config_path)
        api_secret_key = tc_config["secret_key"]
        api_access_id = tc_config["access_id"]
        api_base_url = tc_config["base_url"]
        api_default_org = tc_config["default_org"]
        print(f"Loaded config from: {config_path}")
        print(f"Base URL: {api_base_url}")
        print(f"Access ID: {api_access_id}")
        print(f"Default Org: {api_default_org}")
    except Exception as e:
        print(f"[ERROR] Failed to load configuration: {e}")
        sys.exit(1)

    # Disable SSL verification warnings (use cautiously)
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    # Initialize ThreatConnect session
    try:
        tc = ThreatConnect(api_access_id, api_secret_key, api_default_org, api_base_url)
        print("ThreatConnect initialized.")
    except Exception as e:
        print(f"[ERROR] Failed to initialize ThreatConnect: {e}")
        sys.exit(1)

    # Define the owner (organization scope)
    owner = 'HTOC Org'

    # Create a request object to fetch indicators (or other data)
    try:
        ro = RequestObject()
        ro.set_http_method('GET')
        ro.set_owner(owner)
        ro.set_owner_allowed(True)
        # ro.set_resource_pagination(True)  # Uncomment if needed
        print("RequestObject successfully created.")
    except Exception as e:
        print(f"[ERROR] Failed to initialize RequestObject: {e}")
        sys.exit(1)

    # ═══════════════════════════════════════════════════════════════════════════════
    # DATA COLLECTION - ThreatConnect API Query
    # ═══════════════════════════════════════════════════════════════════════════════
    
    print("Querying ThreatConnect API...")
    
    # Define time period
    # Calculate the start date (2 days ago) at 00:00:00 UTC
    start_date = (datetime.now(pytz.UTC) - timedelta(days=2)).date()

    # Format as 'YYYY-MM-DDT00:00:00Z'
    start = f"{start_date}T00:00:00Z"

    # List of owners to pull from
    list_of_owners = ['HTOC Org']
    final_results = []
    type_names = [
        "Address", "EmailAddress", "File", "Host", "URL", "ASN", "CIDR", 
        "Email Subject", "Hashtag", "Mutex", "Registry Key", "Stripped URL", "User Agent"
    ]
    type_name_condition = ", ".join([f'"{t}"' for t in type_names])

    for owner in list_of_owners:
        print(f"\nQuerying owner: {owner}")
        try:
            tql_raw = (
                f'ownerName EQ "{owner}" AND '
                f'typeName IN ({type_name_condition})'
                f'lastObserved >= "{start}"'
            )
            tql_encoded = urllib.parse.quote(tql_raw)

            ro.set_http_method('GET')
            ro.set_request_uri(
                f'/v3/indicators?tql={tql_encoded}&fields=tags,observations,associatedGroups&resultStart=0&resultLimit=10000'
            )

             # Send the request
            response = tc.api_request(ro)

            # Parse response
            if response.headers.get('content-type') == 'application/json':
                results = response.json()
                final_results.append(results)
            else:
                print(f"Unexpected response format: {response.headers.get('content-type')}")

        except Exception as e:
            print(f"Failed to query indicators for {owner}: {e}")

    # Normalize and clean results
    if final_results:
        # Extract and normalize data only if 'data' key exists and contains 'summary'
        normalized_data = []
        for result in final_results:
            if 'data' in result:
                for item in result['data']:
                    if 'summary' in item:
                        normalized_data.append(item)

        if normalized_data:
            observed_src = pd.json_normalize(normalized_data)
            observed_src['indicator'] = observed_src['summary'].str.split(' ', expand=True)[0].str.strip()
            observed_src = observed_src.drop_duplicates(subset='indicator', keep='first')  # Remove duplicates based on 'indicator'
            observed_src = observed_src[observed_src['lastObserved'] >= start]
            print(f"\nRetrieved {len(observed_src)} unique observed indicators.")
        else:
            print("\nNo valid indicators with 'summary' key retrieved.")
            return
    else:
        print("\nNo indicators retrieved.")
        return

    # ═══════════════════════════════════════════════════════════════════════════════
    # DATA COLLECTION - Observed Data Files
    # ═══════════════════════════════════════════════════════════════════════════════
    
    print("Loading observed data files...")
    
    # Base file path with placeholder for date
    base_path = r"Z:/HTOC/Data_Analytics/Data/OpDiv_Observations/htoc_opdiv_obs_d{date}.csv"
    date_format = "%Y%m%d"

    # Fetch file paths for the last 3 days
    file_paths = get_file_paths(base_path, days=3)

    # Load observed data
    observed_data_df = load_observed_data(file_paths)

    # ═══════════════════════════════════════════════════════════════════════════════
    # DATA PROCESSING PIPELINE
    # ═══════════════════════════════════════════════════════════════════════════════
    
    print("Starting data processing pipeline...")
    
    # Process tags and apply filters
    recent_tags = process_data_pipeline(observed_src, observed_data_df)
    
    if recent_tags.empty:
        print("No indicators met the filtering criteria. Exiting.")
        return

    # ═══════════════════════════════════════════════════════════════════════════════
    # ATTRIBUTE COLLECTION
    # ═══════════════════════════════════════════════════════════════════════════════
    
    print("Collecting indicator attributes...")
    
    # Extract unique indicators from recent_tags
    indicators = recent_tags['summary'].unique()

    # Initialize a list to store attribute data
    attributes_data = []

    # Track indicators with no attributes
    indicators_with_no_attributes = []

    # Iterate over unique indicators
    for indicator in indicators:
        try:
            ro.set_http_method('GET')
            ro.set_request_uri(f'/v3/indicators/{indicator}?fields=attributes&resultStart=0&resultLimit=1000')
            response = tc.api_request(ro)

            if response.headers.get('content-type') == 'application/json':
                data = response.json().get('data', {})
                attributes = data.get('attributes', {}).get('data', [])

                if not attributes:
                    print(f"No attributes found for indicator: {indicator}")
                    # Track indicators with no attributes
                    indicators_with_no_attributes.append(indicator)
                else:
                    for attr in attributes:
                        attr.update({
                            'id': data.get('id'),
                            'summary': data.get('summary'),
                            'type': data.get('type'),
                            'ownerName': data.get('ownerName')
                        })
                        attributes_data.append(attr)
            # Suppress non-JSON and known 400 error responses
        except Exception as e:
            # Suppress error output, including known 400 error
            if hasattr(e, 'response') and getattr(e.response, 'status_code', None) == 400:
                continue
            if "Status Code: 400" in str(e):
                continue
            pass

    # Create attributes 
    attributes_observed_src = pd.json_normalize(attributes_data)

    # Un-nest 'createdBy' and filter out 'SOAR' entries
    if not attributes_observed_src.empty and 'createdBy.lastName' in attributes_observed_src.columns and attributes_observed_src['createdBy.lastName'].notnull().any():
        attributes_observed_src = attributes_observed_src[attributes_observed_src['createdBy.lastName'] != 'SOAR']

    # Drop duplicates based on 'id' to keep unique attribute records
    attributes_observed_src = attributes_observed_src.drop_duplicates(subset='id').reset_index(drop=True)

    # Filter recent_tags for indicators that had attributes
    filtered_with_attrs = recent_tags[recent_tags['summary'].isin(attributes_observed_src['summary'])]

    # Filter recent_tags for indicators that had no attributes
    no_attrs_df = recent_tags[recent_tags['summary'].isin(indicators_with_no_attributes)]

    # Combine both and remove duplicates based on 'summary'
    filtered_recent_tags = pd.concat([filtered_with_attrs, no_attrs_df], ignore_index=True)
    filtered_recent_tags = filtered_recent_tags.drop_duplicates(subset='summary').reset_index(drop=True)
    
    print(f"Final filtered dataset: {len(filtered_recent_tags)} indicators")

    # Check if we have any data before creating Excel file
    if len(filtered_recent_tags) == 0:
        print("No data to export. Excel file will not be created.")
        return

    # ═══════════════════════════════════════════════════════════════════════════════
    # EXCEL OUTPUT GENERATION
    # ═══════════════════════════════════════════════════════════════════════════════
    
    print("Generating Excel output...")
    
    # Save all columns from filtered_recent_tags to Excel with clickable hyperlinks for 'webLink'
    output_dir = r"Z:\HTOC\HTOC Reports\I&W Reports\5. I&W Staging\Spreadsheet"
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(
        output_dir,
        f"I&W_indicators_full_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx"
    )

    try:
        import openpyxl
        from openpyxl.styles import Font

        # Prepare data for Excel - convert all complex data types to strings
        excel_data = filtered_recent_tags.copy()
        
        # Convert timezone-aware datetime columns to timezone-naive
        for col in excel_data.columns:
            if excel_data[col].dtype == 'datetime64[ns, UTC]' or (excel_data[col].dtype == 'object' and 
                                                                 excel_data[col].apply(lambda x: hasattr(x, 'tzinfo') and x.tzinfo is not None).any()):
                excel_data[col] = pd.to_datetime(excel_data[col], errors='coerce').dt.tz_convert(None)
        
        # Convert complex data types to strings
        for col in excel_data.columns:
            if excel_data[col].dtype == 'object':  # Check if column might contain complex objects
                excel_data[col] = excel_data[col].apply(
                    lambda x: ', '.join(map(str, x)) if isinstance(x, list) else str(x) if x is not None else ''
                )

        # Create a new workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "I&W Indicators Full"

        # Write headers
        for col_idx, col_name in enumerate(excel_data.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data rows
        for row_idx, row in enumerate(excel_data.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                col_name = excel_data.columns[col_idx - 1]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Make 'webLink' column clickable
                if col_name == 'webLink' and pd.notna(value) and value != '':
                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")

        # Auto-adjust column widths
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        wb.save(excel_path)
        print(f"Successfully saved I&W indicators data to Excel: {excel_path}")

    except ImportError:
        print("ERROR: openpyxl not available. Install with: pip install openpyxl")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: Failed to create Excel file: {e}")
        sys.exit(1)

    print("I&W Spreadsheet processing completed successfully!")


def get_file_paths(base_path, days=3):
    """Generate file paths for the last `days` days using list comprehension."""
    today = datetime.utcnow()
    dates_to_pull = [(today - timedelta(days=i)).strftime("%Y%m%d") for i in range(days)]
    
    # Construct file paths
    file_paths = [base_path.format(date=dt) for dt in dates_to_pull]
    
    # Filter for existing files
    existing_files = [file_path for file_path in file_paths if os.path.exists(file_path)]
    
    if not existing_files:
        print("No files found for the specified date range.")
    else:
        print(f"Files to be loaded: {existing_files}")
    
    return existing_files


def load_observed_data(file_paths):
    """Load and concatenate observed data from multiple files."""
    data_frames = []

    for file_path in file_paths:
        try:
            df = pd.read_csv(file_path)
            data_frames.append(df)
        except Exception as e:
            print(f"Error reading file {file_path}: {e}")
    
    # Concatenate data
    if data_frames:
        observed_data_df = pd.concat(data_frames, ignore_index=True)
        print(f"Loaded data from {len(data_frames)} files.")
    else:
        observed_data_df = pd.DataFrame()

    return observed_data_df


def process_data_pipeline(observed_src, observed_data_df):
    """Process the data pipeline with filtering and grouping."""
    warnings.simplefilter(action='ignore', category=pd.errors.SettingWithCopyWarning)

    # Time cutoffs
    cutoff = pd.Timestamp.utcnow()
    date_added_cutoff = cutoff - timedelta(days=180)
    cutoff_naive = cutoff.tz_convert(None)

    # Required columns validation
    required_columns = ['indicator', 'OpDiv', 'obs_date']
    missing_columns = [c for c in required_columns if c not in observed_data_df.columns]
    if missing_columns:
        raise ValueError(f"Missing columns in observed data: {missing_columns}")

    print("Starting tag processing pipeline...")

    # Step 1: Process tags from observed_src
    print("Processing tags from observed_src...")
    all_filtered = []

    for _, row in observed_src.iterrows():
        processed_tags = process_tag_row(row, observed_src)
        if processed_tags is not None:
            all_filtered.append(processed_tags)

    # Step 2: Create filtered_tags DataFrame
    print("Creating filtered_tags DataFrame...")
    filtered_tags = pd.concat(all_filtered, ignore_index=True) if all_filtered else pd.DataFrame()

    if not filtered_tags.empty:
        # Ensure datetime columns
        filtered_tags['lastObserved'] = pd.to_datetime(filtered_tags['lastObserved'], errors='coerce', utc=True)
        filtered_tags['dateAdded'] = pd.to_datetime(filtered_tags['dateAdded'], errors='coerce', utc=True)
        print(f"Created filtered_tags with {len(filtered_tags)} rows")
    else:
        print("No filtered tags found")
        return pd.DataFrame()

    # Step 3: Map observed dates
    print("Mapping observed dates...")
    filtered_tags = map_observed_dates(filtered_tags, observed_data_df)

    # Step 3.5: Get indicators with multiple partners from observed_data_df
    print("Identifying indicators with multiple partners from observed_data_df...")
    multi_partner_indicators = get_multi_partner_indicators(observed_data_df, cutoff_naive)
    print(f"Found {len(multi_partner_indicators)} indicators with multiple partners")

    # Step 4: Apply filters and grouping
    print("Applying filters and partner grouping...")
    recent_tags = apply_filters_and_grouping(filtered_tags, cutoff, date_added_cutoff, cutoff_naive, multi_partner_indicators)

    # Step 5: Extract group IDs
    print("Extracting group IDs...")
    recent_tags = extract_group_ids(recent_tags)

    # Step 6: Add I&W column indicating if indicator has been tagged as 'I&W' (used in previous report)
    if 'all_tags' in recent_tags.columns:
        recent_tags['I&W'] = recent_tags['all_tags'].apply(
            lambda x: 'Yes' if isinstance(x, list) and 'I&W' in x else 'No'
        )
    else:
        recent_tags['I&W'] = 'No'

    # Move I&W column to the very end
    cols = [col for col in recent_tags.columns if col != 'I&W'] + ['I&W']
    recent_tags = recent_tags[cols]

    # Final summary
    print(f"Processing complete! Final dataset shape: {recent_tags.shape}")
    if not recent_tags.empty:
        print(f"Partners represented: {recent_tags['partners'].nunique()} unique partner combinations")
        print(f"Date range: {recent_tags['observed_date'].min()} to {recent_tags['observed_date'].max()}")

    return recent_tags


def process_tag_row(row, observed_src):
    """Process a single row from observed_src to extract and filter tags."""
    tags_data = row.get('tags.data')
    if not isinstance(tags_data, list):
        return None

    # Normalize the tags for the current row
    tags_df = pd.json_normalize(tags_data)

    # Ensure string and apply VA rename before filtering
    tags_df['name'] = (
        tags_df['name']
        .astype(str)
        .str.replace('VA CSOC CTS Splunk', 'VA Splunk API', regex=False)
    )

    # Skip if all associated groups are Adversary
    if 'associatedGroups.data' in observed_src.columns:
        ag_data = row.get('associatedGroups.data')
        if isinstance(ag_data, list) and len(ag_data) > 0:
            groups_df = pd.json_normalize(ag_data)
            if 'type' in groups_df.columns and set(groups_df['type']) == {'Adversary'}:
                return None

    # All tag names (for all_tags)
    all_tags_list = tags_df['name'].tolist()

    # Filter for API tags
    api_tags = tags_df[tags_df['name'].str.contains('API', case=False, na=False)].copy()
    if api_tags.empty:
        return None

    # Add metadata columns
    meta_cols = [
        'summary', 'observations', 'description', 'type',
        'dateAdded', 'lastModified', 'lastObserved', 'webLink',
        'rating', 'confidence', 'id', 'associatedGroups.data'
    ]
    for col in meta_cols:
        api_tags[col] = [row.get(col)] * len(api_tags)

    # Attach all tags list
    if len(api_tags) > 0:
        api_tags['all_tags'] = [all_tags_list] * len(api_tags)

    return api_tags


def map_observed_dates(filtered_tags, observed_data_df):
    """Map observed dates from observed_data_df to filtered_tags."""
    if filtered_tags.empty:
        return filtered_tags
    
    # Clean name -> match OpDiv (remove trailing ' Splunk API')
    filtered_tags['cleaned_name'] = filtered_tags['name'].str.replace(r'\s+Splunk API$', '', regex=True)
    filtered_tags['observed_date'] = pd.NaT
    
    # Ensure observed_data_df obs_date is datetime (naive)
    observed_data_df['obs_date'] = pd.to_datetime(observed_data_df['obs_date'], errors='coerce')

    for idx, r in filtered_tags.iterrows():
        summary = r.get('summary')
        cleaned_name = r.get('cleaned_name')
        if pd.isna(summary) or pd.isna(cleaned_name):
            continue
        match = observed_data_df[
            (observed_data_df['indicator'] == summary) &
            (observed_data_df['OpDiv'] == cleaned_name)
        ]
        if not match.empty:
            filtered_tags.at[idx, 'observed_date'] = match['obs_date'].iloc[0]

    filtered_tags['observed_date'] = pd.to_datetime(filtered_tags['observed_date'], errors='coerce')
    
    # Drop helper column
    filtered_tags.drop(columns=['cleaned_name'], inplace=True, errors='ignore')
    
    return filtered_tags


def get_multi_partner_indicators(observed_data_df, cutoff_naive):
    """Get indicators that have multiple partners from observed_data_df."""
    if observed_data_df.empty:
        return pd.DataFrame()
    
    # Ensure obs_date is datetime
    observed_data_df['obs_date'] = pd.to_datetime(observed_data_df['obs_date'], errors='coerce')
    
    # Filter by recent dates (last 3 days)
    recent_obs = observed_data_df[
        observed_data_df['obs_date'] >= cutoff_naive - timedelta(days=3)
    ].copy()
    
    if recent_obs.empty:
        return pd.DataFrame()
    
    # Group by indicator and count unique OpDiv (partners)
    partner_counts = (
        recent_obs.groupby('indicator')['OpDiv']
        .agg(['nunique', lambda s: ', '.join(sorted(set(s.dropna())))]).reset_index()
        .rename(columns={'nunique': 'partner_count', '<lambda_0>': 'partners'})
    )
    
    # Keep only indicators with multiple partners
    multi_partner_indicators = partner_counts[partner_counts['partner_count'] >= 2].copy()
    
    return multi_partner_indicators


def apply_filters_and_grouping(filtered_tags, cutoff, date_added_cutoff, cutoff_naive, multi_partner_indicators):
    """Apply time filters, partner grouping, and other filtering criteria."""
    if filtered_tags.empty:
        return pd.DataFrame()
    
    # Time-based filters
    # Last 2 days by observed_date (naive)
    recent_tags = filtered_tags[filtered_tags['observed_date'] >= cutoff_naive - timedelta(days=2)].copy()

    if recent_tags.empty:
        return recent_tags
    
    # Filter to only include indicators that have multiple partners in observed_data_df
    if not multi_partner_indicators.empty:
        recent_tags = recent_tags[
            recent_tags['summary'].isin(multi_partner_indicators['indicator'])
        ].copy()
    
    if recent_tags.empty:
        return recent_tags
    
    # Partner extraction and grouping from ThreatConnect tags (as fallback)
    recent_tags['partner'] = recent_tags['name'].str.replace(' Splunk API', '', regex=False)

    partner_groups = (
        recent_tags.groupby('summary')['partner']
        .agg(['nunique', lambda s: ', '.join(sorted(set(s.dropna())))]).reset_index()
        .rename(columns={'nunique': 'partner_count', '<lambda_0>': 'partners'})
    )

    recent_tags = recent_tags.merge(partner_groups, on='summary', how='left')
    
    # Merge with multi_partner_indicators to get partners from observed_data_df
    if not multi_partner_indicators.empty:
        recent_tags = recent_tags.merge(
            multi_partner_indicators[['indicator', 'partners', 'partner_count']],
            left_on='summary',
            right_on='indicator',
            how='left',
            suffixes=('', '_from_obs')
        )
        # Use partners from observed_data_df if available, otherwise use from tags
        recent_tags['partners'] = recent_tags['partners_from_obs'].fillna(recent_tags['partners'])
        recent_tags['partner_count'] = recent_tags['partner_count_from_obs'].fillna(recent_tags['partner_count'])
        recent_tags = recent_tags.drop(columns=['indicator', 'partners_from_obs', 'partner_count_from_obs'], errors='ignore')

    # Additional filters
    recent_tags = recent_tags[recent_tags['partner_count'] >= 2]
    recent_tags = recent_tags[recent_tags['observations'] < 15000]
    recent_tags = recent_tags[recent_tags['dateAdded'] >= date_added_cutoff]

    # Deduplicate by summary
    recent_tags = recent_tags.drop_duplicates(subset='summary', keep='first')

    # Drop unneeded columns if present
    cols_to_drop = [
        'techniqueId', 'tactics.data', 'tactics.count',
        'platforms.data', 'platforms.count', 'partner', 'name'
    ]
    recent_tags = recent_tags.drop(columns=[c for c in cols_to_drop if c in recent_tags.columns], errors='ignore')

    # Remove rows where all_tags contains unwanted markers (keep htoc_wl filter, but not I&W)
    if 'all_tags' in recent_tags.columns:
        recent_tags = recent_tags[~recent_tags['all_tags'].apply(lambda x: isinstance(x, list) and 'htoc_wl' in x)]
    
    # Add partners from tags at the end (after all filtering)
    # Re-extract partners from tags for the final filtered dataset
    if not recent_tags.empty and 'all_tags' in recent_tags.columns:
        def extract_partners_from_tags(all_tags_list):
            """Extract partner names from tags that contain 'API'."""
            if not isinstance(all_tags_list, list):
                return []
            api_partners = []
            for tag in all_tags_list:
                if isinstance(tag, str) and 'API' in tag:
                    # Extract partner name (remove ' Splunk API' suffix)
                    partner = tag.replace(' Splunk API', '').replace('VA CSOC CTS Splunk', 'VA').strip()
                    if partner:
                        api_partners.append(partner)
            return sorted(set(api_partners))
        
        # Extract partners from tags for each indicator
        tag_partners_series = recent_tags['all_tags'].apply(extract_partners_from_tags)
        
        # Combine with existing partners from observed_data_df
        def combine_partners(row_idx):
            """Combine partners from observed_data_df and tags for a specific row."""
            obs_partners = recent_tags.loc[row_idx, 'partners']
            tag_partners_list = tag_partners_series.loc[row_idx]
            
            combined = set()
            # Add partners from observed_data_df
            if pd.notna(obs_partners) and obs_partners:
                for p in str(obs_partners).split(', '):
                    if p.strip():
                        combined.add(p.strip())
            # Add partners from tags
            if tag_partners_list:
                for p in tag_partners_list:
                    if p:
                        combined.add(p)
            return ', '.join(sorted(combined)) if combined else obs_partners
        
        recent_tags['partners'] = recent_tags.index.to_series().apply(combine_partners)
        
        # Update partner_count based on combined partners
        recent_tags['partner_count'] = recent_tags['partners'].apply(
            lambda x: len([p for p in str(x).split(', ') if p.strip()]) if pd.notna(x) and x else 0
        )

    return recent_tags


def extract_group_ids(recent_tags):
    """Extract group IDs from associatedGroups.data."""
    if 'associatedGroups.data' in recent_tags.columns:
        recent_tags['group_id'] = recent_tags['associatedGroups.data'].apply(
            lambda x: x[0]['id'] if isinstance(x, list) and len(x) > 0 and 'id' in x[0] else None
        )
        # Convert group_id to string type and remove trailing decimals if it exists
        if 'group_id' in recent_tags.columns:
            recent_tags['group_id'] = recent_tags['group_id'].apply(
                lambda x: str(int(float(x))) if pd.notna(x) and str(x) != 'None' else x
            ).astype(str)
    return recent_tags


if __name__ == "__main__":
    main()