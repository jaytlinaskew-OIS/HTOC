r"""
NextObservedIndicator V4 — performance evaluation module.

Evaluates forecast bands against observed outcomes and writes growing
Excel workbooks under {SAVE_ROOT}\Performance:
  - performance_1day.xlsx (overall + OpDiv sheets)
  - performance_7day.xlsx, etc.

Call init_performance_eval(save_root) before run_performance_evaluation().

Ground truth is gated by a healthy-day mask: an observation day counts for an
OpDiv only if the file exists, the day has settled (OBS_SETTLE_DAYS old), and
that OpDiv appears in it. An OpDiv is scored at horizon H only when every day
in the label window is healthy, because a missing observation is not a negative
observation. Daily runs therefore score OBS_SETTLE_DAYS days back; the most
recent evaluation dates self-correct on later runs, older rows require a
backfill. See the Legend sheet of any workbook for the reader-facing version.
"""

from __future__ import annotations

import os
import re
import sys
import traceback
from datetime import datetime, timedelta

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import noi_v4_feed_health as feed_health
from noi_v4_bands import BAND_HIGH_P_OPDIV
from noi_v4_feed_health import FeedHealth

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HTOC_SHARE_ROOT = os.environ.get("HTOC_SHARE_ROOT", r"\\cscso1fsappv01\data\HTOC").strip()
SAVE_ROOT = os.environ.get(
    "NOI_V4_SAVE_DIR",
    os.path.join(HTOC_SHARE_ROOT, r"JA\NextObserveV4Test"),
).strip()
DATA_PATH = SAVE_ROOT
SAVE_PATH = os.path.join(SAVE_ROOT, "Full Daily Reports")
PERF_DIR = os.path.join(SAVE_ROOT, "Performance")
PERF_ALERTS_DIR = os.path.join(PERF_DIR, "Alerts")
PERF_LOG_DIR = os.path.join(PERF_DIR, "Logs")
EXCLUDE_FOLDERS = {"automation scripts", "Logs", "LogsBackup", "Full Daily Reports",
                   "Performance", "Possibly Active Review", "Alerts"}


def _perf_log(message: str, level: str = "INFO") -> None:
    """Print and append to the daily performance log on the share."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[PERF {ts}] {level}: {message}"
    print(line, flush=True)
    try:
        os.makedirs(PERF_LOG_DIR, exist_ok=True)
        log_fp = os.path.join(PERF_LOG_DIR, f"perf_eval_{datetime.today().strftime('%Y%m%d')}.log")
        with open(log_fp, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception as log_err:
        print(f"[PERF {ts}] WARNING: could not write perf log: {log_err}", flush=True)


def log_perf_error(context: str, exc: BaseException | None = None) -> None:
    """Log an eval error with full traceback to stdout and Performance/Logs."""
    if exc is None:
        tb = traceback.format_exc().strip()
        summary = context
    else:
        tb = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__)).strip()
        summary = f"{context}: {exc}"
    _perf_log(summary, level="ERROR")
    if tb and tb != "NoneType: None":
        for tb_line in tb.splitlines():
            _perf_log(tb_line, level="ERROR")


def init_performance_eval(save_root: str, htoc_share_root: str | None = None) -> None:
    """Configure paths for evaluation (call once before running eval)."""
    global SAVE_ROOT, DATA_PATH, SAVE_PATH, PERF_DIR, PERF_ALERTS_DIR, PERF_LOG_DIR, OBS_TEMPLATE, HTOC_SHARE_ROOT
    SAVE_ROOT = save_root.strip()
    DATA_PATH = SAVE_ROOT
    SAVE_PATH = os.path.join(SAVE_ROOT, "Full Daily Reports")
    PERF_DIR = os.path.join(SAVE_ROOT, "Performance")
    PERF_ALERTS_DIR = os.path.join(PERF_DIR, "Alerts")
    PERF_LOG_DIR = os.path.join(PERF_DIR, "Logs")
    if htoc_share_root:
        HTOC_SHARE_ROOT = htoc_share_root.strip()
    OBS_TEMPLATE = os.environ.get(
        "HTOC_OBS_TEMPLATE",
        os.path.join(HTOC_SHARE_ROOT, r"Data_Analytics\Data\OpDiv_Observations\htoc_opdiv_obs_d{date}.csv"),
    )

V4_NAME = re.compile(r"^.+_output_(\d{8})\.csv$", re.IGNORECASE)
LEGACY_NAME = re.compile(r"^(\d{8})\.csv$", re.IGNORECASE)

OBS_TEMPLATE = os.environ.get(
    "HTOC_OBS_TEMPLATE",
    os.path.join(HTOC_SHARE_ROOT, r"Data_Analytics\Data\OpDiv_Observations\htoc_opdiv_obs_d{date}.csv"),
)
DATE_FMT = "%Y%m%d"
EVAL_HORIZONS = [1, 7, 14, 30, 45]
_horizons_override = os.environ.get("NOI_V4_EVAL_HORIZONS")
if _horizons_override:
    try:
        EVAL_HORIZONS = [int(x.strip()) for x in _horizons_override.split(",") if x.strip()]
    except Exception:
        _perf_log(
            f"could not parse NOI_V4_EVAL_HORIZONS={_horizons_override}; using default {EVAL_HORIZONS}",
            level="WARNING",
        )

ROLLING_BASELINE_DAYS = int(os.environ.get("NOI_V4_PERF_ROLLING_BASELINE_DAYS", "14"))
MIN_DECIDED_COUNT = int(os.environ.get("NOI_V4_PERF_MIN_DECIDED_COUNT", "50"))
HIGH_PREC_ABS_MIN = float(os.environ.get("NOI_V4_PERF_HIGH_PREC_ABS_MIN", "0.90"))
HIGH_PREC_DROP_PP_ROLLING = float(os.environ.get("NOI_V4_PERF_HIGH_PREC_DROP_PP_ROLLING", "0.05"))
HIGH_PREC_DROP_PP_DAY = float(os.environ.get("NOI_V4_PERF_HIGH_PREC_DROP_PP_DAY", "0.03"))
LOW_NEG_PREC_ABS_MIN = float(os.environ.get("NOI_V4_PERF_LOW_NEG_PREC_ABS_MIN", "0.85"))
LOW_NEG_PREC_DROP_PP_ROLLING = float(os.environ.get("NOI_V4_PERF_LOW_NEG_PREC_DROP_PP_ROLLING", "0.05"))
ACC_ABS_MIN = float(os.environ.get("NOI_V4_PERF_ACC_ABS_MIN", "0.80"))
ACC_DROP_PP_ROLLING = float(os.environ.get("NOI_V4_PERF_ACC_DROP_PP_ROLLING", "0.05"))
# Target metric is recall against every actual positive, including abstain.
# Decided-only recall hid the Possibly Active misses. Precision is a floor.
RECALL_ALL_ABS_MIN = float(os.environ.get("NOI_V4_PERF_RECALL_ALL_ABS_MIN", "0.60"))
RECALL_ALL_OPDIV_ABS_MIN = float(os.environ.get("NOI_V4_PERF_RECALL_ALL_OPDIV_ABS_MIN", "0.50"))
RECALL_ALL_DROP_PP_ROLLING = float(os.environ.get("NOI_V4_PERF_RECALL_ALL_DROP_PP_ROLLING", "0.05"))
RECALL_ABS_MIN = float(os.environ.get("NOI_V4_PERF_RECALL_ABS_MIN", "0.60"))

# ---- Forecast coverage ----
# How far back to report on forecast availability.
COVERAGE_LOOKBACK_DAYS = int(os.environ.get("NOI_V4_PERF_COVERAGE_LOOKBACK_DAYS", "14"))
# A gap this recent means the forecast job is failing now and is worth an alert.
COVERAGE_ALERT_DAYS = int(os.environ.get("NOI_V4_PERF_COVERAGE_ALERT_DAYS", "3"))

# ---- Healthy-day mask ----
# Feed health lives in noi_v4_feed_health so the forecast runner's training
# labels and these metrics agree on what "usable ground truth" means.
OBS_SETTLE_DAYS = feed_health.SETTLE_DAYS

# Shown instead of a number when a metric has no meaningful value. Three
# different situations produce a blank-looking cell and they are not the same
# problem, so they do not share a label:
#
#   MISSING_METRIC  the ground truth is genuinely absent -- the feed was down or
#                   the file never arrived. Nothing can fix this but a re-pull.
#   NOT_SETTLED     the data is present but the upstream job is still rewriting
#                   it. Scoring now would count late arrivals as misses. This
#                   one corrects itself once the day ages past the settle window.
#   NOT_APPLICABLE  the data is fine and the day was scored, but this particular
#                   metric has no population -- no High predictions makes
#                   precision 0/0, and no observed positives makes Accuracy and
#                   Negative Precision read 100% purely because there was
#                   nothing to miss. That last case is the dangerous one: left
#                   as a number it renders green, a perfect score for a day that
#                   never tested the model.
MISSING_METRIC = "Missing Data to compute"
NOT_SETTLED = "Not settled yet"
NOT_APPLICABLE = "Not applicable"
_SENTINELS = frozenset({MISSING_METRIC, NOT_SETTLED, NOT_APPLICABLE})

# Provenance for a forecast the scheduled job missed and a replay rebuilt later.
BACKFILL_MARKER_NAME = "backfilled_forecasts.txt"
BACKFILLED_NOTE = "forecast backfilled by as-of replay, not produced live"
_BACKFILL_CACHE: set[str] | None = None

# A row scored off days that reached normal volume before formally settling.
PROVISIONAL_NOTE = "provisional - scored before the label window settled"

# Traffic-light thresholds for percentage metric columns (stored as 0-100 floats).
# Tuple is (green_min, yellow_min): value >= green -> green; >= yellow -> yellow; else red.
METRIC_TRAFFIC_LIGHTS = {
    "Precision - High (%)": (90.0, 85.0),
    "Recall - High vs All Positives (%)": (75.0, 60.0),
    "Recall - 7-Day High vs 1-Day Positives (%)": (85.0, 75.0),
    "Recall - 1-Day or 7-Day High vs All Positives (%)": (85.0, 75.0),
    "Accuracy - High (%)": (85.0, 70.0),
    "Negative Precision - Low (%)": (94.0, 90.0),
    "Accuracy (%)": (95.0, 90.0),
    "Specificity - True Negative Rate (%)": (95.0, 90.0),
    "Balanced Accuracy (%)": (88.0, 80.0),
    "Recall - High (%)": (85.0, 70.0),
    "F1 Score - High (%)": (90.0, 85.0),
    "Coverage (%)": (85.0, 75.0),
}

# Legend copy: what each measure is and why it matters.
METRIC_DESCRIPTIONS = {
    "Precision - High (%)": (
        "Of indicators labeled Highly likely, the percent that were actually observed within the "
        "horizon. Correct High calls divided by all High calls.",
        "This is the noise floor, not the objective. Keep it at or above 90% so analysts still "
        "trust High flags. Do not maximize it: 100% precision with half the sightings missed is "
        "the wrong trade for a catch-the-sightings mission.",
    ),
    "Accuracy - High (%)": (
        "High-class accuracy: True Positives / (True Positives + False Positives + False Negatives). "
        "The Low band's true negatives are left out of the denominator entirely.",
        "This is the number to read instead of Accuracy (%) when judging the High band. Overall "
        "Accuracy is (TP + TN) / decided, and TN usually dwarfs everything else -- CDC on a "
        "typical day is 11 TP and 186 TN, so Accuracy reads 95% while the High class is 11 of 20. "
        "Accuracy - High on that same row is 55%. Precision can also look perfect here (every "
        "High call was right) while half the actual positives were labeled Low; this metric "
        "penalizes both kinds of High error.",
    ),
    "Negative Precision - Low (%)": (
        "Of indicators labeled Low confidence, the percent that stayed unobserved within the horizon.",
        "Shows the Low band is trustworthy for deprioritization. Low values mean we are wrongly dismissing real activity.",
    ),
    "Accuracy (%)": (
        "Overall correct rate on decided cases (High + Low bands only; Possibly active is excluded).",
        "Do not use this to judge High-band performance. The Low band is normally four of every "
        "five decided pairs, so a wall of true negatives will hold this near 95% even when High "
        "is catching barely half of the recurrences. Use Accuracy - High (%) for that.",
    ),
    "Specificity - True Negative Rate (%)": (
        "Of the indicators that did NOT recur, the percent correctly labeled Low confidence. "
        "Denominator = actual negatives among decided pairs (True Negatives + False Positives). "
        "Not the same as Negative Precision, which divides by Low calls made rather than by "
        "negatives that existed.",
        "The negative-class counterpart to Recall. Together the two show whether the model is "
        "equally good at both jobs; a large gap means one class is carrying the headline Accuracy.",
    ),
    "Balanced Accuracy (%)": (
        "The average of Recall - High and Specificity, giving the observed and unobserved classes "
        "equal weight regardless of how many of each there are. Scoped to decided pairs, like Accuracy.",
        "Accuracy is dominated by whichever band is bigger, and here that is almost always Low -- "
        "roughly four of every five decided pairs at the 1-day horizon. A model that simply said "
        "'Low' constantly would still post a high Accuracy. Balanced Accuracy removes that free "
        "credit, so a gap between it and Accuracy is the size of the class-imbalance flattery.",
    ),
    "Recall - High (%)": (
        "Of observed indicators that received a firm High or Low label (Possibly active excluded), "
        "the percent correctly labeled Highly likely. Denominator = High TP + positives wrongly labeled Low.",
        "Does not count positives left in Possibly active. That is why it looks healthier than "
        "the mission metric. Use Recall - High vs All Positives (%) to judge missed sightings.",
    ),
    "Recall - High vs All Positives (%)": (
        "Of every indicator that was actually observed -- High, Low, and Possibly active -- "
        "the percent labeled Highly likely on this horizon. Denominator = all actual positives.",
        "This is the 1-day High / tomorrow-page target. Catch-the-sightings means don't miss "
        "real sightings, and this is the only 1-day High column that counts abstained recurrences "
        "as the misses they are. Precision is the floor; this is the number to raise. Skip-day "
        "regulars that 1-day High cannot call at 90% precision live in Possibly Active and are "
        "already counted on the 7-day High coverage columns -- do not lower 1-day High to chase them.",
    ),
    "Recall - 7-Day High vs 1-Day Positives (%)": (
        "Of indicators actually observed the next day, the percent that were already Highly likely "
        "on the 7-day horizon of the same forecast. 1-day eval only; N/A on other horizons.",
        "This is the skip-day / weekly-board metric. Most 1-day Possibly Active recurrences are "
        "already High for the week -- the model believes they show, just not tomorrow specifically. "
        "Judge that leftover here, not by promoting those rows into 1-day High.",
    ),
    "Recall - 1-Day or 7-Day High vs All Positives (%)": (
        "Of next-day recurrences, the percent that were Highly likely on 1-day High OR 7-day High. "
        "1-day eval only; N/A on other horizons.",
        "On-the-board coverage: metronomes caught tomorrow plus skip-day regulars already flagged "
        "for the week. This is not a license to merge the bands -- 1-day High stays the trusted "
        "tomorrow page.",
    ),
    "Possibly Active Observed already 7-Day High": (
        "Count of 1-day Possibly Active rows that were actually observed the next day and were "
        "already Highly likely on 7-day. 1-day eval only; N/A on other horizons.",
        "The leftover two-thirds that a 1-day High cut cannot take without breaking 90% precision. "
        "They are already on the weekly board. Do not promote them into 1-day High.",
    ),
    "F1 Score - High (%)": (
        "Balance of Precision - High and Recall - High (decided-only recall; harmonic mean).",
        "Not the catch-the-sightings score -- it uses decided-only recall and ignores Possibly "
        "active misses. Read Recall - High vs All Positives (%) instead.",
    ),
    "Coverage (%)": (
        "Percent of scored pairs placed in High or Low (not Possibly active).",
        "The complement is the Possibly Active abstain band. Low coverage with a high Ended-High "
        "rate means catchable-but-unflagged sightings sat between High and Low.",
    ),
    "Possibly Active Ended High Rate (%)": (
        "Of Possibly active predictions, the percent that were later observed within the horizon.",
        "Leftover 1-day recurrences the High band did not take. Judge skip-day coverage on "
        "7-day High, not by promoting these rows into 1-day High.",
    ),
    "Avg Prob - Possibly Active (%)": (
        "Average model probability for all Possibly active predictions.",
        "Shows where the middle band sits between Low and High thresholds (typically ~20–80%).",
    ),
    "Avg Prob - Possibly Active Ended High (%)": (
        "Average probability among Possibly active cases that later became observed.",
        "If much higher than Ended Low, probability ranking inside the gray zone is informative.",
    ),
    "Avg Prob - Possibly Active Ended Low (%)": (
        "Average probability among Possibly active cases that stayed unobserved.",
        "Compare with Ended High average to see if mid-band probabilities separate outcomes.",
    ),
    "Unique Indicators Scored": (
        "Count of distinct indicators scored on the forecast date (100-day lookback set).",
        "Volume context for other rates; large swings can change precision/recall even if the model quality is stable.",
    ),
    "Scored Pairs (Indicator-OpDiv)": (
        "Count of Indicator–OpDiv pairs scored (one indicator can appear under multiple OpDivs).",
        "Denominator for most rates. Use this—not unique indicators—when interpreting pair-level precision/recall.",
    ),
    "Decided (High + Low)": (
        "Count of pairs in Highly likely or Low confidence (abstain / Possibly active excluded).",
        "Sample size behind Accuracy, Precision, and Negative Precision. Small n makes daily swings less reliable.",
    ),
    "Undecided (Possibly Active)": (
        "Count of pairs left in the middle / abstain band.",
        "Workload and uncertainty signal: how much of the scored set is neither a firm High nor Low call.",
    ),
    "Excluded OpDivs (Incomplete Labels)": (
        "OpDivs dropped from this row because their observation feed was missing, empty, or "
        "not yet settled somewhere in the label window.",
        "A missing label is not a negative label. If an OpDiv's feed is down, its indicators look "
        "unobserved and every High call scores as a false positive — which is a data outage, not a "
        "model error. Excluding them keeps the metrics honest; a non-empty value here means this "
        "row does not cover all OpDivs.",
    ),
    "Excluded Pairs (Incomplete Labels)": (
        "Number of Indicator-OpDiv pairs removed from scoring by the exclusions above.",
        "Size of the blind spot. Large values mean the row is based on a materially smaller "
        "population than usual, so compare it to other days with care.",
    ),
    "Data Status": (
        "Whether this OpDiv was scored on this date, and if not, why not.",
        "'Scored' means the metrics on the row are real. Anything else names the days that "
        "blocked scoring and what was wrong with each, so you can tell a feed outage from a "
        f"day that is merely too recent to trust. '{PROVISIONAL_NOTE}' means the OpDiv had "
        "already reached its usual volume for that weekday, so it was scored rather than held "
        "back for the sake of a slower feed. Such a row can only understate the model -- "
        "observations arriving afterwards turn counted misses into hits, never the reverse -- "
        "so it is never alerted on, and it is rewritten with final numbers once the day "
        "settles. A trailing "
        f"'({BACKFILLED_NOTE})' means the scheduled job missed that morning and the forecast "
        "was rebuilt later from data available as of that date. Such a row is honest about "
        "the model but slightly flattering: the rebuild read observation files that had since "
        "settled, which the live run would not have had. Treat it as indicative rather than "
        "as a like-for-like comparison against live days.",
    ),
    MISSING_METRIC: (
        "Placeholder meaning the ground truth for this window is genuinely absent -- the "
        "observation feed was down or the file never arrived.",
        "A missing label is not a negative label. Scored through, an outage makes every "
        "recurring indicator look unobserved and turns correct High calls into false positives, "
        "blaming the model for a data problem. This does not fix itself: the upstream data has "
        "to be re-pulled, then the row corrected with a backfill.",
    ),
    NOT_SETTLED: (
        "Placeholder meaning the observations exist but the upstream job is still rewriting "
        "the file for these days.",
        "Distinct from missing data -- nothing is lost, it is only early. Each day's file keeps "
        f"filling for about {OBS_SETTLE_DAYS} mornings, so scoring it now would count "
        "late-arriving observations as misses and understate precision by roughly 5 percentage "
        "points. These rows correct themselves once the day ages past the settle window; no "
        "action needed.",
    ),
    NOT_APPLICABLE: (
        "Placeholder meaning the day was scored fine, but this particular metric has no "
        "population to measure.",
        "Either the denominator is empty (no High predictions makes precision 0/0), or one "
        "class is absent (no observed positives makes Negative Precision and Accuracy read 100% "
        "purely because there was nothing to miss). That second case is why it is blanked rather "
        "than shown: as a number it renders green, a perfect score for a day that never tested "
        "the model. The counts on the row show which situation applies.",
    ),
}

FILL_GREEN = PatternFill(fill_type="solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill(fill_type="solid", fgColor="FFEB9C")
FILL_RED = PatternFill(fill_type="solid", fgColor="FFC7CE")
FONT_GREEN = Font(color="006100")
FONT_YELLOW = Font(color="9C5700")
FONT_RED = Font(color="9C0006")
FILL_HEADER = PatternFill(fill_type="solid", fgColor="D9E2F3")
FONT_HEADER = Font(bold=True)

# ---- Human-readable column definitions for the performance workbook ----
PERF_COLUMNS = [
    "Evaluation Date",
    "Forecast Date",
    "Horizon (Days)",
    "Data Status",
    "Unique Indicators Scored",
    "Scored Pairs (Indicator-OpDiv)",
    "Decided (High + Low)",
    "Undecided (Possibly Active)",
    "Coverage (%)",
    "Actual Positives",
    "Actual Negatives",
    "Observed Positive Rate (%)",
    # HIGH band
    "Predicted High Count",
    "True Positives (High)",
    "False Positives (High)",
    "False Negatives (High)",
    "Precision - High (%)",
    "Recall - High vs All Positives (%)",
    "Recall - 7-Day High vs 1-Day Positives (%)",
    "Recall - 1-Day or 7-Day High vs All Positives (%)",
    "Accuracy - High (%)",
    "Recall - High (%)",
    "Predicted High Rate (%)",
    # LOW band
    "Predicted Low Count",
    "True Negatives (Low)",
    "False Positives (Low)",
    "Negative Precision - Low (%)",
    "Predicted Low Rate (%)",
    # Possibly active band outcomes
    "Possibly Active Count",
    "Possibly Active Ended High (Observed)",
    "Possibly Active Ended Low (Not Observed)",
    "Possibly Active Ended High Rate (%)",
    "Avg Prob - Possibly Active (%)",
    "Avg Prob - Possibly Active Ended High (%)",
    "Avg Prob - Possibly Active Ended Low (%)",
    "Possibly Active Observed already 7-Day High",
    # Overall
    "Accuracy (%)",
    "Specificity - True Negative Rate (%)",
    "Balanced Accuracy (%)",
    "F1 Score - High (%)",
    # Label-window health
    "Excluded OpDivs (Incomplete Labels)",
    "Excluded Pairs (Incomplete Labels)",
]

PERF_OPDIV_COLUMNS = [
    "Evaluation Date",
    "Forecast Date",
    "Horizon (Days)",
    "OpDiv",
    "Data Status",
    "Unique Indicators Scored",
    "Scored Pairs (Indicator-OpDiv)",
    "Decided (High + Low)",
    "Undecided (Possibly Active)",
    "Coverage (%)",
    "Actual Positives",
    "Actual Negatives",
    "Observed Positive Rate (%)",
    # HIGH band
    "Predicted High Count",
    "True Positives (High)",
    "False Positives (High)",
    "False Negatives (High)",
    "Precision - High (%)",
    "Recall - High vs All Positives (%)",
    "Recall - 7-Day High vs 1-Day Positives (%)",
    "Recall - 1-Day or 7-Day High vs All Positives (%)",
    "Accuracy - High (%)",
    "Recall - High (%)",
    "Predicted High Rate (%)",
    # LOW band
    "Predicted Low Count",
    "True Negatives (Low)",
    "False Positives (Low)",
    "Negative Precision - Low (%)",
    "Predicted Low Rate (%)",
    # Possibly active band outcomes
    "Possibly Active Count",
    "Possibly Active Ended High (Observed)",
    "Possibly Active Ended Low (Not Observed)",
    "Possibly Active Ended High Rate (%)",
    "Avg Prob - Possibly Active (%)",
    "Avg Prob - Possibly Active Ended High (%)",
    "Avg Prob - Possibly Active Ended Low (%)",
    "Possibly Active Observed already 7-Day High",
    # Overall
    "Accuracy (%)",
    "Specificity - True Negative Rate (%)",
    "Balanced Accuracy (%)",
    "F1 Score - High (%)",
]

_HORIZON_FILE_LABELS = {
    1: "1day", 7: "7day", 14: "14day", 30: "30day", 45: "45day",
}


def _pct_or_missing(value: float, defined: bool):
    """Percentage float when the metric is meaningful, else the sentinel.

    Used on rows that were scored, so an undefined metric here means the
    population is empty -- not that the data is missing.
    """
    if not defined:
        return NOT_APPLICABLE
    return _pct_f(value)


def _pct(v: float) -> str:
    """Format a 0-1 ratio as a percentage string like '91.25%', or '' if NaN."""
    if pd.isna(v):
        return ""
    return f"{v * 100:.2f}%"


def _pct_f(v: float) -> float:
    """Round a 0-1 ratio to a percentage float like 91.25, or NaN."""
    if pd.isna(v):
        return float("nan")
    return round(v * 100, 2)


def _parse_prob_pct(series: pd.Series) -> pd.Series:
    """Parse '29.57%' / '29.57' / 0.2957 into a 0-100 percentage float series."""
    if series is None or series.empty:
        return pd.Series(dtype=float)
    s = series.astype(str).str.strip().str.replace("%", "", regex=False)
    vals = pd.to_numeric(s, errors="coerce")
    # If values look like 0-1 probabilities, convert to percent.
    if vals.notna().any() and float(vals.max(skipna=True)) <= 1.0:
        vals = vals * 100.0
    return vals


def _mean_prob(series: pd.Series) -> float:
    if series is None or len(series) == 0:
        return float("nan")
    if pd.api.types.is_numeric_dtype(series):
        vals = pd.to_numeric(series, errors="coerce")
    else:
        vals = _parse_prob_pct(series)
    if vals.notna().sum() == 0:
        return float("nan")
    return round(float(vals.mean()), 2)


def _week_horizon_coverage(df: pd.DataFrame, horizon_days: int) -> dict:
    """7-day High coverage of this horizon's actual positives.

    1-day High is the tomorrow page. 7-day High is the weekly board. Skip-day
    regulars sit in Possibly Active on 1-day while already High for the week.
    Only defined for 1-day eval; other horizons return N/A.
    """
    na = NOT_APPLICABLE
    out = {
        "recall_7d": na,
        "recall_union": na,
        "pa_obs_7d_high": na,
        "_raw_recall_7d": float("nan"),
        "_raw_recall_union": float("nan"),
    }
    if int(horizon_days) != 1 or df is None or df.empty or "band_7" not in df.columns:
        return out
    if df["band_7"].eq("").all():
        return out
    pos = df["Observed"] == 1
    n_pos = int(pos.sum())
    if n_pos == 0:
        out["pa_obs_7d_high"] = 0
        return out
    high7 = df["band_7"].eq("Highly likely")
    high1 = df["band_h"].eq("Highly likely")
    rec7 = float((high7 & pos).sum() / n_pos)
    rec_u = float(((high1 | high7) & pos).sum() / n_pos)
    pa_hi7 = int((df["band_h"].eq("Possibly active") & pos & high7).sum())
    out["recall_7d"] = _pct_f(rec7)
    out["recall_union"] = _pct_f(rec_u)
    out["pa_obs_7d_high"] = pa_hi7
    out["_raw_recall_7d"] = rec7
    out["_raw_recall_union"] = rec_u
    return out


def _build_metrics_row(
    *,
    eval_date_str: str,
    forecast_date_str: str,
    horizon_days: int,
    unique_indicators: int,
    total_pairs: int,
    high_n: int,
    high_tp: int,
    high_fp: int,
    low_n: int,
    low_tn: int,
    low_fp: int,
    actual_pos: int,
    actual_neg: int,
    poss_n: int = 0,
    poss_ended_high: int = 0,
    poss_ended_low: int = 0,
    avg_prob_poss: float = float("nan"),
    avg_prob_poss_high: float = float("nan"),
    avg_prob_poss_low: float = float("nan"),
    week_cov: dict | None = None,
    opdiv: str | None = None,
) -> dict:
    """Build one overall or OpDiv performance row with grouped HIGH/LOW/Possibly metrics."""
    decided = int(high_n + low_n)
    undecided = int(poss_n if poss_n else max(total_pairs - decided, 0))
    high_fn = int(low_fp)  # positives missed because labeled Low confidence
    # Standard FN vs all positives also includes Possibly active that became observed.
    high_fn_all = int(low_fp + poss_ended_high)

    precision = (high_tp / high_n) if high_n else float("nan")
    recall = (high_tp / (high_tp + high_fn)) if (high_tp + high_fn) else float("nan")
    recall_all = (high_tp / actual_pos) if actual_pos else float("nan")
    # High-class accuracy excludes Low-band true negatives. Denominator is every
    # decided pair that was either called High or actually positive -- TP, FP, FN.
    # CDC 11 TP / 9 FN / 0 FP is 11/20 = 55%, not the 95% overall Accuracy.
    high_union = int(high_tp + high_fp + high_fn)
    accuracy_high = (high_tp / high_union) if high_union else float("nan")
    neg_prec = (low_tn / low_n) if low_n else float("nan")
    accuracy = ((high_tp + low_tn) / decided) if decided else float("nan")
    # Accuracy is dominated by whichever band is larger, and the Low band usually
    # is. Specificity and balanced accuracy re-weight the two classes equally so a
    # big, easy negative class cannot carry the headline number on its own. Both
    # are scoped to decided pairs, like Accuracy itself.
    decided_neg = int(low_tn + high_fp)
    specificity = (low_tn / decided_neg) if decided_neg else float("nan")
    balanced_acc = (
        (recall + specificity) / 2
        if pd.notna(recall) and pd.notna(specificity)
        else float("nan")
    )
    f1 = (
        (2 * precision * recall) / (precision + recall)
        if pd.notna(precision) and pd.notna(recall) and (precision + recall) > 0
        else float("nan")
    )
    coverage = (decided / total_pairs) if total_pairs else float("nan")
    pred_high_rate = (high_n / total_pairs) if total_pairs else float("nan")
    obs_pos_rate = (actual_pos / total_pairs) if total_pairs else float("nan")
    pred_low_rate = (low_n / total_pairs) if total_pairs else float("nan")
    poss_high_rate = (poss_ended_high / undecided) if undecided else float("nan")

    # A rate needs a non-empty denominator to be defined, and both classes to
    # be present to mean anything. Negative Precision and Accuracy compute
    # cleanly when actual_pos == 0 -- they just report that every negative call
    # was right, which is guaranteed when there is nothing positive to miss.
    has_pairs = total_pairs > 0
    has_pos = actual_pos > 0
    has_high = high_n > 0
    has_low = low_n > 0
    has_decided = decided > 0
    has_decided_pos = (high_tp + high_fn) > 0

    row = {
        "Evaluation Date": eval_date_str,
        "Forecast Date": forecast_date_str,
        "Horizon (Days)": horizon_days,
        "Unique Indicators Scored": unique_indicators,
        "Scored Pairs (Indicator-OpDiv)": total_pairs,
        "Decided (High + Low)": decided,
        "Undecided (Possibly Active)": undecided,
        "Coverage (%)": _pct_or_missing(coverage, has_pairs),
        "Actual Positives": actual_pos,
        "Actual Negatives": actual_neg,
        "Observed Positive Rate (%)": _pct_or_missing(obs_pos_rate, has_pairs),
        "Predicted High Count": high_n,
        "True Positives (High)": high_tp,
        "False Positives (High)": high_fp,
        "False Negatives (High)": high_fn,
        "Precision - High (%)": _pct_or_missing(precision, has_high),
        "Accuracy - High (%)": _pct_or_missing(accuracy_high, high_union > 0),
        "Recall - High (%)": _pct_or_missing(recall, has_decided_pos),
        "Recall - High vs All Positives (%)": _pct_or_missing(recall_all, has_pos),
        "Recall - 7-Day High vs 1-Day Positives (%)": (week_cov or {}).get(
            "recall_7d", NOT_APPLICABLE
        ),
        "Recall - 1-Day or 7-Day High vs All Positives (%)": (week_cov or {}).get(
            "recall_union", NOT_APPLICABLE
        ),
        "Predicted High Rate (%)": _pct_or_missing(pred_high_rate, has_pairs),
        "Predicted Low Count": low_n,
        "True Negatives (Low)": low_tn,
        "False Positives (Low)": low_fp,
        "Negative Precision - Low (%)": _pct_or_missing(neg_prec, has_low and has_pos),
        "Predicted Low Rate (%)": _pct_or_missing(pred_low_rate, has_pairs),
        "Possibly Active Count": undecided,
        "Possibly Active Ended High (Observed)": int(poss_ended_high),
        "Possibly Active Ended Low (Not Observed)": int(poss_ended_low),
        "Possibly Active Ended High Rate (%)": _pct_or_missing(poss_high_rate, undecided > 0),
        "Avg Prob - Possibly Active (%)": avg_prob_poss,
        "Avg Prob - Possibly Active Ended High (%)": avg_prob_poss_high,
        "Avg Prob - Possibly Active Ended Low (%)": avg_prob_poss_low,
        "Possibly Active Observed already 7-Day High": (week_cov or {}).get(
            "pa_obs_7d_high", NOT_APPLICABLE
        ),
        "Accuracy (%)": _pct_or_missing(accuracy, has_decided and has_pos),
        "Specificity - True Negative Rate (%)": _pct_or_missing(
            specificity, decided_neg > 0
        ),
        "Balanced Accuracy (%)": _pct_or_missing(
            balanced_acc, has_decided_pos and decided_neg > 0
        ),
        "F1 Score - High (%)": _pct_or_missing(f1, has_high and has_decided_pos),
        # _raw_* stay numeric NaN so alert thresholds and rolling baselines
        # skip them instead of averaging in a vacuous 100%.
        "_raw_precision": precision if has_high else float("nan"),
        "_raw_accuracy_high": accuracy_high if high_union > 0 else float("nan"),
        "_raw_recall": recall if has_decided_pos else float("nan"),
        "_raw_recall_all": recall_all if has_pos else float("nan"),
        "_raw_recall_7d": (week_cov or {}).get("_raw_recall_7d", float("nan")),
        "_raw_recall_union": (week_cov or {}).get("_raw_recall_union", float("nan")),
        "_raw_neg_prec": neg_prec if (has_low and has_pos) else float("nan"),
        "_raw_accuracy": accuracy if (has_decided and has_pos) else float("nan"),
        "_raw_balanced_acc": (
            balanced_acc if (has_decided_pos and decided_neg > 0) else float("nan")
        ),
        "_raw_fn_all": high_fn_all,
    }
    row["Data Status"] = "Scored"
    if opdiv is not None:
        row["OpDiv"] = opdiv
    return row


_RAW_COLS = ["_raw_precision", "_raw_accuracy_high", "_raw_recall", "_raw_recall_all",
             "_raw_recall_7d", "_raw_recall_union",
             "_raw_neg_prec", "_raw_accuracy", "_raw_balanced_acc", "_raw_fn_all"]


def _sentinel_for_reasons(reasons) -> str:
    """Which placeholder describes why a window could not be scored.

    A window blocked only by days that are still being rewritten is a timing
    problem and will resolve on its own, so calling it missing data would be
    wrong -- the observations are sitting right there in the file. Anything
    else in the mix means at least one day's ground truth is genuinely absent,
    and that is the more serious of the two, so it wins.
    """
    statuses = feed_health.statuses_in(reasons)
    if statuses and statuses <= {feed_health.UNSETTLED}:
        return NOT_SETTLED
    return MISSING_METRIC


# The only counts that are honestly zero on an unscored row are the two that
# describe what the evaluation itself did. Every other count is a claim we
# cannot support: "Actual Positives = 0" asserts nothing recurred that day when
# the truth is we never looked, and "Predicted High Count = 0" is simply false
# -- the model did make those calls, they just could not be graded. A zero in
# those cells reads as a finding rather than an absence, which is the same
# mistake as printing a green 100% Accuracy.
_ZERO_ON_UNSCORED = frozenset({
    "Unique Indicators Scored",
    "Scored Pairs (Indicator-OpDiv)",
})


def _build_unscorable_row(columns: list[str], base: dict, sentinel: str) -> dict:
    """Row for a population that could not be scored at all.

    Writing the row rather than skipping it is the point: rows are upserted on
    (Evaluation Date, Forecast Date, Horizon[, OpDiv]), so a day that produces
    no row leaves whatever was written for it earlier sitting in the sheet as
    stale numbers that read like a real measurement.
    """
    row = dict(base)
    for col in columns:
        if col not in row:
            row[col] = 0 if col in _ZERO_ON_UNSCORED else sentinel
    for raw in _RAW_COLS:
        row.setdefault(raw, float("nan"))
    return row


def _build_excluded_opdiv_row(
    *,
    eval_date_str: str,
    forecast_date_str: str,
    horizon_days: int,
    opdiv: str,
    unique_indicators: int,
    total_pairs: int,
    reasons: list[str],
) -> dict:
    return _build_unscorable_row(
        PERF_OPDIV_COLUMNS,
        {
            "Evaluation Date": eval_date_str,
            "Forecast Date": forecast_date_str,
            "Horizon (Days)": horizon_days,
            "OpDiv": opdiv,
            "Data Status": (
                f"Not scored - {total_pairs:,} pairs / {unique_indicators:,} indicators "
                f"withheld: {'; '.join(reasons[:4])}"
            ),
        },
        _sentinel_for_reasons(reasons),
    )


# ========================== data loading ==========================

def load_all_csvs_from_folders(
    root_path: str,
    today_only: bool = True,
    file_date: str | None = None,
) -> pd.DataFrame:
    all_dfs: list[pd.DataFrame] = []
    target_date = file_date or datetime.today().strftime("%Y%m%d")

    if not os.path.isdir(root_path):
        _perf_log(f"data root does not exist: {root_path}", level="WARNING")
        return pd.DataFrame()

    for dirpath, dirnames, filenames in os.walk(root_path):
        parts = set(os.path.normpath(dirpath).split(os.sep))
        if parts & EXCLUDE_FOLDERS:
            continue
        partner = os.path.basename(dirpath)
        for fname in filenames:
            m = V4_NAME.match(fname) or LEGACY_NAME.match(fname)
            if not m:
                continue
            found_date = m.group(1)
            if today_only and found_date != target_date:
                continue
            fpath = os.path.join(dirpath, fname)
            try:
                df = pd.read_csv(fpath)
                df["Partner"] = partner
                df["FileDate"] = file_date
                all_dfs.append(df)
                print(f"loaded {fpath} ({len(df)} rows)")
            except Exception as e:
                _perf_log(f"Skipping {fpath}: {e}", level="WARNING")

    if all_dfs:
        return pd.concat(all_dfs, ignore_index=True)
    print("No CSV files found for today.")
    return pd.DataFrame()


def save_daily_report(df: pd.DataFrame, save_path: str, today_str: str) -> str:
    os.makedirs(save_path, exist_ok=True)
    output_path = os.path.join(save_path, f"full_daily_report_{today_str}.csv")
    df.to_csv(output_path, index=False)
    print(f"Saved to {output_path} ({len(df)} rows)")
    return output_path


def consolidate_daily_report(date_str: str | None = None) -> str | None:
    """Build full_daily_report_{date}.csv from per-OpDiv forecast CSVs."""
    date_str = date_str or datetime.today().strftime(DATE_FMT)
    daily_search = load_all_csvs_from_folders(DATA_PATH, today_only=True, file_date=date_str)
    if daily_search.empty:
        _perf_log(f"no OpDiv forecast files found for {date_str}; skip consolidate.", level="WARNING")
        return None
    try:
        return save_daily_report(daily_search, SAVE_PATH, date_str)
    except Exception as e:
        log_perf_error(f"consolidation failed for {date_str}", e)
        return None


# ========================== observation ground truth ==========================

def _load_observations_as_df(obs_date_str: str) -> pd.DataFrame:
    obs_fp = OBS_TEMPLATE.format(date=obs_date_str)
    if not os.path.exists(obs_fp):
        _perf_log(f"observation file missing for {obs_date_str}: {obs_fp}", level="WARNING")
        return pd.DataFrame(columns=["Indicator", "Partner", "Observed"])

    p = pd.read_csv(obs_fp, usecols=["indicator", "obs_date", "OpDiv"])
    p["Indicator"] = p["indicator"].astype(str).str.strip()
    p["Partner"] = p["OpDiv"].astype(str).str.strip()
    p["date"] = pd.to_datetime(p["obs_date"], errors="coerce").dt.normalize()
    p = p[p["Indicator"].ne("") & p["Indicator"].ne("nan") & p["Partner"].ne("") & p["Partner"].ne("nan")]
    if p.empty:
        return pd.DataFrame(columns=["Indicator", "Partner", "Observed"])

    eval_date = pd.to_datetime(obs_date_str, format=DATE_FMT, errors="coerce")
    if pd.notna(eval_date):
        p = p[p["date"] == eval_date]

    p = p.drop_duplicates(["Indicator", "Partner"])[["Indicator", "Partner"]]
    p["Observed"] = 1
    return p


_FEED_HEALTH: FeedHealth | None = None


def get_feed_health() -> FeedHealth:
    """Shared feed-health model, rebuilt if the observation template changes."""
    global _FEED_HEALTH
    if _FEED_HEALTH is None or _FEED_HEALTH.obs_template != OBS_TEMPLATE:
        _FEED_HEALTH = FeedHealth.from_files(OBS_TEMPLATE)
    return _FEED_HEALTH


def _healthy_opdivs_for_window(
    start_date_exclusive,
    end_date_inclusive,
    candidate_opdivs,
    today=None,
) -> tuple[set[str], dict]:
    """OpDivs whose label window is fully trustworthy, plus exclusion reasons.

    Delegates to the shared feed-health model so the metrics and the forecast
    runner's training labels apply the same definition of usable ground truth.
    """
    fh = get_feed_health()
    if today is not None:
        fh._today = today
    return fh.healthy_opdivs(
        candidate_opdivs=candidate_opdivs,
        start_exclusive=start_date_exclusive,
        end_inclusive=end_date_inclusive,
    )


def _load_observed_union_between(start_date_exclusive, end_date_inclusive) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    cur = start_date_exclusive + timedelta(days=1)
    while cur <= end_date_inclusive:
        dstr = cur.strftime(DATE_FMT)
        obs = _load_observations_as_df(dstr)
        if not obs.empty:
            frames.append(obs[["Indicator", "Partner"]])
        cur += timedelta(days=1)

    if not frames:
        return pd.DataFrame(columns=["Indicator", "Partner", "Observed"])

    out = pd.concat(frames, ignore_index=True).drop_duplicates(["Indicator", "Partner"])
    out["Observed"] = 1
    return out


# ========================== evaluation ==========================

def _forecast_report_path(d) -> str:
    return os.path.join(SAVE_PATH, f"full_daily_report_{d.strftime(DATE_FMT)}.csv")


def _backfilled_forecast_dates() -> set[str]:
    """Forecast dates rebuilt after the fact by an as-of replay.

    A replay reads observation files as they stand now, fully settled, while the
    live run that morning would have seen the last few days still filling. The
    resulting rows are therefore mildly optimistic and are labelled rather than
    left to pass as ordinary history.
    """
    global _BACKFILL_CACHE
    if _BACKFILL_CACHE is not None:
        return _BACKFILL_CACHE
    fp = os.path.join(DATA_PATH, BACKFILL_MARKER_NAME)
    dates: set[str] = set()
    try:
        if os.path.exists(fp):
            with open(fp, encoding="utf-8") as fh:
                dates = {ln.strip() for ln in fh if ln.strip()}
    except Exception as e:
        _perf_log(f"could not read backfill marker {fp}: {e}", level="WARNING")
    _BACKFILL_CACHE = dates
    return dates


def _tag_backfilled(overall: dict, forecast_date_str: str) -> dict:
    """Note the replayed provenance on every row a replayed forecast produced."""
    if forecast_date_str not in _backfilled_forecast_dates():
        return overall
    for row in [overall, *overall.get("_opdiv_rows", [])]:
        row["Data Status"] = f"{row.get('Data Status', 'Scored')} ({BACKFILLED_NOTE})"
    return overall


def _tag_provisional(overall: dict, provisional_info: dict) -> dict:
    """Mark rows resting on days that had not formally settled when scored.

    Applied per OpDiv, because the whole point of early settlement is that one
    slow feed no longer speaks for the rest: on the same date some OpDivs are
    final and others are still provisional.
    """
    if not provisional_info:
        return overall
    for row in overall.get("_opdiv_rows", []):
        days = provisional_info.get(str(row.get("OpDiv", "")).strip())
        if days:
            row["Data Status"] = (
                f"{row.get('Data Status', 'Scored')} ({PROVISIONAL_NOTE}; "
                f"{'; '.join(days[:4])})"
            )
    overall["Data Status"] = (
        f"{overall.get('Data Status', 'Scored')} ({PROVISIONAL_NOTE} for "
        f"{', '.join(sorted(provisional_info))})"
    )
    return overall


def check_forecast_coverage(today=None) -> list[str]:
    """Alert when recent forecast days are absent, and log longer-run coverage.

    A missing forecast is invisible in the metrics: no forecast means no row,
    and a sheet with no row for a date looks identical to a date nobody has
    reached yet. The Aug 2026 case ran six days that way -- the forecast job
    exited FATAL every morning on an unreachable share path, and the gap only
    surfaced later as holes in the 7-day sheet.

    This cannot report an outage while it is happening: the forecast runner and
    this evaluation are separate scheduled tasks, but they read the same share,
    so whatever kills one kills the other. What it does catch is the recovery.
    The first run after a gap names every day that was lost, which is the
    difference between finding out on day one and finding out in a month.

    Only recent gaps alert. Older ones cannot be filled without a leakage-free
    replay, so repeating them daily would be noise that trains people to ignore
    the alert file.
    """
    today = today or datetime.today().date()
    missing = [
        today - timedelta(days=k)
        for k in range(COVERAGE_LOOKBACK_DAYS + 1)
        if not os.path.exists(_forecast_report_path(today - timedelta(days=k)))
    ]
    expected = COVERAGE_LOOKBACK_DAYS + 1
    if not missing:
        _perf_log(f"forecast coverage: all {expected} of the last {expected} days present")
        return []

    missing_str = ", ".join(d.strftime(DATE_FMT) for d in sorted(missing))
    _perf_log(
        f"forecast coverage: {expected - len(missing)}/{expected} days present; "
        f"missing {missing_str}",
        level="WARNING",
    )

    recent = [d for d in missing if (today - d).days < COVERAGE_ALERT_DAYS]
    if not recent:
        return []
    return [
        f"PIPELINE_ALERT: no forecast written for "
        f"{', '.join(d.strftime(DATE_FMT) for d in sorted(recent))} "
        f"(checked {SAVE_PATH}). The forecast job did not produce output -- check its run "
        f"log for a FATAL exit. Days older than {COVERAGE_ALERT_DAYS} are reported in the "
        f"performance log but not alerted; total gap over the last {expected} days: "
        f"{missing_str}"
    ]


def _evaluate_horizon(horizon_days: int, eval_date) -> dict | None:
    forecast_date = eval_date - timedelta(days=horizon_days)
    forecast_date_str = forecast_date.strftime(DATE_FMT)
    eval_date_str = eval_date.strftime(DATE_FMT)
    forecast_fp = _forecast_report_path(forecast_date)
    if not os.path.exists(forecast_fp):
        _perf_log(f"forecast daily report missing: {forecast_fp}", level="WARNING")
        return None

    df_pred = pd.read_csv(forecast_fp)

    conf_col = f"Confidence: {horizon_days}-Day"
    prob_col = f"Probability: {horizon_days}-Day"
    required_cols = {"Indicator", "Partner", conf_col}
    missing = required_cols - set(df_pred.columns)
    if missing:
        raise RuntimeError(f"PERF: missing columns in {forecast_fp}: {sorted(missing)}")

    df_obs = _load_observed_union_between(
        start_date_exclusive=forecast_date,
        end_date_inclusive=eval_date,
    )
    if df_obs.empty:
        _perf_log(
            "no observation ground-truth in window "
            f"({forecast_date_str}, {eval_date_str}] for H={horizon_days}; skipping eval.",
            level="WARNING",
        )
        return None

    df_pred["Indicator"] = df_pred["Indicator"].astype(str).str.strip()
    df_pred["Partner"] = df_pred["Partner"].astype(str).str.strip()

    healthy_opdivs, mask_info = _healthy_opdivs_for_window(
        start_date_exclusive=forecast_date,
        end_date_inclusive=eval_date,
        candidate_opdivs=df_pred["Partner"].unique(),
    )
    excluded_info = mask_info["excluded"]
    provisional_info = mask_info.get("provisional", {})
    if excluded_info:
        for opdiv in sorted(excluded_info):
            _perf_log(
                f"H={horizon_days} eval={eval_date_str}: excluding {opdiv} -- "
                f"incomplete label window ({'; '.join(excluded_info[opdiv][:4])})",
                level="WARNING",
            )
    if provisional_info:
        _perf_log(
            f"H={horizon_days} eval={eval_date_str}: scoring "
            f"{', '.join(sorted(provisional_info))} on days that have delivered but "
            f"not formally settled; rows are marked provisional with their "
            f"completeness and will be rewritten once those days settle"
        )

    excluded_mask = df_pred["Partner"].isin(excluded_info.keys())
    excluded_pairs = int(excluded_mask.sum())
    excluded_pop = (
        df_pred[excluded_mask]
        .groupby("Partner")
        .agg(pairs=("Indicator", "size"), uniq=("Indicator", "nunique"))
        .to_dict("index")
    )
    df_pred = df_pred[df_pred["Partner"].isin(healthy_opdivs)]
    if df_pred.empty:
        _perf_log(
            f"H={horizon_days} eval={eval_date_str}: every OpDiv has an incomplete "
            f"label window over ({forecast_date_str}, {eval_date_str}]; nothing scorable. "
            f"unsettled={len(mask_info['unsettled_days'])} missing={len(mask_info['missing_days'])}",
            level="WARNING",
        )
        all_reasons = [r for rs in excluded_info.values() for r in rs]
        overall = _build_unscorable_row(
            PERF_COLUMNS,
            {
                "Evaluation Date": eval_date_str,
                "Forecast Date": forecast_date_str,
                "Horizon (Days)": horizon_days,
                "Data Status": (
                    f"Not scored - no OpDiv has a complete label window over "
                    f"({forecast_date_str}, {eval_date_str}]; "
                    f"blocked by: {', '.join(sorted(feed_health.statuses_in(all_reasons)))}"
                ),
                "Excluded OpDivs (Incomplete Labels)": ", ".join(sorted(excluded_info)),
                "Excluded Pairs (Incomplete Labels)": excluded_pairs,
            },
            _sentinel_for_reasons(all_reasons),
        )
        overall["_opdiv_rows"] = [
            _build_excluded_opdiv_row(
                eval_date_str=eval_date_str,
                forecast_date_str=forecast_date_str,
                horizon_days=horizon_days,
                opdiv=opdiv,
                unique_indicators=int(excluded_pop.get(opdiv, {}).get("uniq", 0)),
                total_pairs=int(excluded_pop.get(opdiv, {}).get("pairs", 0)),
                reasons=reasons,
            )
            for opdiv, reasons in sorted(excluded_info.items())
        ]
        return _tag_backfilled(overall, forecast_date_str)

    df_scored = df_pred.merge(df_obs, on=["Indicator", "Partner"], how="left")
    df_scored["Observed"] = df_scored["Observed"].fillna(0).astype(int)

    tag_prefix = f"{horizon_days}-Day:"
    df_scored["band_h"] = (
        df_scored[conf_col]
        .astype(str)
        .str.replace(tag_prefix, "", regex=False)
        .str.strip()
    )
    conf7 = "Confidence: 7-Day"
    if conf7 in df_scored.columns:
        df_scored["band_7"] = (
            df_scored[conf7]
            .astype(str)
            .str.replace("7-Day:", "", regex=False)
            .str.strip()
        )
    else:
        df_scored["band_7"] = ""
        if horizon_days == 1:
            _perf_log(
                "7-Day confidence column missing; skip-day coverage columns will be N/A",
                level="WARNING",
            )
    if prob_col in df_scored.columns:
        df_scored["_prob_pct"] = _parse_prob_pct(df_scored[prob_col])
    else:
        df_scored["_prob_pct"] = float("nan")
        _perf_log(f"probability column missing: {prob_col}; avg probs will be blank", level="WARNING")

    total = int(len(df_scored))
    unique_indicators = int(df_scored["Indicator"].nunique())

    high = df_scored[df_scored["band_h"].eq("Highly likely")]
    high_n = len(high)
    high_tp = int(high["Observed"].sum()) if high_n else 0
    high_fp = int((high["Observed"] == 0).sum()) if high_n else 0

    low = df_scored[df_scored["band_h"].eq("Low confidence")]
    low_n = len(low)
    low_tn = int((low["Observed"] == 0).sum()) if low_n else 0
    low_fp = int((low["Observed"] == 1).sum()) if low_n else 0

    poss = df_scored[df_scored["band_h"].eq("Possibly active")]
    poss_n = len(poss)
    poss_ended_high = int(poss["Observed"].sum()) if poss_n else 0
    poss_ended_low = int((poss["Observed"] == 0).sum()) if poss_n else 0
    avg_prob_poss = _mean_prob(poss["_prob_pct"]) if poss_n else float("nan")
    avg_prob_poss_high = (
        _mean_prob(poss.loc[poss["Observed"] == 1, "_prob_pct"]) if poss_ended_high else float("nan")
    )
    avg_prob_poss_low = (
        _mean_prob(poss.loc[poss["Observed"] == 0, "_prob_pct"]) if poss_ended_low else float("nan")
    )

    actual_pos = int(df_scored["Observed"].sum())
    actual_neg = int(total - actual_pos)

    def _poss_stats(g: pd.DataFrame) -> tuple[int, int, int, float, float, float]:
        g_poss = g[g["band_h"].eq("Possibly active")]
        n = len(g_poss)
        ended_high = int(g_poss["Observed"].sum()) if n else 0
        ended_low = int((g_poss["Observed"] == 0).sum()) if n else 0
        avg_all = _mean_prob(g_poss["_prob_pct"]) if n else float("nan")
        avg_high = (
            _mean_prob(g_poss.loc[g_poss["Observed"] == 1, "_prob_pct"]) if ended_high else float("nan")
        )
        avg_low = (
            _mean_prob(g_poss.loc[g_poss["Observed"] == 0, "_prob_pct"]) if ended_low else float("nan")
        )
        return n, ended_high, ended_low, avg_all, avg_high, avg_low

    # Per-OpDiv (Partner) breakdown
    opdiv_rows: list[dict] = []
    for opdiv, g in df_scored.groupby("Partner"):
        g_total = int(len(g))
        g_high = g[g["band_h"].eq("Highly likely")]
        g_high_n = len(g_high)
        g_high_tp = int(g_high["Observed"].sum()) if g_high_n else 0
        g_high_fp = int((g_high["Observed"] == 0).sum()) if g_high_n else 0

        g_low = g[g["band_h"].eq("Low confidence")]
        g_low_n = len(g_low)
        g_low_tn = int((g_low["Observed"] == 0).sum()) if g_low_n else 0
        g_low_fp = int((g_low["Observed"] == 1).sum()) if g_low_n else 0

        g_poss_n, g_poss_hi, g_poss_lo, g_avg, g_avg_hi, g_avg_lo = _poss_stats(g)

        opdiv_rows.append(
            _build_metrics_row(
                eval_date_str=eval_date_str,
                forecast_date_str=forecast_date_str,
                horizon_days=horizon_days,
                unique_indicators=int(g["Indicator"].nunique()),
                total_pairs=g_total,
                high_n=g_high_n,
                high_tp=g_high_tp,
                high_fp=g_high_fp,
                low_n=g_low_n,
                low_tn=g_low_tn,
                low_fp=g_low_fp,
                actual_pos=int(g["Observed"].sum()),
                actual_neg=int(g_total - g["Observed"].sum()),
                poss_n=g_poss_n,
                poss_ended_high=g_poss_hi,
                poss_ended_low=g_poss_lo,
                avg_prob_poss=g_avg,
                avg_prob_poss_high=g_avg_hi,
                avg_prob_poss_low=g_avg_lo,
                week_cov=_week_horizon_coverage(g, horizon_days),
                opdiv=opdiv,
            )
        )

    for opdiv, reasons in sorted(excluded_info.items()):
        pop = excluded_pop.get(opdiv, {"pairs": 0, "uniq": 0})
        opdiv_rows.append(
            _build_excluded_opdiv_row(
                eval_date_str=eval_date_str,
                forecast_date_str=forecast_date_str,
                horizon_days=horizon_days,
                opdiv=opdiv,
                unique_indicators=int(pop["uniq"]),
                total_pairs=int(pop["pairs"]),
                reasons=reasons,
            )
        )

    overall = _build_metrics_row(
        eval_date_str=eval_date_str,
        forecast_date_str=forecast_date_str,
        horizon_days=horizon_days,
        unique_indicators=unique_indicators,
        total_pairs=total,
        high_n=high_n,
        high_tp=high_tp,
        high_fp=high_fp,
        low_n=low_n,
        low_tn=low_tn,
        low_fp=low_fp,
        actual_pos=actual_pos,
        actual_neg=actual_neg,
        poss_n=poss_n,
        poss_ended_high=poss_ended_high,
        poss_ended_low=poss_ended_low,
        avg_prob_poss=avg_prob_poss,
        avg_prob_poss_high=avg_prob_poss_high,
        avg_prob_poss_low=avg_prob_poss_low,
        week_cov=_week_horizon_coverage(df_scored, horizon_days),
    )
    # "(none)" rather than "" or "None" so the value survives a pandas
    # read_excel round-trip instead of coming back as NaN.
    overall["Excluded OpDivs (Incomplete Labels)"] = (
        ", ".join(sorted(excluded_info)) if excluded_info else "(none)"
    )
    overall["Excluded Pairs (Incomplete Labels)"] = excluded_pairs
    overall["_opdiv_rows"] = opdiv_rows
    _tag_provisional(overall, provisional_info)
    return _tag_backfilled(overall, forecast_date_str)


def _fmt_metric(row: dict, key: str) -> str:
    """Value with a % suffix, or the sentinel verbatim when not computable."""
    v = row.get(key)
    return str(v) if v in _SENTINELS else f"{v}%"


def _print_summary(row: dict) -> None:
    h = row["Horizon (Days)"]
    print(
        f"PERF SUMMARY  Horizon={h}-Day  "
        f"Eval={row['Evaluation Date']}  Forecast={row['Forecast Date']}  "
        f"UniqueIndicators={row['Unique Indicators Scored']}  "
        f"Pairs={row['Scored Pairs (Indicator-OpDiv)']}  "
        f"Decided={row['Decided (High + Low)']}  "
        f"Coverage={_fmt_metric(row, 'Coverage (%)')}"
    )
    print(
        f"  TARGET Recall vs all positives={_fmt_metric(row, 'Recall - High vs All Positives (%)')}  "
        f"Precision floor={_fmt_metric(row, 'Precision - High (%)')}  "
        f"AccuracyHigh={_fmt_metric(row, 'Accuracy - High (%)')}  "
        f"Recall(decided)={_fmt_metric(row, 'Recall - High (%)')}"
    )
    if int(row["Horizon (Days)"]) == 1:
        print(
            f"  SKIP-DAY 7-Day High vs 1-Day pos="
            f"{_fmt_metric(row, 'Recall - 7-Day High vs 1-Day Positives (%)')}  "
            f"1-Day or 7-Day High="
            f"{_fmt_metric(row, 'Recall - 1-Day or 7-Day High vs All Positives (%)')}  "
            f"PA observed already 7-Day High="
            f"{row.get('Possibly Active Observed already 7-Day High')}"
        )
    print(
        f"  Accuracy={_fmt_metric(row, 'Accuracy (%)')}  "
        f"(do not use for High-band health)  "
        f"F1={_fmt_metric(row, 'F1 Score - High (%)')}"
    )
    print(
        f"  Specificity={_fmt_metric(row, 'Specificity - True Negative Rate (%)')}  "
        f"BalancedAcc={_fmt_metric(row, 'Balanced Accuracy (%)')}"
    )
    print(
        f"  HIGH band: TP={row['True Positives (High)']}/{row['Predicted High Count']}  "
        f"FP={row['False Positives (High)']}  FN={row['False Negatives (High)']}"
    )
    print(
        f"  LOW  band: TN={row['True Negatives (Low)']}/{row['Predicted Low Count']}  "
        f"FP={row['False Positives (Low)']}  "
        f"Neg Precision={_fmt_metric(row, 'Negative Precision - Low (%)')}"
    )
    print(
        f"  POSS band: n={row['Possibly Active Count']}  "
        f"EndedHigh={row['Possibly Active Ended High (Observed)']}  "
        f"EndedLow={row['Possibly Active Ended Low (Not Observed)']}  "
        f"EndedHighRate={_fmt_metric(row, 'Possibly Active Ended High Rate (%)')}  "
        f"AvgProb={_fmt_metric(row, 'Avg Prob - Possibly Active (%)')}  "
        f"(EndedHighAvg={_fmt_metric(row, 'Avg Prob - Possibly Active Ended High (%)')}, "
        f"EndedLowAvg={_fmt_metric(row, 'Avg Prob - Possibly Active Ended Low (%)')})"
    )


# ========================== alerting ==========================

def _maybe_alert(perf_df: pd.DataFrame, row: dict, opdiv: str | None = None) -> list[str]:
    # An unscored row has no measurement to alert on, and its count cells hold
    # the sentinel rather than a number. The feed problem behind it is already
    # logged as a WARNING when the window is rejected.
    status = str(row.get("Data Status", ""))
    if status.startswith("Not scored"):
        return []
    # A provisional row can only understate the model: observations that arrive
    # after it was scored turn counted misses into hits, never the reverse. The
    # Aug 2026 precision scare was exactly that artifact, so these rows are
    # published but never alerted on. The settled rewrite alerts if it is real.
    if PROVISIONAL_NOTE in status:
        return []

    h = int(row["Horizon (Days)"])
    tag = f"{h}-Day" if not opdiv else f"{h}-Day {opdiv}"
    precision = pd.to_numeric(row.get("_raw_precision"), errors="coerce")
    recall_all = pd.to_numeric(row.get("_raw_recall_all"), errors="coerce")
    neg_prec = pd.to_numeric(row.get("_raw_neg_prec"), errors="coerce")
    decided = pd.to_numeric(row.get("Decided (High + Low)"), errors="coerce")
    low_n = pd.to_numeric(row.get("Predicted Low Count"), errors="coerce")
    actual_pos = pd.to_numeric(row.get("Actual Positives"), errors="coerce")
    if pd.isna(decided):
        return []
    decided = int(decided)
    low_n = int(low_n) if pd.notna(low_n) else 0
    actual_pos = int(actual_pos) if pd.notna(actual_pos) else 0

    if not perf_df.empty:
        rolling = perf_df.tail(ROLLING_BASELINE_DAYS)
        roll_prec = float(rolling["_raw_precision"].mean())
        roll_neg = float(rolling["_raw_neg_prec"].mean())
        roll_recall_all = (
            float(rolling["_raw_recall_all"].mean())
            if "_raw_recall_all" in rolling.columns
            else float("nan")
        )
    else:
        roll_prec = float("nan")
        roll_neg = float("nan")
        roll_recall_all = float("nan")

    alerts: list[str] = []
    recall_floor = RECALL_ALL_OPDIV_ABS_MIN if opdiv else RECALL_ALL_ABS_MIN

    if decided >= MIN_DECIDED_COUNT:
        # Precision floor: overall, and OpDivs whose High cut we lowered.
        watch_prec = opdiv is None or int(h) in BAND_HIGH_P_OPDIV.get(
            str(opdiv).strip().upper(), {}
        )
        if watch_prec and precision < HIGH_PREC_ABS_MIN:
            alerts.append(
                f"PERFORMANCE_ALERT ({tag}): Precision dropped to {_pct(precision)} "
                f"(floor: {_pct(HIGH_PREC_ABS_MIN)}, n={decided})"
            )
            if pd.notna(roll_prec) and precision < (roll_prec - HIGH_PREC_DROP_PP_ROLLING):
                alerts.append(
                    f"PERFORMANCE_ALERT ({tag}): Precision {_pct(precision)} fell below "
                    f"{ROLLING_BASELINE_DAYS}-day rolling average {_pct(roll_prec)} "
                    f"by more than {HIGH_PREC_DROP_PP_ROLLING*100:.0f}pp while under the floor "
                    f"(n={decided})"
                )
        # Target: recall against all positives, including Possibly Active misses.
        if actual_pos > 0 and pd.notna(recall_all) and recall_all < recall_floor:
            alerts.append(
                f"PERFORMANCE_ALERT ({tag}): Recall vs all positives dropped to "
                f"{_pct(recall_all)} (target minimum: {_pct(recall_floor)}, "
                f"positives={actual_pos})"
            )
        if (
            not opdiv
            and actual_pos > 0
            and pd.notna(recall_all)
            and pd.notna(roll_recall_all)
            and recall_all < (roll_recall_all - RECALL_ALL_DROP_PP_ROLLING)
        ):
            alerts.append(
                f"PERFORMANCE_ALERT ({tag}): Recall vs all {_pct(recall_all)} fell below "
                f"{ROLLING_BASELINE_DAYS}-day rolling average {_pct(roll_recall_all)} "
                f"by more than {RECALL_ALL_DROP_PP_ROLLING*100:.0f}pp (positives={actual_pos})"
            )

    if low_n >= MIN_DECIDED_COUNT:
        if neg_prec < LOW_NEG_PREC_ABS_MIN:
            alerts.append(
                f"PERFORMANCE_ALERT ({tag}): Low-band negative precision dropped to "
                f"{_pct(neg_prec)} (minimum standard: {_pct(LOW_NEG_PREC_ABS_MIN)}, n={low_n})"
            )
        if pd.notna(roll_neg) and neg_prec < (roll_neg - LOW_NEG_PREC_DROP_PP_ROLLING):
            alerts.append(
                f"PERFORMANCE_ALERT ({tag}): Low-band negative precision {_pct(neg_prec)} "
                f"fell below {ROLLING_BASELINE_DAYS}-day rolling average {_pct(roll_neg)} "
                f"by more than {LOW_NEG_PREC_DROP_PP_ROLLING*100:.0f}pp (n={low_n})"
            )

    return alerts


# ========================== persist + run loop ==========================

def _perf_file_name(horizon_days: int) -> str:
    label = _HORIZON_FILE_LABELS.get(horizon_days, f"{horizon_days}day")
    return f"performance_{label}.csv"


def _normalize_perf_history(df: pd.DataFrame) -> pd.DataFrame:
    """Map legacy column names when reloading workbook history."""
    if df.empty:
        return df
    if "Total Indicators Scored" in df.columns:
        if "Scored Pairs (Indicator-OpDiv)" not in df.columns:
            df = df.rename(columns={"Total Indicators Scored": "Scored Pairs (Indicator-OpDiv)"})
        else:
            df = df.drop(columns=["Total Indicators Scored"])
        if "Unique Indicators Scored" not in df.columns:
            df["Unique Indicators Scored"] = pd.NA
    if "False Negatives (Low)" in df.columns and "False Negatives (High)" not in df.columns:
        df = df.rename(columns={"False Negatives (Low)": "False Negatives (High)"})
    if "F1 Score (%)" in df.columns and "F1 Score - High (%)" not in df.columns:
        df = df.rename(columns={"F1 Score (%)": "F1 Score - High (%)"})
    df = _backfill_balanced_accuracy(df)
    df = _backfill_accuracy_high(df)
    return df


def _backfill_accuracy_high(df: pd.DataFrame) -> pd.DataFrame:
    """Derive Accuracy - High for rows written before the column existed.

    TP / (TP + FP + FN). Low-band true negatives are not in the denominator.
    """
    need = ["True Positives (High)", "False Positives (High)", "False Negatives (High)"]
    if any(c not in df.columns for c in need):
        return df
    tp, fp, fn = (pd.to_numeric(df[c], errors="coerce") for c in need)
    union = tp + fp + fn
    acc_h = (tp / union).where(union > 0) * 100
    col = "Accuracy - High (%)"
    rounded = acc_h.round(2)
    if col not in df.columns:
        df[col] = rounded
    else:
        existing = pd.to_numeric(df[col], errors="coerce")
        df[col] = df[col].where(existing.notna(), rounded)
    df[col] = df[col].where(df[col].notna(), NOT_APPLICABLE)
    if "_raw_accuracy_high" in df.columns:
        raw = pd.to_numeric(df["_raw_accuracy_high"], errors="coerce")
        df["_raw_accuracy_high"] = raw.where(raw.notna(), acc_h / 100)
    else:
        df["_raw_accuracy_high"] = acc_h / 100
    return df


def _backfill_balanced_accuracy(df: pd.DataFrame) -> pd.DataFrame:
    """Derive Specificity and Balanced Accuracy for rows written before they existed.

    Every input is already on the historical row as a count, so this reconstructs
    the two columns exactly rather than approximating them. Rows that were never
    scored keep their sentinel: their counts are zero, and a zero denominator here
    means the metric is undefined, not that it is zero.
    """
    need = ["True Negatives (Low)", "False Positives (High)",
            "True Positives (High)", "False Positives (Low)"]
    if any(c not in df.columns for c in need):
        return df

    tn, fp, tp, fn = (pd.to_numeric(df[c], errors="coerce") for c in need)
    neg, pos = tn + fp, tp + fn
    spec = (tn / neg).where(neg > 0) * 100
    rec = (tp / pos).where(pos > 0) * 100
    bal = (spec + rec) / 2

    for col, vals in (
        ("Specificity - True Negative Rate (%)", spec),
        ("Balanced Accuracy (%)", bal),
    ):
        rounded = vals.round(2)
        if col not in df.columns:
            df[col] = rounded
        else:
            # Only fill gaps; never overwrite a value the evaluator computed.
            existing = pd.to_numeric(df[col], errors="coerce")
            df[col] = df[col].where(existing.notna(), rounded)
        df[col] = df[col].where(df[col].notna(), NOT_APPLICABLE)

    if "_raw_balanced_acc" in df.columns:
        raw = pd.to_numeric(df["_raw_balanced_acc"], errors="coerce")
        df["_raw_balanced_acc"] = raw.where(raw.notna(), bal / 100)
    else:
        df["_raw_balanced_acc"] = bal / 100
    return df


def _traffic_light_for_value(value, green_min: float, yellow_min: float):
    """Return (fill, font) for a metric value, or None if not colorable."""
    try:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        v = float(value)
    except (TypeError, ValueError):
        return None
    if v >= green_min:
        return FILL_GREEN, FONT_GREEN
    if v >= yellow_min:
        return FILL_YELLOW, FONT_YELLOW
    return FILL_RED, FONT_RED


def _apply_perf_sheet_formatting(ws) -> None:
    """Bold header, freeze pane, traffic-light color key metric cells."""
    if ws.max_row < 1 or ws.max_column < 1:
        return

    headers = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        name = str(cell.value) if cell.value is not None else ""
        headers[name] = col
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 36

    for col_name, (green_min, yellow_min) in METRIC_TRAFFIC_LIGHTS.items():
        col_idx = headers.get(col_name)
        if not col_idx:
            continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, col_idx)
            style = _traffic_light_for_value(cell.value, green_min, yellow_min)
            if style is None:
                continue
            fill, font = style
            cell.fill = fill
            cell.font = font

    # Reasonable column widths for readability
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col).value or "")
        width = min(max(len(header) + 2, 12), 42)
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_legend_sheet(wb) -> None:
    """Add/replace a Legend sheet explaining traffic lights and each measure."""
    if "Legend" in wb.sheetnames:
        del wb["Legend"]
    ws = wb.create_sheet("Legend", 0)

    ws["A1"] = "Performance workbook legend"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        "Mission: catch real indicator sightings without missing them, while keeping High flags "
        "trusted. The target metric for the tomorrow page is Recall - High vs All Positives (%). "
        "Precision - High is the floor (green at 90%), not a number to maximize. Do not lower "
        "1-day High to chase skip-day leftovers -- those are the 7-Day High coverage columns. "
        "Accuracy (%) is dominated by Low-band true negatives and is not a High-band signal. "
        "Three jobs: 1-day High = page for tomorrow; 7-day High = already on the weekly board; "
        "Possibly active = abstain (not a worklist). Bands: Highly likely = predicted positive, "
        "Low confidence = predicted negative, Possibly active = abstain."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 96

    ws["A4"] = "Color key"
    ws["A4"].font = Font(bold=True, size=12)
    ws["A5"] = "Color"
    ws["B5"] = "Meaning"
    ws["C5"] = "Rule"
    for col in range(1, 4):
        ws.cell(5, col).fill = FILL_HEADER
        ws.cell(5, col).font = FONT_HEADER

    ws["A6"] = "Green"
    ws["A6"].fill = FILL_GREEN
    ws["A6"].font = FONT_GREEN
    ws["B6"] = "Good"
    ws["C6"] = "At or above the green threshold"

    ws["A7"] = "Yellow"
    ws["A7"].fill = FILL_YELLOW
    ws["A7"].font = FONT_YELLOW
    ws["B7"] = "Watch / middle"
    ws["C7"] = "At or above yellow, below green"

    ws["A8"] = "Red"
    ws["A8"].fill = FILL_RED
    ws["A8"].font = FONT_RED
    ws["B8"] = "Bad"
    ws["C8"] = "Below the yellow threshold"

    ws["A10"] = "Traffic-light metrics"
    ws["A10"].font = Font(bold=True, size=12)
    headers = [
        "Metric",
        "What it measures",
        "Why it matters",
        "Green >=",
        "Yellow >=",
        "Red <",
    ]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(11, col, title)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    row = 12
    for metric, (green_min, yellow_min) in METRIC_TRAFFIC_LIGHTS.items():
        what, why = METRIC_DESCRIPTIONS.get(metric, ("", ""))
        ws.cell(row, 1, metric)
        ws.cell(row, 2, what).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 3, why).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 4, green_min)
        ws.cell(row, 5, yellow_min)
        ws.cell(row, 6, yellow_min)
        ws.row_dimensions[row].height = 60
        row += 1

    row += 1
    ws.cell(row, 1, "Other key measures (not traffic-lighted)")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1
    other_headers = ["Metric", "What it measures", "Why it matters"]
    for col, title in enumerate(other_headers, start=1):
        cell = ws.cell(row, col, title)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    row += 1

    other_metrics = [
        "Unique Indicators Scored",
        "Scored Pairs (Indicator-OpDiv)",
        "Decided (High + Low)",
        "Undecided (Possibly Active)",
        "Possibly Active Ended High Rate (%)",
        "Avg Prob - Possibly Active (%)",
        "Avg Prob - Possibly Active Ended High (%)",
        "Avg Prob - Possibly Active Ended Low (%)",
        "Possibly Active Observed already 7-Day High",
        "Excluded OpDivs (Incomplete Labels)",
        "Excluded Pairs (Incomplete Labels)",
        "Data Status",
        MISSING_METRIC,
        NOT_SETTLED,
        NOT_APPLICABLE,
    ]
    for metric in other_metrics:
        what, why = METRIC_DESCRIPTIONS.get(metric, ("", ""))
        ws.cell(row, 1, metric)
        ws.cell(row, 2, what).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 3, why).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 48
        row += 1

    row += 1
    ws.cell(row, 1, "How rows get written and corrected")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1
    for note in [
        f"Each daily run scores the evaluation date {OBS_SETTLE_DAYS} days back, not today. The upstream "
        "observation feed keeps rewriting each day's file for about two to three mornings, so scoring a "
        "day before it settles counts late-arriving observations as misses and understates precision by "
        "roughly 5 percentage points.",
        "Rows are upserted on (Evaluation Date, Forecast Date, Horizon). A row therefore gets rewritten "
        f"if a later run happens to score the same date -- which is why the {OBS_SETTLE_DAYS} most recent "
        "evaluation dates self-correct as the lag window catches up to them.",
        "IMPORTANT: rows older than the lag window are never revisited automatically. If you see old "
        "outliers that never corrected themselves, they were written by an earlier run against unsettled "
        "or broken data and can only be fixed by a backfill "
        "(set NOI_V4_PERF_BACKFILL_START / NOI_V4_PERF_BACKFILL_END and rerun).",
        "An OpDiv is dropped from a row when its observation feed was missing, empty, or unsettled "
        "anywhere in the label window -- see the two Excluded columns above. A missing label is not a "
        "negative label, so scoring through a feed outage would blame the model for a data problem.",
    ]:
        ws.cell(row, 1, note).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 46
        row += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.freeze_panes = "A12"


def _format_perf_workbook(wb) -> None:
    """Apply legend + traffic-light formatting to all data sheets."""
    _write_legend_sheet(wb)
    for name in wb.sheetnames:
        if name == "Legend":
            continue
        _apply_perf_sheet_formatting(wb[name])


def _horizon_workbook_path(horizon_days: int) -> str:
    label = _HORIZON_FILE_LABELS.get(horizon_days, f"{horizon_days}day")
    return os.path.join(PERF_DIR, f"performance_{label}.xlsx")

def _safe_sheet_name(name: str) -> str:
    # Excel sheet constraints: <=31 chars, no []:*?/\
    bad = "[]:*?/\\"
    for ch in bad:
        name = name.replace(ch, "_")
    return name[:31]


def _export_horizon_workbooks() -> None:
    """
    Exports one growing workbook per horizon.
    For each horizon workbook:
      - sheet 'overall' with day-by-day aggregate metrics
      - one sheet per OpDiv with that OpDiv's day-by-day metrics
    """
    # Workbook is now written directly in _run_eval_loop.
    return

    try:
        for horizon_days in EVAL_HORIZONS:
            label = _HORIZON_FILE_LABELS.get(horizon_days, f"{horizon_days}day")
            overall_fp = os.path.join(PERF_DIR, _perf_file_name(horizon_days))
            opdiv_fp = os.path.join(PERF_DIR, f"_tmp_performance_{label}_by_opdiv.csv")

            # Only create workbook when we have at least one source file.
            if not os.path.exists(overall_fp) and not os.path.exists(opdiv_fp):
                continue

            horizon_xlsx = os.path.join(PERF_DIR, f"performance_{label}.xlsx")
            with pd.ExcelWriter(horizon_xlsx) as writer:
                if os.path.exists(overall_fp):
                    df_overall = pd.read_csv(overall_fp)
                    if "Evaluation Date" in df_overall.columns:
                        df_overall = df_overall.sort_values("Evaluation Date")
                    df_overall.to_excel(writer, sheet_name="overall", index=False)

                if os.path.exists(opdiv_fp):
                    df_opdiv = pd.read_csv(opdiv_fp)
                    if "Evaluation Date" in df_opdiv.columns:
                        df_opdiv = df_opdiv.sort_values(["OpDiv", "Evaluation Date"])

                    if "OpDiv" in df_opdiv.columns:
                        for opdiv, g in df_opdiv.groupby("OpDiv", dropna=False):
                            sheet = _safe_sheet_name(str(opdiv) if pd.notna(opdiv) else "UNKNOWN")
                            g.to_excel(writer, sheet_name=sheet, index=False)
                    try:
                        os.remove(opdiv_fp)
                    except Exception:
                        pass
    except Exception as e:
        # Non-fatal: the pipeline is still useful with the CSVs.
        print(f"PERF: Excel export skipped (non-fatal): {e}")


def _run_eval_loop(eval_date, all_alerts: list[str]) -> int:
    """Evaluate all horizons for one eval date. Returns error count."""
    error_count = 0
    for horizon_days in EVAL_HORIZONS:
        try:
            row = _evaluate_horizon(horizon_days=horizon_days, eval_date=eval_date)
            if row is None:
                continue

            wb_path = _horizon_workbook_path(horizon_days)
            overall_hist = pd.DataFrame()
            opdiv_hist = pd.DataFrame()
            if os.path.exists(wb_path):
                try:
                    with pd.ExcelFile(wb_path) as xf:
                        if "overall" in xf.sheet_names:
                            overall_hist = _normalize_perf_history(
                                pd.read_excel(wb_path, sheet_name="overall")
                            )
                    opdiv_frames = []
                    for sn in xf.sheet_names:
                        if sn in {"overall", "Legend"}:
                            continue
                        d = _normalize_perf_history(pd.read_excel(wb_path, sheet_name=sn))
                        if "OpDiv" not in d.columns:
                            d["OpDiv"] = sn
                        opdiv_frames.append(d)
                        if opdiv_frames:
                            opdiv_hist = pd.concat(opdiv_frames, ignore_index=True)
                except Exception as e:
                    log_perf_error(f"could not read existing workbook {wb_path}; starting fresh", e)
                    overall_hist = pd.DataFrame()
                    opdiv_hist = pd.DataFrame()

            perf_df = overall_hist.copy()
            perf_df_new = pd.concat([overall_hist, pd.DataFrame([row])], ignore_index=True)
            save_cols = [c for c in PERF_COLUMNS if c in perf_df_new.columns]
            raw_cols = [c for c in perf_df_new.columns if c.startswith("_raw_")]
            dedupe_keys = [k for k in ["Evaluation Date", "Forecast Date", "Horizon (Days)"] if k in perf_df_new.columns]
            for k in dedupe_keys:
                perf_df_new[k] = perf_df_new[k].astype(str).str.strip()
            if dedupe_keys:
                perf_df_new = perf_df_new.drop_duplicates(subset=dedupe_keys, keep="last")
            if "Evaluation Date" in perf_df_new.columns:
                perf_df_new = perf_df_new.assign(
                    _sort_eval_date=perf_df_new["Evaluation Date"].astype(str).str.strip()
                ).sort_values("_sort_eval_date").drop(columns=["_sort_eval_date"])
            perf_df_new = perf_df_new.reindex(columns=save_cols + raw_cols)

            _print_summary(row)

            alerts = _maybe_alert(perf_df, row)
            if alerts:
                for line in alerts:
                    print(line)
                all_alerts.extend(alerts)

            opdiv_rows = row.get("_opdiv_rows", [])
            for opdiv_row in opdiv_rows:
                opd = str(opdiv_row.get("OpDiv", "")).strip()
                if not opd:
                    continue
                hist = (
                    opdiv_hist[opdiv_hist["OpDiv"].astype(str).str.strip() == opd]
                    if (not opdiv_hist.empty and "OpDiv" in opdiv_hist.columns)
                    else pd.DataFrame()
                )
                opdiv_alerts = _maybe_alert(hist, opdiv_row, opdiv=opd)
                if opdiv_alerts:
                    for line in opdiv_alerts:
                        print(line)
                    all_alerts.extend(opdiv_alerts)
            opdiv_df_new = opdiv_hist.copy()
            if opdiv_rows:
                opdiv_df_new = pd.concat([opdiv_df_new, pd.DataFrame(opdiv_rows)], ignore_index=True)
                save_cols = [c for c in PERF_OPDIV_COLUMNS if c in opdiv_df_new.columns]
                raw_cols = [c for c in opdiv_df_new.columns if c.startswith("_raw_")]
                dedupe_keys = [k for k in ["Evaluation Date", "Forecast Date", "Horizon (Days)", "OpDiv"] if k in opdiv_df_new.columns]
                for k in dedupe_keys:
                    opdiv_df_new[k] = opdiv_df_new[k].astype(str).str.strip()
                if dedupe_keys:
                    opdiv_df_new = opdiv_df_new.drop_duplicates(subset=dedupe_keys, keep="last")
                if {"OpDiv", "Evaluation Date"}.issubset(set(opdiv_df_new.columns)):
                    opdiv_df_new = opdiv_df_new.assign(
                        _sort_eval_date=opdiv_df_new["Evaluation Date"].astype(str).str.strip()
                    ).sort_values(["OpDiv", "_sort_eval_date"]).drop(columns=["_sort_eval_date"])
                opdiv_df_new = opdiv_df_new.reindex(columns=save_cols + raw_cols)

            with pd.ExcelWriter(wb_path, engine="openpyxl") as writer:
                perf_df_new.to_excel(writer, sheet_name="overall", index=False)
                if not opdiv_df_new.empty and "OpDiv" in opdiv_df_new.columns:
                    for opdiv, g in opdiv_df_new.groupby("OpDiv", dropna=False):
                        sheet = _safe_sheet_name(str(opdiv) if pd.notna(opdiv) else "UNKNOWN")
                        g.to_excel(writer, sheet_name=sheet, index=False)
                _format_perf_workbook(writer.book)
        except Exception as e:
            log_perf_error(
                f"{horizon_days}-day horizon failed for eval_date={eval_date.strftime(DATE_FMT)}",
                e,
            )
            error_count += 1
    return error_count


def run_performance_evaluation(eval_date=None) -> bool:
    """
    Evaluate all configured horizons for eval_date (default today).
    Supports backfill via NOI_V4_PERF_BACKFILL_START / END env vars.

    Returns True when no errors occurred; False if any horizon/day failed.
    Never raises — all failures are logged to stdout and Performance/Logs.
    """
    total_errors = 0
    try:
        os.makedirs(PERF_DIR, exist_ok=True)
        os.makedirs(PERF_ALERTS_DIR, exist_ok=True)
        os.makedirs(PERF_LOG_DIR, exist_ok=True)
        _perf_log("starting performance evaluation")

        backfill_start_str = os.environ.get("NOI_V4_PERF_BACKFILL_START")
        backfill_end_str = os.environ.get("NOI_V4_PERF_BACKFILL_END")
        if backfill_start_str:
            backfill_start_str = backfill_start_str.strip()
        if backfill_end_str:
            backfill_end_str = backfill_end_str.strip()

        if backfill_start_str and backfill_end_str:
            start_date = datetime.strptime(backfill_start_str, DATE_FMT).date()
            end_date = datetime.strptime(backfill_end_str, DATE_FMT).date()
            if end_date < start_date:
                raise RuntimeError("NOI_V4_PERF_BACKFILL_END must be >= NOI_V4_PERF_BACKFILL_START")
            cur = start_date
            while cur <= end_date:
                all_alerts: list[str] = []
                try:
                    total_errors += _run_eval_loop(eval_date=cur, all_alerts=all_alerts)
                    if all_alerts:
                        alert_fp = os.path.join(PERF_ALERTS_DIR, f"alert_{cur.strftime(DATE_FMT)}.txt")
                        with open(alert_fp, "w", encoding="utf-8") as f:
                            for line in all_alerts:
                                f.write(line + "\n")
                except Exception as e:
                    log_perf_error(f"backfill day {cur.strftime(DATE_FMT)} failed", e)
                    total_errors += 1
                cur += timedelta(days=1)
        else:
            if eval_date is None:
                # Score the settled date, then walk forward over the dates still
                # inside the settle window. Those newer dates yield a row for
                # every OpDiv that has already reached its usual volume, so a
                # single slow feed no longer blanks the whole day.
                #
                # Self-heal boundary: rows are upserted on (Evaluation Date,
                # Forecast Date, Horizon), so each of these dates is rewritten
                # every morning until it settles, and the provisional numbers
                # are replaced by final ones. Anything older than the window is
                # never revisited and stays as originally written -- including
                # rows produced before this mask existed. Correct those with a
                # backfill via NOI_V4_PERF_BACKFILL_START / _END.
                settled = datetime.today().date() - timedelta(days=OBS_SETTLE_DAYS)
                newest = datetime.today().date() - timedelta(
                    days=max(feed_health.EARLY_SETTLE_MIN_AGE, 1)
                )
                eval_dates = [
                    settled + timedelta(days=k)
                    for k in range((newest - settled).days + 1)
                ] if feed_health.EARLY_SETTLE else [settled]
                _perf_log(
                    f"scoring {eval_dates[0].strftime(DATE_FMT)} (settled) plus "
                    f"{len(eval_dates) - 1} provisional date(s) up to "
                    f"{eval_dates[-1].strftime(DATE_FMT)}"
                )
            else:
                settled = eval_date
                eval_dates = [eval_date]

            for d in eval_dates:
                # Only the settled date alerts. A provisional row is scored
                # against files that may still grow, which can only invent
                # misses -- alerting on it reproduces the false alarm this
                # whole mask exists to prevent.
                day_alerts: list[str] = []
                total_errors += _run_eval_loop(eval_date=d, all_alerts=day_alerts)
                if day_alerts and d == settled:
                    alert_fp = os.path.join(PERF_ALERTS_DIR, f"alert_{d.strftime(DATE_FMT)}.txt")
                    with open(alert_fp, "w", encoding="utf-8") as f:
                        for line in day_alerts:
                            f.write(line + "\n")
                    _perf_log(f"wrote metric alerts to {alert_fp}")
            eval_date = settled

            # Filed under today rather than the eval date, and under a distinct
            # name: this is a statement about the pipeline right now, not about
            # how the model scored three days ago. Skipped during backfill,
            # where forecast availability is a fixed historical fact.
            coverage_alerts = check_forecast_coverage()
            if coverage_alerts:
                today_str = datetime.today().strftime(DATE_FMT)
                cov_fp = os.path.join(PERF_ALERTS_DIR, f"alert_pipeline_{today_str}.txt")
                with open(cov_fp, "w", encoding="utf-8") as f:
                    for line in coverage_alerts:
                        f.write(line + "\n")
                _perf_log(f"wrote pipeline alerts to {cov_fp}", level="WARNING")

        if total_errors:
            _perf_log(f"evaluation finished with {total_errors} error(s)", level="WARNING")
            return False
        _perf_log("evaluation finished successfully")
        return True
    except Exception as e:
        log_perf_error("performance evaluation aborted", e)
        return False


def run_eval_after_forecast(
    stamp: str,
    save_dir: str,
    htoc_share_root: str | None = None,
    consolidate_only: bool = False,
) -> bool:
    """
    Consolidate today's forecasts and run evaluation.
    Non-fatal wrapper for the forecast scheduled task; logs all issues.

    `consolidate_only` is for as-of replays, which need the daily report built
    but should not each trigger a scoring pass over the same settled date; the
    caller scores the affected range once at the end.
    """
    try:
        init_performance_eval(save_dir, htoc_share_root)
        _perf_log(f"starting post-forecast evaluation (stamp={stamp})")
        out = consolidate_daily_report(stamp)
        if out is None:
            _perf_log(f"consolidation produced no report for stamp={stamp}", level="WARNING")
        else:
            _perf_log(f"consolidated daily report: {out}")
        if consolidate_only:
            _perf_log(f"consolidate-only mode (stamp={stamp}); skipping scoring pass")
            return out is not None
        return run_performance_evaluation()
    except Exception as e:
        log_perf_error("post-forecast evaluation failed (non-fatal)", e)
        return False
