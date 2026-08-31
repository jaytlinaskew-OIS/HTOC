"""Merge scored rows into the PRISM Excel workbook (test path by default)."""
from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from htoc_ml.prism.engine import COLUMN_RENAME, EXPORT_COLUMNS

FILLS = {
    "low": PatternFill(start_color="83de85", end_color="83de85", fill_type="solid"),
    "medium": PatternFill(start_color="eef084", end_color="eef084", fill_type="solid"),
    "high": PatternFill(start_color="f29953", end_color="f29953", fill_type="solid"),
    "critical": PatternFill(start_color="e83f3f", end_color="e83f3f", fill_type="solid"),
}


def export_frame(df_scored: pd.DataFrame) -> pd.DataFrame:
    columns = [c for c in EXPORT_COLUMNS if c in df_scored.columns]
    frame = df_scored[columns].copy()
    if "PRISM Score (Final)" in df_scored.columns:
        frame["PRISM Score"] = df_scored["PRISM Score (Final)"]
    if "Severity (Final)" in df_scored.columns:
        frame["Severity"] = df_scored["Severity (Final)"]
    if "VirusTotal Malicious Score" not in frame.columns and "VT Display" in df_scored.columns:
        frame["VirusTotal Malicious Score"] = df_scored["VT Display"]
    for col in frame.columns:
        if pd.api.types.is_datetime64_any_dtype(frame[col]) and getattr(frame[col].dt, "tz", None) is not None:
            frame[col] = frame[col].dt.tz_convert("UTC").dt.tz_localize(None)
    return frame


def merge_into_workbook(df_scored: pd.DataFrame, excel_path: Path) -> Path:
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    df_export = export_frame(df_scored)
    columns_to_save = list(df_export.columns)
    if excel_path.exists():
        existing = pd.read_excel(excel_path, engine="openpyxl")
        rename = {
            old: new
            for old, new in COLUMN_RENAME.items()
            if old in existing.columns and new not in existing.columns
        }
        if rename:
            existing.rename(columns=rename, inplace=True)
        for col in columns_to_save:
            if col not in existing.columns:
                existing[col] = pd.NaT if col == "Last Observed" else (
                    0 if col in {
                        "VirusTotal Malicious Score", "Observation Yearly Count",
                        "ThreatConnect Rating", "Observation Penalty Multiplier",
                        "Botnet Flag", "False Positives", "PRISM Score", "ThreatConnect Score",
                    } else ""
                )
        existing = existing[[c for c in columns_to_save if c in existing.columns]].copy()
        existing_set = set(existing["Indicator"].values)
        new_set = set(df_export["Indicator"].values)
        existing_idx = existing.set_index("Indicator").sort_index()
        export_idx = df_export.set_index("Indicator").sort_index()
        to_update = [
            i for i in existing_set & new_set
            if not existing_idx.loc[i].equals(export_idx.loc[i])
        ]
        unchanged = [i for i in existing_set & new_set if i not in to_update]
        combined = pd.concat([
            existing[existing["Indicator"].isin(unchanged)],
            df_export[df_export["Indicator"].isin(to_update)],
            df_export[df_export["Indicator"].isin(new_set - existing_set)],
            existing[~existing["Indicator"].isin(new_set)],
        ], ignore_index=True).drop_duplicates(subset="Indicator", keep="last")
        print(f"Updated: {len(to_update)} | Added: {len(new_set - existing_set)} | Total: {len(combined)}")
    else:
        combined = df_export.drop_duplicates(subset="Indicator", keep="last").copy()
        print(f"Created new file with {len(combined)} indicators")

    history = None
    if excel_path.exists():
        try:
            history = pd.read_excel(excel_path, sheet_name="Complete History", engine="openpyxl")
        except (ValueError, KeyError, OSError):
            history = None

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        combined.to_excel(writer, index=False, sheet_name="PRISM Scores")
        if "ThreatConnect Score" in combined.columns and "PRISM Score" in combined.columns:
            compare = combined[["Indicator", "ThreatConnect Score", "PRISM Score"]].copy()
            compare["ThreatConnect Score"] = pd.to_numeric(compare["ThreatConnect Score"], errors="coerce").fillna(0)
            compare["PRISM Score"] = pd.to_numeric(compare["PRISM Score"], errors="coerce").fillna(0)
            compare["Difference"] = compare["PRISM Score"] - compare["ThreatConnect Score"]
            compare.to_excel(writer, index=False, sheet_name="Score Comparison")
        if history is not None:
            history.to_excel(writer, index=False, sheet_name="Complete History")
        worksheet = writer.sheets["PRISM Scores"]
        for row_idx, severity in enumerate(combined["Severity"], start=2):
            fill = FILLS.get(str(severity).lower())
            if fill:
                for col_idx in range(1, len(combined.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = fill

    _append_history(df_scored, excel_path)
    print(f"Saved {len(combined)} indicators to {excel_path}")
    return excel_path


def _append_history(df_scored: pd.DataFrame, excel_path: Path) -> None:
    run_timestamp = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S")
    history_columns = ["Scoring Date", "Indicator", "Indicator Type", "PRISM Score", "Severity", "Explanation"]
    mapping = {
        "Indicator": "Indicator",
        "Indicator Type": "Indicator Type",
        "PRISM Score (Final)": "PRISM Score",
        "Severity (Final)": "Severity",
        "Explanation": "Explanation",
    }
    available = {k: v for k, v in mapping.items() if k in df_scored.columns}
    slice_df = df_scored[list(available.keys())].rename(columns=available)
    for col in slice_df.select_dtypes(["category"]).columns:
        slice_df[col] = slice_df[col].astype(str)
    current = slice_df.fillna("").assign(**{"Scoring Date": run_timestamp})
    current = current[[c for c in history_columns if c in current.columns]]

    if excel_path.exists():
        try:
            history_all = pd.read_excel(excel_path, sheet_name="Complete History", engine="openpyxl")
            for col in history_all.columns:
                if pd.api.types.is_datetime64_any_dtype(history_all[col]) and getattr(history_all[col].dtype, "tz", None) is not None:
                    history_all[col] = history_all[col].dt.tz_convert("UTC").dt.tz_localize(None)
            history_all = pd.concat([
                history_all.dropna(axis=1, how="all"),
                current.dropna(axis=1, how="all"),
            ], ignore_index=True)
        except (ValueError, KeyError, OSError):
            history_all = current.copy()
    else:
        history_all = current.copy()

    history_all["_date_only"] = pd.to_datetime(history_all["Scoring Date"]).dt.date
    history_all = history_all.drop_duplicates(subset=["Indicator", "_date_only"], keep="last").drop(columns=["_date_only"])
    history_all = history_all.sort_values(["Indicator", "Scoring Date"], ascending=[True, False])
    history_all = history_all[[c for c in history_columns if c in history_all.columns]]

    existing_sheets = {}
    if excel_path.exists():
        try:
            book = pd.ExcelFile(excel_path, engine="openpyxl")
            for name in book.sheet_names:
                if name not in {"Complete History", "Latest Scores"}:
                    existing_sheets[name] = pd.read_excel(excel_path, sheet_name=name, engine="openpyxl")
            book.close()
        except (ValueError, OSError) as exc:
            print(f"Warning: Could not read existing sheets: {exc}")

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for name, sheet in existing_sheets.items():
            sheet.to_excel(writer, index=False, sheet_name=name)
            if name == "PRISM Scores" and "Severity" in sheet.columns:
                ws = writer.sheets["PRISM Scores"]
                for row_idx, severity in enumerate(sheet["Severity"], start=2):
                    fill = FILLS.get(str(severity).lower())
                    if fill:
                        for col_idx in range(1, len(sheet.columns) + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
        history_all.to_excel(writer, sheet_name="Complete History", index=False)
        ws_h = writer.sheets["Complete History"]
        for dim, width in [("A", 20), ("B", 20), ("C", 15), ("D", 18), ("E", 12), ("F", 60)]:
            ws_h.column_dimensions[dim].width = width
        for cell in ws_h[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    print(
        f"Scoring history updated — {len(history_all)} total records, "
        f"{history_all['Indicator'].nunique()} unique indicators"
    )
