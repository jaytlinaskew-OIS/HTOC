"""Growing Excel workbooks with traffic-light metric columns and a legend."""
from __future__ import annotations

from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FILL_GREEN = PatternFill(fill_type="solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill(fill_type="solid", fgColor="FFEB9C")
FILL_RED = PatternFill(fill_type="solid", fgColor="FFC7CE")
FONT_GREEN = Font(color="006100")
FONT_YELLOW = Font(color="9C5700")
FONT_RED = Font(color="9C0006")
FILL_HEADER = PatternFill(fill_type="solid", fgColor="D9E2F3")
FONT_HEADER = Font(bold=True)


@dataclass(frozen=True)
class LegendSpec:
    title: str
    intro: str
    traffic_lights: Mapping[str, tuple[float, float]]
    descriptions: Mapping[str, tuple[str, str]]
    other_metrics: tuple[str, ...]
    footnotes: tuple[str, ...]


def traffic_light_for_value(value, green_min: float, yellow_min: float):
    try:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    if numeric >= green_min:
        return FILL_GREEN, FONT_GREEN
    if numeric >= yellow_min:
        return FILL_YELLOW, FONT_YELLOW
    return FILL_RED, FONT_RED


def safe_sheet_name(name: str) -> str:
    cleaned = str(name)
    for char in "[]:*?/\\":
        cleaned = cleaned.replace(char, "_")
    return cleaned[:31]


def upsert_history(
    existing: pd.DataFrame,
    new_rows: pd.DataFrame | list[dict],
    *,
    keys: Sequence[str],
    columns: Sequence[str],
    sort_by: Sequence[str],
) -> pd.DataFrame:
    incoming = pd.DataFrame(new_rows)
    if existing is None or existing.empty:
        combined = incoming.copy()
    else:
        combined = pd.concat([existing, incoming], ignore_index=True)
    if combined.empty:
        return combined
    present_keys = [key for key in keys if key in combined.columns]
    for key in present_keys:
        combined[key] = combined[key].astype(str).str.strip()
    if present_keys:
        combined = combined.drop_duplicates(subset=list(present_keys), keep="last")
    present_sort = [key for key in sort_by if key in combined.columns]
    if present_sort:
        combined = combined.assign(
            **{f"_sort_{key}": combined[key].astype(str).str.strip() for key in present_sort}
        ).sort_values([f"_sort_{key}" for key in present_sort])
        combined = combined.drop(columns=[f"_sort_{key}" for key in present_sort])
    save_cols = [col for col in columns if col in combined.columns]
    raw_cols = [col for col in combined.columns if str(col).startswith("_raw_")]
    return combined.reindex(columns=save_cols + raw_cols)


def apply_sheet_formatting(ws, traffic_lights: Mapping[str, tuple[float, float]]) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        name = str(cell.value) if cell.value is not None else ""
        headers[name] = col
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 36
    for col_name, (green_min, yellow_min) in traffic_lights.items():
        col_idx = headers.get(col_name)
        if not col_idx:
            continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, col_idx)
            style = traffic_light_for_value(cell.value, green_min, yellow_min)
            if style is None:
                continue
            fill, font = style
            cell.fill = fill
            cell.font = font
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col).value or "")
        ws.column_dimensions[get_column_letter(col)].width = min(max(len(header) + 2, 12), 42)


def write_legend_sheet(workbook, spec: LegendSpec) -> None:
    if "Legend" in workbook.sheetnames:
        del workbook["Legend"]
    ws = workbook.create_sheet("Legend", 0)
    ws["A1"] = spec.title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = spec.intro
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 96

    ws["A4"] = "Color key"
    ws["A4"].font = Font(bold=True, size=12)
    ws["A5"] = "Color"
    ws["B5"] = "Meaning"
    ws["C5"] = "Rule"
    for col in range(1, 4):
        ws.cell(5, col).fill = FILL_HEADER
        ws.cell(5, col).font = FONT_HEADER
    ws["A6"] = "Green"
    ws["A6"].fill = FILL_GREEN
    ws["A6"].font = FONT_GREEN
    ws["B6"] = "Good"
    ws["C6"] = "At or above the green threshold"
    ws["A7"] = "Yellow"
    ws["A7"].fill = FILL_YELLOW
    ws["A7"].font = FONT_YELLOW
    ws["B7"] = "Watch / middle"
    ws["C7"] = "At or above yellow, below green"
    ws["A8"] = "Red"
    ws["A8"].fill = FILL_RED
    ws["A8"].font = FONT_RED
    ws["B8"] = "Bad"
    ws["C8"] = "Below the yellow threshold"

    ws["A10"] = "Traffic-light metrics"
    ws["A10"].font = Font(bold=True, size=12)
    for col, title in enumerate(
        ["Metric", "What it measures", "Why it matters", "Green >=", "Yellow >=", "Red <"],
        start=1,
    ):
        cell = ws.cell(11, col, title)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    row = 12
    for metric, (green_min, yellow_min) in spec.traffic_lights.items():
        what, why = spec.descriptions.get(metric, ("", ""))
        ws.cell(row, 1, metric)
        ws.cell(row, 2, what).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 3, why).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 4, green_min)
        ws.cell(row, 5, yellow_min)
        ws.cell(row, 6, yellow_min)
        ws.row_dimensions[row].height = 60
        row += 1

    row += 1
    ws.cell(row, 1, "Other key measures (not traffic-lighted)")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1
    for col, title in enumerate(["Metric", "What it measures", "Why it matters"], start=1):
        cell = ws.cell(row, col, title)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    row += 1
    for metric in spec.other_metrics:
        what, why = spec.descriptions.get(metric, ("", ""))
        ws.cell(row, 1, metric)
        ws.cell(row, 2, what).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 3, why).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 48
        row += 1

    row += 1
    ws.cell(row, 1, "How rows get written and corrected")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1
    for note in spec.footnotes:
        ws.cell(row, 1, note).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 46
        row += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.freeze_panes = "A12"


def write_grouped_workbook(
    path: str | Path,
    overall: pd.DataFrame,
    grouped: pd.DataFrame,
    *,
    group_col: str,
    legend: LegendSpec,
) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        overall.to_excel(writer, sheet_name="overall", index=False)
        if not grouped.empty and group_col in grouped.columns:
            for name, group in grouped.groupby(group_col, dropna=False):
                sheet = safe_sheet_name(str(name) if pd.notna(name) else "UNKNOWN")
                group.to_excel(writer, sheet_name=sheet, index=False)
        write_legend_sheet(writer.book, legend)
        for sheet_name in writer.book.sheetnames:
            if sheet_name == "Legend":
                continue
            apply_sheet_formatting(writer.book[sheet_name], legend.traffic_lights)
